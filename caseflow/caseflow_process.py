from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import subprocess
import zipfile
from collections import defaultdict
from copy import copy
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path, PurePosixPath
from typing import Any, TypedDict

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover
    raise SystemExit("VARTA потребує openpyxl. Запустіть інсталер ще раз або встановіть openpyxl.") from exc

PdfReader: Any = None
try:
    from pypdf import PdfReader as _PdfReader
except ImportError:  # pragma: no cover
    pass
else:
    PdfReader = _PdfReader


class Candidate(TypedDict):
    path: Path
    scan_root: str
    relative: Path
    proceeding_folder: str
    flow_folder: str
    flow: str
    channel: str
    component: str
    sha256: str
    uploaded_at: str | None


class LogicalGroup(TypedDict):
    main: Candidate
    files: list[Candidate]


DOCUMENT_EXTENSIONS = {".pdf", ".html", ".htm", ".docx", ".doc", ".rtf", ".txt"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff"}
SIGNATURE_EXTENSIONS = {".p7s", ".p7s.2"}
PROCEEDING_RE = re.compile(r"(?<![\w/])\d{1,3}(?:-[А-Яа-яІіЇїЄєҐґ]+)?/\d{3}/\d+/\d{2,4}(?![\w/])")
DATE_RE = re.compile(r"(?<!\d)(20\d{2})[-_. ](0?[1-9]|1[0-2])[-_. ](0?[1-9]|[12]\d|3[01])(?!\d)")
MAX_ZIP_DEPTH = 5
MAX_ZIP_ENTRIES = 10000

SHEET_HEADERS = {
    "Партії": ["ID партії", "Провадження", "Джерело", "№ партії", "Дата завантаження", "Очікувано документів", "Фактично документів", "Статус", "Дата опрацювання", "Примітки"],
    "Документи": ["ID документа", "ID ЄСІТС", "№ документа", "Дата документа", "Дата надходження/подання", "Провадження", "Суд / джерело", "Потік", "Назва документа", "Тип документа", "Опис / пов’язана подія", "ID партії", "Основний файл", "Додатків очікується", "Додатків фактично", "Картка руху", "Протокол КЕП", "Підпис", "Статус комплектності", "Наступна дія", "Відносна папка", "Google Drive URL", "Airtable record ID", "Примітки", "Класифікація", "Статус звірки"],
    "Хронологія": ["ID події", "Дата / час", "Провадження", "Напрям / джерело", "Учасник / орган", "Тип події", "ID документа", "Документ / подія", "Статус / результат", "Наступна дія", "Строк", "ID партії", "Посилання / шлях", "Примітки"],
    "Файли": ["ID файла", "ID документа", "Компонент", "№ компонента", "Оригінальна назва", "Нормалізована назва", "Відносний шлях", "Повний локальний шлях", "Розширення", "MIME", "Розмір, байт", "SHA-256", "Дублікат?", "Цілісність", "Дата завантаження", "Дата опрацювання", "Google Drive URL", "Airtable record ID", "ID партії", "Провадження", "Примітки"],
    "Провадження": ["Папка", "Номер провадження", "Тип", "Суд / орган", "Статус", "Поточний потік", "Документів від суду", "Моїх документів", "Всього документів", "Комплектних", "Потребують перевірки", "Остання дата документа", "Примітки"],
    "Довідники": ["Провадження", "Потік", "Компонент", "Тип документа", "Статус партії", "Основний файл", "Картка руху", "Протокол КЕП", "Підпис", "Цілісність", "Дублікат", "Статус провадження", "Комплектність", "Правило назви", "judgment_code", "name", "Класифікація", "Статус звірки", "Канал надходження", "Пояснення каналу"],
}


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def read_json(path: Path, default):
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return default


def write_json(path: Path, value) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(path)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def compound_suffix(path: Path) -> str:
    return ".p7s.2" if path.name.lower().endswith(".p7s.2") else path.suffix.lower()


def safe_segment(value: str, fallback: str = "БЕЗ_НАЗВИ") -> str:
    value = " ".join(str(value or "").strip().split())
    for char in '<>:"/\\|?*':
        value = value.replace(char, "_")
    value = re.sub(r"_+", "_", value).strip(" ._")[:120]
    return value or fallback


def normal_token(value: str) -> str:
    value = safe_segment(value).upper().replace(" ", "_")
    return re.sub(r"_+", "_", value).strip("_")


def folder_to_proceeding(folder: str) -> str:
    match = re.fullmatch(r"(\d{1,3})_(\d{3})_(\d+)_(\d{2,4})", folder)
    return "/".join(match.groups()) if match else folder.replace("_Ц_", "-ц/").replace("_", "/")


def stream_from_path(relative: Path) -> tuple[str, str]:
    for part in relative.parts:
        folded = part.casefold()
        if folded == "01_від_суду".casefold():
            return "01_ВІД_СУДУ", "Від суду"
        if folded == "02_мої_документи".casefold():
            return "02_МОЇ_ДОКУМЕНТИ", "Мої документи"
    return "00_НЕВИЗНАЧЕНО", "Невизначено"


def component(path: Path) -> str:
    name = path.name.casefold()
    suffix = compound_suffix(path)
    if suffix in SIGNATURE_EXTENSIONS:
        return "ПІДПИС"
    if "протокол" in name and ("кеп" in name or "підпис" in name):
        return "ПРОТОКОЛ_КЕП"
    if "картк" in name and ("рух" in name or "реєстрац" in name or "документ" in name):
        return "КАРТКА_РУХУ"
    if suffix in IMAGE_EXTENSIONS:
        return "СКРІНШОТ"
    if suffix in DOCUMENT_EXTENSIONS:
        return "ОСНОВНИЙ"
    return "ІНШЕ"


def mime_for(path: Path) -> str:
    import mimetypes

    return mimetypes.guess_type(path.name)[0] or "application/octet-stream"


def unsafe_zip_name(name: str) -> bool:
    pure = PurePosixPath(name.replace("\\", "/"))
    return pure.is_absolute() or ".." in pure.parts


def decoded_zip_name(info: zipfile.ZipInfo) -> str:
    """Recover Windows-1251 names from ESUD ZIP files without the UTF-8 flag."""
    name = info.filename
    if info.flag_bits & 0x800:
        return name
    try:
        candidate = name.encode("cp437").decode("windows-1251")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return name
    def cyrillic(value: str) -> int:
        return sum("А" <= char <= "я" or char in "ІіЇїЄєҐґ" for char in value)

    return candidate if cyrillic(candidate) > cyrillic(name) else name


def find_7zip(root: Path) -> Path | None:
    config = read_json(root / ".caseflow" / "config.json", {})
    candidates = [
        str(config.get("seven_zip_path", "")),
        os.environ.get("VARTA_7Z", ""),
        os.environ.get("CASEFLOW_7Z", ""),  # legacy compatibility
    ]
    for variable in ("ProgramFiles", "ProgramFiles(x86)"):
        base = os.environ.get(variable)
        if base:
            candidates.append(str(Path(base) / "7-Zip" / "7z.exe"))
    candidates.extend(
        [
            r"C:\Program Files\3uToolsV3\files\patchtools\7z-64\7z.exe",
            r"C:\Program Files\3uToolsV3\files\patchtools\7z-32\7z.exe",
            r"C:\Program Files\Lenovo\Lenovo Bootable Generator\7z.exe",
        ]
    )
    command = shutil.which("7z") or shutil.which("7z.exe")
    if command:
        candidates.append(command)
    for candidate in candidates:
        if candidate and Path(candidate).is_file():
            return Path(candidate).resolve()
    return None


def seven_zip_members(source: Path, executable: Path) -> list[dict]:
    result = subprocess.run(
        [str(executable), "l", "-slt", "-ba", "-sccUTF-8", str(source)],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=120,
        check=False,
    )
    if result.returncode not in {0, 1}:
        raise RuntimeError((result.stderr or result.stdout or "7-Zip не прочитав архів")[-2000:])
    members: list[dict] = []
    record: dict[str, str] = {}
    for line in result.stdout.splitlines() + [""]:
        if not line.strip():
            if record and record.get("Path") and record.get("Folder") != "+":
                members.append(record)
            record = {}
            continue
        key, separator, value = line.partition(" = ")
        if separator:
            record[key.strip()] = value.strip()
    return members


def extract_rar_safe(
    source: Path,
    destination: Path,
    executable: Path,
    depth: int = 0,
    counter: list[int] | None = None,
) -> list[dict]:
    """Validate all RAR paths, extract to staging, then recurse into nested archives."""
    if counter is None:
        counter = [0]
    if depth > MAX_ZIP_DEPTH:
        return [{"path": str(source), "status": "skipped", "reason": "archive_depth_limit"}]
    try:
        members = seven_zip_members(source, executable)
    except (OSError, RuntimeError, subprocess.SubprocessError) as exc:
        return [{"path": str(source), "status": "error", "reason": str(exc)}]
    unsafe = [item.get("Path", "") for item in members if unsafe_zip_name(item.get("Path", ""))]
    if unsafe:
        return [{"path": str(source), "status": "skipped", "reason": "unsafe_path", "entries": unsafe[:20]}]
    counter[0] += len(members)
    if counter[0] > MAX_ZIP_ENTRIES:
        return [{"path": str(source), "status": "skipped", "reason": "archive_entry_limit"}]
    destination.mkdir(parents=True, exist_ok=True)
    try:
        result = subprocess.run(
            [str(executable), "x", "-y", "-aoa", "-sccUTF-8", f"-o{destination}", str(source)],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=30 * 60,
            check=False,
        )
    except (OSError, subprocess.SubprocessError) as exc:
        return [{"path": str(source), "status": "error", "reason": str(exc)}]
    if result.returncode not in {0, 1}:
        return [{"path": str(source), "status": "error", "reason": (result.stderr or result.stdout)[-2000:]}]
    results: list[dict] = []
    destination_resolved = destination.resolve()
    for target in sorted(destination.rglob("*"), key=lambda item: str(item).casefold()):
        if target.is_symlink():
            target.unlink(missing_ok=True)
            results.append({"entry": str(target.relative_to(destination)), "status": "skipped", "reason": "symbolic_link"})
            continue
        if not target.is_file():
            continue
        resolved = target.resolve()
        if destination_resolved not in resolved.parents:
            results.append({"entry": str(target), "status": "skipped", "reason": "outside_destination"})
            continue
        results.append({"entry": str(target.relative_to(destination)), "path": str(target), "status": "extracted"})
        if target.suffix.casefold() == ".zip":
            results.extend(extract_zip_safe(target, target.parent / f"{target.stem}__nested", depth + 1, counter))
        elif target.suffix.casefold() == ".rar":
            results.extend(extract_rar_safe(target, target.parent / f"{target.stem}__nested", executable, depth + 1, counter))
    return results


def extract_zip_safe(source: Path, destination: Path, depth: int = 0, counter: list[int] | None = None) -> list[dict]:
    if counter is None:
        counter = [0]
    results = []
    if depth > MAX_ZIP_DEPTH:
        return [{"path": str(source), "status": "skipped", "reason": "zip_depth_limit"}]
    destination.mkdir(parents=True, exist_ok=True)
    try:
        archive = zipfile.ZipFile(source)
    except (zipfile.BadZipFile, OSError) as exc:
        return [{"path": str(source), "status": "error", "reason": str(exc)}]
    with archive:
        for info in archive.infolist():
            if info.is_dir():
                continue
            counter[0] += 1
            if counter[0] > MAX_ZIP_ENTRIES:
                results.append({"path": str(source), "status": "skipped", "reason": "zip_entry_limit"})
                break
            entry_name = decoded_zip_name(info)
            if unsafe_zip_name(entry_name):
                results.append({"entry": entry_name, "status": "skipped", "reason": "unsafe_path"})
                continue
            target = (destination / Path(*PurePosixPath(entry_name.replace("\\", "/")).parts)).resolve()
            if destination.resolve() not in target.parents:
                results.append({"entry": entry_name, "status": "skipped", "reason": "outside_destination"})
                continue
            target.parent.mkdir(parents=True, exist_ok=True)
            with archive.open(info) as input_stream, target.open("wb") as output_stream:
                shutil.copyfileobj(input_stream, output_stream, length=1024 * 1024)
            results.append({"entry": entry_name, "path": str(target), "status": "extracted"})
            if target.suffix.lower() == ".zip":
                nested_destination = target.parent / f"{target.stem}__nested"
                results.extend(extract_zip_safe(target, nested_destination, depth + 1, counter))
    return results


def text_from_file(path: Path, max_chars: int = 50000) -> str:
    suffix = compound_suffix(path)
    try:
        if suffix == ".pdf" and PdfReader is not None:
            reader = PdfReader(path, strict=False)
            return " ".join((page.extract_text() or "") for page in reader.pages[:12])[:max_chars]
        if suffix in {".html", ".htm", ".txt", ".rtf"}:
            raw = path.read_bytes()
            for encoding in ("utf-8-sig", "utf-8", "windows-1251"):
                try:
                    text = raw.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                text = raw.decode("utf-8", errors="replace")
            if suffix in {".html", ".htm"}:
                text = re.sub(r"(?is)<script.*?</script>|<style.*?</style>", " ", text)
                text = re.sub(r"(?s)<[^>]+>", " ", text)
            return text[:max_chars]
        if suffix == ".docx":
            with zipfile.ZipFile(path) as archive:
                xml = archive.read("word/document.xml").decode("utf-8", errors="replace")
            return re.sub(r"(?s)<[^>]+>", " ", xml)[:max_chars]
    except Exception:
        return ""
    return ""


def parse_document_date(path: Path, text: str) -> datetime | None:
    match = DATE_RE.search(path.stem) or DATE_RE.search(text[:5000])
    if not match:
        return None
    try:
        return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def parse_submission_time(text: str) -> datetime | None:
    normalized = " ".join(text.split())
    patterns = [
        r"Надіслано до суду.{0,180}?(\d{2})[.](\d{2})[.](20\d{2}).{0,40}?(\d{2})[:](\d{2})",
        r"(\d{2})[.](\d{2})[.](20\d{2}).{0,40}?(\d{2})[:](\d{2}).{0,180}?Надіслано до суду",
        r"Надіслано до суду.{0,180}?(20\d{2})[-](\d{2})[-](\d{2}).{0,40}?(\d{2})[:](\d{2})",
    ]
    for index, pattern in enumerate(patterns):
        match = re.search(pattern, normalized, flags=re.IGNORECASE)
        if not match:
            continue
        values = [int(value) for value in match.groups()]
        if index < 2:
            day, month, year, hour, minute = values
        else:
            year, month, day, hour, minute = values
        try:
            return datetime(year, month, day, hour, minute)
        except ValueError:
            continue
    return None


def infer_type(path: Path) -> str:
    name = path.stem.casefold()
    rules = [
        ("касаційн", "Касаційна скарга"), ("апеляційн", "Апеляційна скарга"),
        ("позов", "Позовна заява"), ("клопотан", "Клопотання"), ("запереч", "Заперечення"),
        ("відзив", "Відзив"), ("скарг", "Скарга"), ("заяв", "Заява"),
        ("ухвал", "Ухвала"), ("рішенн", "Рішення"), ("постанов", "Постанова"),
        ("вирок", "Вирок"), ("судовий наказ", "Судовий наказ"),
    ]
    return next((kind for token, kind in rules if token in name), "Інше")


def short_title(path: Path, doc_type: str) -> str:
    stem = re.sub(r"[0-9a-f]{8,}", " ", path.stem, flags=re.IGNORECASE)
    stem = re.sub(r"20\d{2}[-_. ]\d{1,2}[-_. ]\d{1,2}", " ", stem)
    words = [word for word in re.split(r"[_\s.,;()\[\]{}]+", stem) if len(word) > 1 and not word.isdigit()]
    if not words:
        words = doc_type.split()
    return " ".join(words[:8])[:160]


def next_id(values, prefix: str, width: int = 4) -> str:
    maximum = 0
    pattern = re.compile(rf"^{re.escape(prefix)}_(\d+)$", re.IGNORECASE)
    for value in values:
        match = pattern.match(str(value or "").strip())
        if match:
            maximum = max(maximum, int(match.group(1)))
    return f"{prefix}_{maximum + 1:0{width}d}"


def increment_id(current: str) -> str:
    prefix, number = current.rsplit("_", 1)
    return f"{prefix}_{int(number) + 1:0{len(number)}d}"


def ensure_workbook(template: Path | None):
    if template:
        return load_workbook(template)
    workbook = Workbook()
    if workbook.worksheets:
        workbook.remove(workbook.worksheets[0])
    control = workbook.create_sheet("Контроль")
    control.append(["VARTA — РЕЄСТР"])
    for table_index, (name, headers) in enumerate(SHEET_HEADERS.items(), 1):
        sheet = workbook.create_sheet(name)
        sheet.append(headers)
        table = Table(displayName=f"VartaTable{table_index}", ref=f"A1:{sheet.cell(1, len(headers)).column_letter}1")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        sheet.add_table(table)
        sheet.freeze_panes = "A2"
    return workbook


def find_template(root: Path) -> Path | None:
    pointer = root / "03_РЕЄСТР" / "ОСТАННІЙ_РЕЄСТР.txt"
    if pointer.exists():
        try:
            pointed = Path(pointer.read_text(encoding="utf-8-sig").strip())
            if not pointed.is_absolute():
                pointed = root / pointed
            if pointed.exists() and pointed.suffix.lower() == ".xlsx" and not pointed.name.startswith("~$"):
                return pointed.resolve()
        except OSError:
            pass
    candidates = [path for path in (root / "03_РЕЄСТР" / "exports").glob("*.xlsx") if not path.name.startswith("~$")]
    candidates += [path for path in (root / "03_РЕЄСТР").glob("*.xlsx") if not path.name.startswith("~$")]
    if not candidates:
        candidates = list((root / "outputs").glob("**/*легенда*часи*.xlsx"))
    return max(candidates, key=lambda path: path.stat().st_mtime) if candidates else None


def append_styled(sheet, row: list):
    new_row = sheet.max_row + 1
    sheet.append(row)
    source_row = 2 if sheet.max_row > 2 else 1
    if source_row != new_row:
        for col in range(1, min(len(row), sheet.max_column) + 1):
            source = sheet.cell(source_row, col)
            target = sheet.cell(new_row, col)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            if source.alignment:
                target.alignment = copy(source.alignment)
    return new_row


def expand_tables(sheet) -> None:
    for table in sheet.tables.values():
        start, _ = table.ref.split(":")
        table.ref = f"{start}:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}"


def sort_timeline(sheet) -> None:
    """Keep chronology in factual event-time order without changing row styles."""
    headers = {str(sheet.cell(1, column).value or "").strip(): column for column in range(1, sheet.max_column + 1)}
    time_column = headers.get("Дата / час")
    id_column = headers.get("ID події")
    if not time_column or sheet.max_row < 3:
        return

    def key(values: list):
        raw = values[time_column - 1]
        if isinstance(raw, datetime):
            moment = raw
        else:
            moment = None
            text = str(raw or "").strip()
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%Y-%m-%d", "%d.%m.%Y"):
                try:
                    moment = datetime.strptime(text, fmt)
                    break
                except ValueError:
                    continue
        precise = bool(re.search(r"\d{1,2}:\d{2}", str(raw))) or (isinstance(raw, datetime) and raw.time() != datetime.min.time())
        return (
            1 if moment is None else 0,
            moment.date() if moment else datetime.max.date(),
            0 if precise else 1,
            moment.time() if moment and precise else datetime.max.time(),
            str(values[(id_column or 1) - 1] or ""),
        )

    values = [[sheet.cell(row, column).value for column in range(1, sheet.max_column + 1)] for row in range(2, sheet.max_row + 1)]
    values.sort(key=key)
    for row_offset, row_values in enumerate(values, 2):
        for column, value in enumerate(row_values, 1):
            sheet.cell(row_offset, column).value = value


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Універсальний локальний конвеєр VARTA")
    parser.add_argument("--root", type=Path, required=True)
    parser.add_argument("--settings-json", default="{}")
    args = parser.parse_args(argv)
    root = args.root.resolve()
    settings = {
        "includeOriginals": True,
        "extractArchives": True,
        "includeCards": True,
        "includeSignatures": True,
        "includeScreenshots": True,
        "autoRegister": True,
        **json.loads(args.settings_json or "{}"),
    }
    started = datetime.now()
    stamp = started.strftime("%Y%m%d_%H%M%S_%f")
    inbox = root / "00_INBOX"
    unpacked_root = root / "02_РОЗПАКОВАНО"
    processed_root = root / "01_ОПРАЦЬОВАНО"
    review_root = root / "99_ПОТРЕБУЄ_ПЕРЕВІРКИ"
    for folder in [inbox, unpacked_root, processed_root, review_root, root / "03_РЕЄСТР" / "exports", root / "tmp" / "caseflow_runs", root / "tmp" / "summary_agent"]:
        folder.mkdir(parents=True, exist_ok=True)

    extraction_log: list[dict[str, Any]] = []
    if settings["extractArchives"]:
        for archive in sorted(inbox.rglob("*.zip"), key=lambda p: str(p).casefold()):
            relative = archive.relative_to(inbox)
            digest = sha256(archive)[:12]
            destination = unpacked_root / relative.parent / f"{safe_segment(archive.stem)}__{digest}"
            if not destination.exists():
                extraction_log.extend(extract_zip_safe(archive, destination))
        rar_executable = find_7zip(root)
        for archive in sorted(inbox.rglob("*.rar"), key=lambda p: str(p).casefold()):
            relative = archive.relative_to(inbox)
            digest = sha256(archive)[:12]
            destination = unpacked_root / relative.parent / f"{safe_segment(archive.stem)}__{digest}"
            if not destination.exists():
                if rar_executable:
                    extraction_log.extend(extract_rar_safe(archive, destination, rar_executable))
                else:
                    extraction_log.append({"path": str(archive), "status": "error", "reason": "7zip_not_found"})

    manifests: dict[Path, dict[str, Any]] = {}
    for manifest in inbox.rglob("caseflow_upload.json"):
        payload = read_json(manifest, {})
        manifests[manifest.parent.resolve()] = payload if isinstance(payload, dict) else {}

    def manifest_for(path: Path) -> dict[str, Any]:
        current = path.parent.resolve()
        while current != root and root in current.parents:
            if current in manifests:
                return manifests[current]
            current = current.parent
        return {}

    candidates: list[Candidate] = []
    for scan_root in [inbox, unpacked_root]:
        for path in sorted(scan_root.rglob("*"), key=lambda p: str(p).casefold()):
            if not path.is_file() or path.name == "caseflow_upload.json":
                continue
            comp = component(path)
            if comp == "СКРІНШОТ" and not settings["includeScreenshots"]:
                continue
            if comp == "ПІДПИС" and not settings["includeSignatures"]:
                continue
            if comp in {"КАРТКА_РУХУ", "ПРОТОКОЛ_КЕП"} and not settings["includeCards"]:
                continue
            relative = path.relative_to(scan_root)
            proceeding_folder = relative.parts[0] if len(relative.parts) > 1 else "НОВЕ_ПРОВАДЖЕННЯ"
            flow_folder, flow = stream_from_path(relative)
            manifest = manifest_for(path) if scan_root == inbox else {}
            raw_channel = manifest.get("channel") or next(
                (part.split("__", 1)[1] for part in relative.parts if "__" in part and part[:8].isdigit()),
                "ІНШЕ",
            )
            uploaded_at = manifest.get("uploaded_at")
            candidates.append({
                "path": path,
                "scan_root": scan_root.name,
                "relative": relative,
                "proceeding_folder": proceeding_folder,
                "flow_folder": flow_folder,
                "flow": flow,
                "channel": str(raw_channel),
                "component": comp,
                "sha256": sha256(path),
                "uploaded_at": uploaded_at if isinstance(uploaded_at, str) else None,
            })

    template = find_template(root)
    workbook = ensure_workbook(template)
    for name, headers in SHEET_HEADERS.items():
        if name not in workbook.sheetnames:
            sheet = workbook.create_sheet(name)
            sheet.append(headers)
    docs_sheet = workbook["Документи"]
    files_sheet = workbook["Файли"]
    timeline_sheet = workbook["Хронологія"]
    batches_sheet = workbook["Партії"]
    existing_doc_ids = [docs_sheet.cell(row, 1).value for row in range(2, docs_sheet.max_row + 1)]
    existing_file_ids = [files_sheet.cell(row, 1).value for row in range(2, files_sheet.max_row + 1)]
    existing_event_ids = [timeline_sheet.cell(row, 1).value for row in range(2, timeline_sheet.max_row + 1)]
    existing_batch_ids = [batches_sheet.cell(row, 1).value for row in range(2, batches_sheet.max_row + 1)]
    doc_cursor = next_id(existing_doc_ids, "DOC")
    file_cursor = next_id(existing_file_ids, "FILE")
    event_cursor = next_id(existing_event_ids, "EVT")
    batch_cursor = next_id(existing_batch_ids, "BATCH")
    existing_hashes = {str(files_sheet.cell(row, 12).value or "").upper() for row in range(2, files_sheet.max_row + 1)}
    index_path = root / ".caseflow" / "index.json"
    index = read_json(index_path, {"hashes": {}, "batches": {}})
    hash_index = index.setdefault("hashes", {})
    batch_index = index.setdefault("batches", {})

    by_folder: defaultdict[Path, list[Candidate]] = defaultdict(list)
    for item in candidates:
        by_folder[item["path"].parent].append(item)
    logical_groups: list[LogicalGroup] = []
    evidence_only: list[Candidate] = []
    for folder, items in by_folder.items():
        mains = [item for item in items if item["component"] == "ОСНОВНИЙ"]
        auxiliaries = [item for item in items if item["component"] != "ОСНОВНИЙ"]
        if not mains:
            evidence_only.extend(auxiliaries)
            continue
        groups: list[LogicalGroup] = [{"main": main, "files": [main]} for main in mains]
        for auxiliary in auxiliaries:
            target_group = max(groups, key=lambda group: SequenceMatcher(None, auxiliary["path"].stem.casefold(), group["main"]["path"].stem.casefold()).ratio())
            target_group["files"].append(auxiliary)
        logical_groups.extend(groups)

    batch_groups: defaultdict[str, list[LogicalGroup]] = defaultdict(list)
    for group in logical_groups:
        main = group["main"]
        upload_folder = next((parent for parent in [main["path"].parent, *main["path"].parents] if parent.resolve() in manifests), main["path"].parent)
        batch_groups[str(upload_folder)].append(group)

    added_docs = added_files = added_events = duplicates = cross_proceeding = 0
    summary_tasks: list[dict[str, Any]] = []
    run_documents: list[dict[str, Any]] = []
    for batch_key, groups in batch_groups.items():
        batch_id = batch_index.get(batch_key)
        if not batch_id:
            batch_id = batch_cursor
            batch_index[batch_key] = batch_id
            batch_cursor = increment_id(batch_cursor)
        new_groups = []
        for group in groups:
            main = group["main"]
            if main["sha256"] in existing_hashes or main["sha256"] in hash_index:
                duplicates += 1
                continue
            new_groups.append(group)
        if not new_groups:
            continue
        first = new_groups[0]["main"]
        source_label = "Від суду + Мої документи" if len({g["main"]["flow"] for g in new_groups}) > 1 else first["flow"]
        uploaded_at = first["uploaded_at"]
        upload_dt = datetime.fromisoformat(uploaded_at) if uploaded_at else started
        append_styled(batches_sheet, [batch_id, folder_to_proceeding(first["proceeding_folder"]), source_label, int(batch_id.split("_")[1]), upload_dt.date(), len(new_groups), len(new_groups), "Опрацьовано", started.date(), "Створено універсальним конвеєром VARTA; обхід рекурсивний."])
        for group in new_groups:
            main = group["main"]
            doc_id = doc_cursor
            doc_cursor = increment_id(doc_cursor)
            file_texts = [text_from_file(item["path"]) for item in group["files"] if item["component"] in {"ОСНОВНИЙ", "КАРТКА_РУХУ"}]
            combined_text = " ".join(file_texts)
            # Preserve textual order: for a user document the addressee/header number is
            # more informative than an alphabetically sorted reference elsewhere in text.
            candidates_in_doc = list(dict.fromkeys(PROCEEDING_RE.findall(combined_text)))
            folder_proceeding = folder_to_proceeding(main["proceeding_folder"])
            primary_proceeding = candidates_in_doc[0] if main["flow"] == "Мої документи" and candidates_in_doc else (candidates_in_doc[0] if candidates_in_doc else folder_proceeding)
            is_cross = primary_proceeding != folder_proceeding and folder_proceeding not in {"НОВЕ/ПРОВАДЖЕННЯ", "НОВЕ_ПРОВАДЖЕННЯ"}
            if is_cross:
                cross_proceeding += 1
            doc_type = infer_type(main["path"])
            title = short_title(main["path"], doc_type)
            document_date = parse_document_date(main["path"], combined_text)
            submitted_at = parse_submission_time(combined_text)
            destination = processed_root / safe_segment(main["proceeding_folder"]) / main["flow_folder"] / doc_id
            destination.mkdir(parents=True, exist_ok=True)
            added_components: list[dict[str, str]] = []
            for component_number, component_item in enumerate(group["files"], 1):
                comp = component_item["component"]
                component_token = "ДОДАТОК" if comp == "СКРІНШОТ" else comp
                effective_date = submitted_at or document_date
                date_token = effective_date.strftime("%Y-%m-%d") if effective_date else "БЕЗ_ДАТИ"
                court_code = "СУД" if main["flow"] == "Від суду" else "КОРИСТУВАЧ"
                normalized = f"{date_token}__{court_code}__{normal_token(main['channel'])}__{normal_token(title)}__{doc_id}__{component_token}{compound_suffix(component_item['path'])}"
                target = destination / normalized
                if settings["includeOriginals"]:
                    shutil.copy2(component_item["path"], target)
                else:
                    target = component_item["path"]
                duplicate = component_item["sha256"] in existing_hashes or component_item["sha256"] in hash_index
                file_id = file_cursor
                file_cursor = increment_id(file_cursor)
                append_styled(files_sheet, [file_id, doc_id, component_token, component_number, component_item["path"].name, normalized, str(target.relative_to(root)), str(target), compound_suffix(component_item["path"]).lstrip("."), mime_for(component_item["path"]), component_item["path"].stat().st_size, component_item["sha256"], "Так" if duplicate else "Ні", "Відкривається", upload_dt.date(), started.date(), None, None, batch_id, primary_proceeding, f"Джерело: {component_item['scan_root']}; канал: {main['channel']}."])
                hash_index[component_item["sha256"]] = {"doc_id": doc_id, "file_id": file_id, "source": str(component_item["path"].relative_to(root)), "processed": str(target.relative_to(root)), "added_at": now_iso()}
                existing_hashes.add(component_item["sha256"])
                added_files += 1
                added_components.append({"component": component_token, "path": str(target.relative_to(root)), "sha256": component_item["sha256"]})
            card_state = "Є" if any(item["component"] == "КАРТКА_РУХУ" for item in group["files"]) else "Відсутня"
            protocol_state = "Є" if any(item["component"] == "ПРОТОКОЛ_КЕП" for item in group["files"]) else "Відсутній"
            signature_state = "Є" if any(item["component"] == "ПІДПИС" for item in group["files"]) else "Відсутній"
            notes = [f"Канал: {main['channel']}. Фактична папка: {folder_proceeding}."]
            if main["flow"] == "Мої документи" and not candidates_in_doc:
                notes.append("Первинне провадження не вдалося прочитати з документа; потрібна ручна перевірка.")
            if is_cross:
                notes.append(f"cross_proceeding: первинне {primary_proceeding}; фактичне розміщення {folder_proceeding}.")
            if len(candidates_in_doc) > 1:
                notes.append("У документі виявлено кілька номерів проваджень: " + ", ".join(candidates_in_doc) + ".")
            append_styled(docs_sheet, [doc_id, None, None, document_date, submitted_at, primary_proceeding, "Суд" if main["flow"] == "Від суду" else "Користувач", main["flow"], title, doc_type, "Потребує короткого фактологічного змісту", batch_id, "Є", None, sum(item["component"] not in {"ОСНОВНИЙ", "КАРТКА_РУХУ", "ПРОТОКОЛ_КЕП", "ПІДПИС"} for item in group["files"]), card_state, protocol_state, signature_state, "ПЕРЕВІРИТИ", "Перевірити реквізити та короткий зміст", str(destination.relative_to(root)), None, None, " ".join(notes), "Процесуальний зміст", "cross_proceeding" if is_cross else "Не звірено"])
            added_docs += 1
            if submitted_at:
                event_id = event_cursor
                event_cursor = increment_id(event_cursor)
                append_styled(timeline_sheet, [event_id, submitted_at, primary_proceeding, "Мій документ" if main["flow"] == "Мої документи" else "Від суду", "Користувач" if main["flow"] == "Мої документи" else "Суд", "Подання документа" if main["flow"] == "Мої документи" else "Надходження документа", doc_id, f"{doc_type}: {title}", "Потребує перевірки", "Звірити картку руху", None, batch_id, str(destination.relative_to(root)), "Час взято з події «Надіслано до суду» у локальній картці руху."])
                added_events += 1
            else:
                review_marker = review_root / safe_segment(main["proceeding_folder"]) / f"{doc_id}__ПОТРІБЕН_ЧАС_ПОДАННЯ.txt"
                review_marker.parent.mkdir(parents=True, exist_ok=True)
                review_marker.write_text(f"Документ: {main['path']}\nПричина: точний час подання/надходження не підтверджено локальним джерелом.\n", encoding="utf-8")
            summary_tasks.append({"doc_id": doc_id, "source_path": str(main["path"]), "short_summary": "", "confidence": "low", "note": "Очікує агента короткого змісту"})
            run_documents.append({"doc_id": doc_id, "primary_proceeding": primary_proceeding, "folder_proceeding": folder_proceeding, "cross_proceeding": is_cross, "submitted_at": submitted_at.isoformat(timespec="minutes") if submitted_at else None, "components": added_components})

    sort_timeline(timeline_sheet)
    for sheet in [batches_sheet, docs_sheet, timeline_sheet, files_sheet]:
        expand_tables(sheet)
    if hasattr(workbook, "calculation"):
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    export_dir = root / "03_РЕЄСТР" / "exports"
    case_number = read_json(root / ".caseflow" / "config.json", {}).get("case_number", root.name)
    output_book = export_dir / f"Реєстр_{safe_segment(case_number).replace('/', '_')}__{stamp}.xlsx"
    if settings["autoRegister"]:
        workbook.save(output_book)
        validation = load_workbook(output_book, read_only=True, data_only=False)
        validation.close()
        (root / "03_РЕЄСТР" / "ОСТАННІЙ_РЕЄСТР.txt").write_text(str(output_book), encoding="utf-8")

    write_json(index_path, index)
    write_json(root / "tmp" / "summary_agent" / "pending_tasks.json", {"schema_version": 1, "generated_at": now_iso(), "items": summary_tasks})
    run_statistics = {
        "documents_added": added_docs,
        "files_added": added_files,
        "events_added": added_events,
        "duplicates_skipped": duplicates,
        "cross_proceeding": cross_proceeding,
        "evidence_without_document": len(evidence_only),
        "archives_entries": len(extraction_log),
    }
    run: dict[str, Any] = {
        "run_id": stamp,
        "started_at": started.astimezone().isoformat(timespec="seconds"),
        "finished_at": now_iso(),
        "settings": settings,
        "statistics": run_statistics,
        "register": str(output_book.relative_to(root)) if settings["autoRegister"] else None,
        "template": str(template.relative_to(root)) if template and root in template.parents else str(template) if template else None,
        "documents": run_documents,
        "extraction_log": extraction_log,
    }
    run_path = root / "tmp" / "caseflow_runs" / f"{stamp}.json"
    write_json(run_path, run)
    print(json.dumps({"run": str(run_path.relative_to(root)), "register": run["register"], **run_statistics}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
