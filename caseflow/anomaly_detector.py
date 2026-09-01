from __future__ import annotations

import argparse
import hashlib
import json
import re
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PdfReader: Any = None
try:
    from pypdf import PdfReader as _PdfReader
except ImportError:  # pragma: no cover
    pass
else:
    PdfReader = _PdfReader


PROCEEDING_RE = re.compile(r"(?<![\w/])\d{1,3}(?:-[А-Яа-яІіЇїЄєҐґ]+)?/\d{3}/\d+/\d{2,4}(?![\w/])")
HASH_RE = re.compile(r"^[0-9A-F]{64}$", re.IGNORECASE)
SEVERITY_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}
SEVERITY_WEIGHT = {"critical": 100, "high": 40, "medium": 15, "low": 5, "info": 0}
NEUTRAL_NOTICE = (
    "Автоматичний сигнал показує неузгодженість джерел. Він не доводить умисел, "
    "порушення чи втручання без додаткової перевірки первинних матеріалів."
)
DETECTOR_VERSION = "1.1.0"


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def read_json(path: Path, default):
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return default


def write_json(path: Path, value) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(path)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def clean(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(key): json_value(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [json_value(item) for item in value]
    return value


def parse_number(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = clean(value)
    if not text:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def folder_to_proceeding(folder: str) -> str:
    match = re.fullmatch(r"(\d{1,3})_(\d{3})_(\d+)_(\d{2,4})", folder)
    return "/".join(match.groups()) if match else folder


def resolve_local_path(root: Path, row: dict) -> Path | None:
    full = clean(row.get("Повний локальний шлях"))
    relative = clean(row.get("Відносний шлях"))
    candidates = []
    if full:
        candidates.append(Path(full))
    if relative:
        candidates.append(root / relative)
    for path in candidates:
        try:
            resolved = path.resolve()
        except OSError:
            continue
        if root == resolved or root in resolved.parents:
            return resolved
    return None


def text_from_file(path: Path, limit: int = 80000) -> str:
    suffix = path.suffix.lower()
    try:
        if suffix == ".pdf" and PdfReader is not None:
            reader = PdfReader(path, strict=False)
            return " ".join((page.extract_text() or "") for page in reader.pages[:16])[:limit]
        if suffix in {".html", ".htm", ".txt", ".rtf"}:
            raw = path.read_bytes()
            for encoding in ("utf-8-sig", "utf-8", "windows-1251"):
                try:
                    text = raw.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                text = raw.decode("utf-8", errors="replace")
            if suffix in {".html", ".htm"}:
                text = re.sub(r"(?is)<script.*?</script>|<style.*?</style>", " ", text)
                text = re.sub(r"(?s)<[^>]+>", " ", text)
            return text[:limit]
        if suffix == ".docx":
            with zipfile.ZipFile(path) as archive:
                xml = archive.read("word/document.xml").decode("utf-8", errors="replace")
            return re.sub(r"(?s)<[^>]+>", " ", xml)[:limit]
    except Exception:
        return ""
    return ""


def infer_document_type(text: str) -> str | None:
    folded = text.casefold()
    rules = [
        ("касаційн", "Касаційна скарга"),
        ("апеляційн", "Апеляційна скарга"),
        ("позов", "Позовна заява"),
        ("клопотан", "Клопотання"),
        ("запереч", "Заперечення"),
        ("відзив", "Відзив"),
        ("скарг", "Скарга"),
        ("заяв", "Заява"),
        ("ухвал", "Ухвала"),
        ("рішенн", "Рішення"),
        ("постанов", "Постанова"),
        ("вирок", "Вирок"),
        ("судовий наказ", "Судовий наказ"),
    ]
    matches = [(folded.find(token), -len(token), kind) for token, kind in rules if token in folded]
    return min(matches)[2] if matches else None


def proceeding_after(text: str, labels: tuple[str, ...]) -> str | None:
    for label in labels:
        match = re.search(label, text, flags=re.IGNORECASE)
        if not match:
            continue
        proceeding = PROCEEDING_RE.search(text[match.end() : match.end() + 180])
        if proceeding:
            return proceeding.group(0)
    return None


def has_time_precision(value: Any) -> bool:
    if isinstance(value, datetime):
        return value.time() != datetime.min.time()
    if isinstance(value, date):
        return False
    return bool(re.search(r"\d{1,2}:\d{2}", clean(value)))


def sheet_rows(workbook, name: str) -> list[dict]:
    if name not in workbook.sheetnames:
        return []
    sheet = workbook[name]
    headers = [clean(cell.value) for cell in sheet[1]]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append({headers[index]: values[index] if index < len(values) else None for index in range(len(headers)) if headers[index]})
    return rows


def find_register(root: Path, explicit: Path | None) -> Path:
    if explicit:
        candidate = explicit if explicit.is_absolute() else root / explicit
        if candidate.exists():
            return candidate.resolve()
        raise FileNotFoundError(candidate)
    pointer = root / "03_РЕЄСТР" / "ОСТАННІЙ_РЕЄСТР.txt"
    if pointer.exists():
        candidate = Path(pointer.read_text(encoding="utf-8-sig").strip())
        if candidate.exists():
            return candidate.resolve()
    candidates = [path for path in (root / "03_РЕЄСТР" / "exports").glob("*.xlsx") if not path.name.startswith("~$")]
    candidates += [path for path in (root / "03_РЕЄСТР").glob("*.xlsx") if not path.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError("У 03_РЕЄСТР не знайдено книги .xlsx")
    return max(candidates, key=lambda path: path.stat().st_mtime).resolve()


class Findings:
    def __init__(self, statuses: dict):
        self.items: list[dict] = []
        self.seen: set[str] = set()
        self.statuses = statuses

    def add(
        self,
        rule_id: str,
        title: str,
        severity: str,
        confidence: str,
        category: str,
        discrepancy: str,
        facts: list[dict],
        *,
        doc_ids: list[str] | None = None,
        proceedings: dict | None = None,
        why_flagged: str,
        next_check: str,
    ) -> None:
        identity = json.dumps(
            {
                "rule": rule_id,
                "docs": sorted(doc_ids or []),
                "facts": [{"source": fact.get("source_path"), "field": fact.get("field"), "value": clean(fact.get("value"))} for fact in facts],
            },
            ensure_ascii=False,
            sort_keys=True,
        )
        fingerprint = hashlib.sha256(identity.encode("utf-8")).hexdigest().upper()[:24]
        if fingerprint in self.seen:
            return
        self.seen.add(fingerprint)
        status_meta = self.statuses.get(fingerprint, {})
        self.items.append(
            {
                "anomaly_id": f"ANOM_{fingerprint[:12]}",
                "fingerprint": fingerprint,
                "detector_version": DETECTOR_VERSION,
                "detected_at": now_iso(),
                "rule_id": rule_id,
                "title": title,
                "severity": severity,
                "confidence": confidence,
                "category": category,
                "status": status_meta.get("status", "open"),
                "status_note": status_meta.get("note", ""),
                "doc_ids": sorted(set(doc_ids or [])),
                "proceedings": proceedings or {},
                "discrepancy": discrepancy,
                "facts": facts,
                "why_flagged": why_flagged,
                "next_check": next_check,
                "possible_benign_explanations": [
                    "технічна помилка імпорту або ручного обліку",
                    "коректна версійність, приєднання чи перенесення, яке ще не підтверджене локальним рішенням",
                ],
                "neutral_notice": NEUTRAL_NOTICE,
            }
        )


def fact(source_type: str, source_path: str, field: str, value: Any, observed_at: str | None = None) -> dict:
    return {
        "source_type": source_type,
        "source_path": source_path,
        "field": field,
        "value": json_value(value),
        "observed_at": observed_at or now_iso(),
    }


def event_stage(row: dict) -> int | None:
    text = f"{clean(row.get('Тип події'))} {clean(row.get('Документ / подія'))}".casefold()
    if any(token in text for token in ("підпис", "кеп")):
        return 1
    if any(token in text for token in ("подан", "надіслано до суду", "направлен")):
        return 2
    if any(token in text for token in ("достав", "надходжен")):
        return 3
    if any(token in text for token in ("реєстрац", "зареєстр")):
        return 4
    return None


def run_detector(root: Path, register_path: Path) -> dict:
    generated_at = now_iso()
    status_path = root / ".caseflow" / "anomaly_status.json"
    statuses = read_json(status_path, {})
    findings = Findings(statuses)
    text_cache_path = root / ".caseflow" / "anomaly_text_cache.json"
    text_cache = read_json(text_cache_path, {})
    computed_hashes: dict[str, str] = {}
    workbook = load_workbook(register_path, data_only=False, read_only=False)
    documents = sheet_rows(workbook, "Документи")
    files = sheet_rows(workbook, "Файли")
    events = sheet_rows(workbook, "Хронологія")
    register_rel = str(register_path.relative_to(root)) if root in register_path.parents else str(register_path)
    docs_by_id = {clean(row.get("ID документа")): row for row in documents if clean(row.get("ID документа"))}
    files_by_doc: dict[str, list[dict]] = defaultdict(list)
    events_by_doc: dict[str, list[dict]] = defaultdict(list)
    for row in files:
        files_by_doc[clean(row.get("ID документа"))].append(row)
    for row in events:
        events_by_doc[clean(row.get("ID документа"))].append(row)

    # Stable ID collisions.
    for sheet_name, rows, field_name in (("Документи", documents, "ID документа"), ("Файли", files, "ID файла"), ("Хронологія", events, "ID події")):
        counts = Counter(clean(row.get(field_name)) for row in rows if clean(row.get(field_name)))
        for value, count in counts.items():
            if count > 1:
                findings.add(
                    "ID_COLLISION",
                    f"Повторний ідентифікатор {value}",
                    "high",
                    "high",
                    "registry_integrity",
                    f"У вкладці «{sheet_name}» ідентифікатор {value} використано {count} рази.",
                    [fact("registry", register_rel, f"{sheet_name}.{field_name}", value)],
                    doc_ids=[value] if field_name == "ID документа" else [],
                    why_flagged="Стабільний ідентифікатор має однозначно посилатися на один запис.",
                    next_check="Порівняти рядки з однаковим ID та встановити, чи це дублікат, версія або помилка імпорту.",
                )

    # Document-level semantic and chronology checks.
    for doc_id, row in docs_by_id.items():
        proceeding = clean(row.get("Провадження"))
        flow = clean(row.get("Потік"))
        status = clean(row.get("Статус звірки")).casefold()
        notes = clean(row.get("Примітки"))
        source_path = register_rel
        if status == "cross_proceeding" or "cross_proceeding" in notes.casefold():
            stated_primary = proceeding_after(
                notes,
                (r"primary proceeding\s*[:—-]?", r"первинн(?:е|ий|ого)?(?:\s+провадження|\s+зв['’]язок)?\s*[:—-]?", r"основний текст\s*[:—-]?"),
            )
            actual_proceeding = proceeding_after(
                notes,
                (r"cabinet proceeding\s*[:—-]?", r"фактичн(?:е|а)(?:\s+розміщення|\s+папка|\s+апеляційне провадження)?\s*[:—-]?", r"суд створив\s*[:—-]?"),
            )
            proceedings = {"registry": proceeding}
            if stated_primary:
                proceedings["stated_primary"] = stated_primary
            if actual_proceeding:
                proceedings["actual"] = actual_proceeding
            labels = {"registry": "Реєстр", "stated_primary": "первинне за документом/приміткою", "actual": "фактичне розміщення"}
            described = "; ".join(f"{labels.get(key, key)}: {value}" for key, value in proceedings.items())
            findings.add(
                "CROSS_PROCEEDING",
                "Документ пов’язано з іншим провадженням",
                "high",
                "medium",
                "proceeding",
                described + ".",
                [fact("registry", source_path, "Документи.Провадження", proceeding), fact("registry", source_path, "Документи.Примітки", notes)],
                doc_ids=[doc_id],
                proceedings=proceedings,
                why_flagged="Перенесення документа до іншого провадження може змінити контекст його розгляду й потребує окремого пояснення.",
                next_check="Відкрити сам документ, картку руху та процесуальне рішення про приєднання або передачу матеріалу.",
            )
        if status in {"mismatch", "невідповідність"}:
            findings.add(
                "REGISTERED_MISMATCH",
                "У Реєстрі вже зафіксована невідповідність",
                "high",
                "medium",
                "source_mismatch",
                f"Для {doc_id} встановлено статус «{row.get('Статус звірки')}».",
                [fact("registry", source_path, "Документи.Статус звірки", row.get("Статус звірки")), fact("registry", source_path, "Документи.Примітки", notes)],
                doc_ids=[doc_id],
                proceedings={"primary": proceeding},
                why_flagged="Статус mismatch означає, що щонайменше два джерела не узгоджуються.",
                next_check="Встановити точні поля розбіжності та додати посилання на обидва первинні джерела.",
            )
        expected = parse_number(row.get("Додатків очікується"))
        actual_attachment_count = parse_number(row.get("Додатків фактично"))
        if (
            expected is not None
            and actual_attachment_count is not None
            and expected != actual_attachment_count
        ):
            severity = "high" if actual_attachment_count < expected else "medium"
            findings.add(
                "ATTACHMENT_COUNT_MISMATCH",
                "Не збігається кількість додатків",
                severity,
                "high",
                "attachments",
                f"Заявлено додатків: {expected}; фактично знайдено: {actual_attachment_count}.",
                [fact("registry", source_path, "Додатків очікується", expected), fact("registry", source_path, "Додатків фактично", actual_attachment_count)],
                doc_ids=[doc_id],
                proceedings={"primary": proceeding},
                why_flagged="Відсутній або зайвий додаток може змінювати повноту поданого пакета.",
                next_check="Звірити перелік додатків у тексті документа, картці руху, ZIP та папці DOC_ID.",
            )
        missing_fields = []
        for field_name in ("Основний файл", "Картка руху", "Протокол КЕП", "Підпис"):
            value = clean(row.get(field_name)).casefold()
            if value in {"відсутній", "відсутня", "пошкоджений", "потребує перевірки"}:
                missing_fields.append((field_name, row.get(field_name)))
        if missing_fields:
            findings.add(
                "MISSING_OR_UNVERIFIED_COMPONENT",
                "Відсутній або неперевірений компонент документа",
                "medium",
                "high",
                "completeness",
                "; ".join(f"{name}: {value}" for name, value in missing_fields),
                [fact("registry", source_path, name, value) for name, value in missing_fields],
                doc_ids=[doc_id],
                proceedings={"primary": proceeding},
                why_flagged="Без основного файла, картки руху або підпису неможливо повністю відтворити походження та рух документа.",
                next_check="Знайти відсутній компонент у вихідному пакеті або зафіксувати, що він не надавався.",
            )
        document_date = parse_datetime(row.get("Дата документа"))
        submitted_at = parse_datetime(row.get("Дата надходження/подання"))
        if document_date and submitted_at and document_date > submitted_at:
            findings.add(
                "DOCUMENT_AFTER_SUBMISSION",
                "Дата документа пізніша за час його подання",
                "high",
                "high",
                "chronology",
                f"Дата документа {document_date.isoformat(sep=' ')} пізніша за подання {submitted_at.isoformat(sep=' ')}.",
                [fact("registry", source_path, "Дата документа", row.get("Дата документа")), fact("registry", source_path, "Дата надходження/подання", row.get("Дата надходження/подання"))],
                doc_ids=[doc_id],
                proceedings={"primary": proceeding},
                why_flagged="Документ не може бути остаточно створений після підтвердженого моменту його подання без пояснення версії або виправлення дати.",
                next_check="Перевірити метадані файла, КЕП, картку руху та чи не є це повторною версією документа.",
            )

        # Read primary document only to verify proceeding/type claims.
        main_rows = [file_row for file_row in files_by_doc.get(doc_id, []) if clean(file_row.get("Компонент")).upper() == "ОСНОВНИЙ"]
        main_paths = [resolve_local_path(root, file_row) for file_row in main_rows]
        existing_main_paths = [path for path in main_paths if path and path.exists()]
        if existing_main_paths:
            main_path = existing_main_paths[0]
            cache_key = sha256_file(main_path)
            computed_hashes[str(main_path)] = cache_key
            cached_text = text_cache.get(cache_key, {}) if isinstance(text_cache, dict) else {}
            if cached_text:
                mentioned = list(cached_text.get("proceedings", []))
                inferred_type = cached_text.get("inferred_type")
            else:
                text = text_from_file(main_path)
                mentioned = sorted(set(PROCEEDING_RE.findall(text)), key=str.casefold)
                inferred_type = infer_document_type(f"{main_path.name} {text[:700]}")
                text_cache[cache_key] = {
                    "source_name": main_path.name,
                    "proceedings": mentioned,
                    "inferred_type": inferred_type,
                    "cached_at": generated_at,
                }
            if flow.casefold() == "мої документи".casefold() and mentioned and proceeding not in mentioned:
                findings.add(
                    "PRIMARY_PROCEEDING_CONFLICT",
                    "Провадження Реєстру не підтверджується текстом документа користувача",
                    "high",
                    "medium",
                    "proceeding",
                    f"Реєстр: {proceeding}; у тексті документа: {', '.join(mentioned)}.",
                    [fact("registry", source_path, "Документи.Провадження", proceeding), fact("document", str(existing_main_paths[0].relative_to(root)), "Номери у тексті", mentioned)],
                    doc_ids=[doc_id],
                    proceedings={"registry": proceeding, "mentioned": mentioned},
                    why_flagged="Для документа користувача первинний номер визначається самим документом, а не папкою чи карткою кабінету.",
                    next_check="Встановити, до якого номера належить прохальна частина документа, і окремо зафіксувати фактичне розміщення судом.",
                )
            if len(mentioned) > 1:
                findings.add(
                    "MULTIPLE_PROCEEDINGS_IN_DOCUMENT",
                    "У документі виявлено кілька номерів проваджень",
                    "medium",
                    "medium",
                    "proceeding",
                    f"Текст містить: {', '.join(mentioned)}.",
                    [fact("document", str(existing_main_paths[0].relative_to(root)), "Номери у тексті", mentioned)],
                    doc_ids=[doc_id],
                    proceedings={"registry": proceeding, "mentioned": mentioned},
                    why_flagged="Кілька номерів можуть бути коректними посиланнями, але автоматичне визначення первинного зв’язку стає неоднозначним.",
                    next_check="Прочитати вступну та прохальну частини й відокремити основне провадження від згаданих пов’язаних справ.",
                )
            registered_type = clean(row.get("Тип документа"))
            if inferred_type and registered_type and inferred_type.casefold() != registered_type.casefold() and registered_type.casefold() != "інше":
                findings.add(
                    "DOCUMENT_TYPE_CONFLICT",
                    "Тип документа не узгоджується з його назвою або текстом",
                    "medium",
                    "medium",
                    "metadata",
                    f"У Реєстрі: {registered_type}; за назвою/текстом: {inferred_type}.",
                    [fact("registry", source_path, "Тип документа", registered_type), fact("document", str(existing_main_paths[0].relative_to(root)), "Автоматично розпізнаний тип", inferred_type)],
                    doc_ids=[doc_id],
                    proceedings={"primary": proceeding},
                    why_flagged="Зміна типу може впливати на спосіб відображення та пошуку документа.",
                    next_check="Звірити офіційний заголовок першої сторінки й назву в картці руху.",
                )

        staged = []
        for event in events_by_doc.get(doc_id, []):
            stage = event_stage(event)
            moment = parse_datetime(event.get("Дата / час"))
            if stage is not None and moment and has_time_precision(event.get("Дата / час")):
                staged.append((stage, moment, event))
        for left in staged:
            for right in staged:
                if left[0] < right[0] and left[1] > right[1]:
                    findings.add(
                        "EVENT_SEQUENCE_CONTRADICTION",
                        "Етапи руху документа розташовані у неможливому порядку",
                        "high",
                        "high",
                        "chronology",
                        f"Етап «{left[2].get('Тип події')}» має час {left[1]}, а пізніший етап «{right[2].get('Тип події')}» — {right[1]}.",
                        [fact("registry", source_path, f"Хронологія.{left[2].get('ID події')}", left[1].isoformat(sep=" ")), fact("registry", source_path, f"Хронологія.{right[2].get('ID події')}", right[1].isoformat(sep=" "))],
                        doc_ids=[doc_id],
                        proceedings={"primary": proceeding},
                        why_flagged="Підписання, подання, доставка та реєстрація повинні мати послідовний часовий порядок.",
                        next_check="Перевірити точні мітки часу у картці руху та не змішувати дату документа з датою реєстрації.",
                    )
                    break

    # Duplicate or version conflicts by file metadata.
    by_hash: dict[str, list[dict]] = defaultdict(list)
    by_name: dict[str, list[dict]] = defaultdict(list)
    for row in files:
        digest = clean(row.get("SHA-256")).upper()
        if HASH_RE.fullmatch(digest):
            by_hash[digest].append(row)
        name_key = clean(row.get("Оригінальна назва")).casefold()
        if name_key:
            by_name[name_key].append(row)
    for digest, rows in by_hash.items():
        doc_ids = sorted({clean(row.get("ID документа")) for row in rows if clean(row.get("ID документа"))})
        duplicate_proceedings = sorted({clean(row.get("Провадження")) for row in rows if clean(row.get("Провадження"))})
        if len(doc_ids) > 1 and len(duplicate_proceedings) > 1:
            findings.add(
                "SAME_FILE_DIFFERENT_CONTEXT",
                "Той самий файл обліковано в різних документах або провадженнях",
                "medium",
                "high",
                "duplicates",
                f"SHA-256 {digest[:16]}… пов’язаний з DOC_ID {', '.join(doc_ids)} і провадженнями {', '.join(duplicate_proceedings)}.",
                [fact("registry", register_rel, "Файли.SHA-256", digest), fact("registry", register_rel, "Пов’язані DOC_ID", doc_ids)],
                doc_ids=doc_ids,
                proceedings={"listed": duplicate_proceedings},
                why_flagged="Повторне використання байт-в-байт однакового файла може бути законним, але потребує пояснення контексту.",
                next_check="Порівняти роль файла в кожному пакеті та перевірити, чи не створено зайві логічні документи.",
            )
    for name, rows in by_name.items():
        components = {clean(row.get("Компонент")).upper() for row in rows}
        if components and components <= {"ПІДПИС", "КАРТКА_РУХУ", "ПРОТОКОЛ_КЕП", "СКРІНШОТ"}:
            continue
        hashes = sorted({clean(row.get("SHA-256")).upper() for row in rows if HASH_RE.fullmatch(clean(row.get("SHA-256")).upper())})
        if len(hashes) > 1:
            doc_ids = sorted({clean(row.get("ID документа")) for row in rows if clean(row.get("ID документа"))})
            findings.add(
                "SAME_NAME_DIFFERENT_CONTENT",
                "Однакова назва використана для різного вмісту",
                "low",
                "high",
                "versions",
                f"Назва «{rows[0].get('Оригінальна назва')}» має {len(hashes)} різні SHA-256.",
                [fact("registry", register_rel, "Оригінальна назва", rows[0].get("Оригінальна назва")), fact("registry", register_rel, "Різні SHA-256", hashes)],
                doc_ids=doc_ids,
                why_flagged="Однакова назва приховує різні версії або різні документи й ускладнює доказування походження.",
                next_check="Відкрити обидві версії, порівняти зміст і час появи, після чого позначити версії окремо.",
            )

    for doc_id, rows in files_by_doc.items():
        main_hashes = sorted({clean(row.get("SHA-256")).upper() for row in rows if clean(row.get("Компонент")).upper() == "ОСНОВНИЙ" and HASH_RE.fullmatch(clean(row.get("SHA-256")).upper())})
        if len(main_hashes) > 1:
            findings.add(
                "MULTIPLE_MAIN_VERSIONS",
                "Для одного DOC_ID існує кілька різних основних файлів",
                "high",
                "high",
                "versions",
                f"{doc_id} має {len(main_hashes)} різні хеші основного файла.",
                [fact("registry", register_rel, "DOC_ID", doc_id), fact("registry", register_rel, "SHA-256 основних файлів", main_hashes)],
                doc_ids=[doc_id],
                why_flagged="Різні основні файли під одним ID можуть бути версіями, але мають бути явно розмежовані.",
                next_check="Порівняти версії, визначити первинну та пізнішу, не перезаписуючи історію.",
            )

    # File-system integrity and longitudinal snapshot.
    snapshot_state_path = root / ".caseflow" / "anomaly_snapshot.json"
    previous_state = read_json(snapshot_state_path, {"runs": 0, "paths": {}})
    previous_paths = previous_state.get("paths", {})
    current_paths: dict[str, dict] = {}
    for row in files:
        local_path = resolve_local_path(root, row)
        relative = clean(row.get("Відносний шлях"))
        registered_hash = clean(row.get("SHA-256")).upper()
        doc_id = clean(row.get("ID документа"))
        if local_path is None:
            continue
        safe_relative = str(local_path.relative_to(root)) if root in local_path.parents else relative
        if not local_path.exists():
            findings.add(
                "REGISTERED_FILE_MISSING",
                "Файл із Реєстру відсутній на диску",
                "critical",
                "high",
                "filesystem",
                f"Реєстр посилається на {safe_relative}, але файла немає.",
                [fact("registry", register_rel, "Файли.Відносний шлях", safe_relative), fact("filesystem", safe_relative, "Наявність", "відсутній")],
                doc_ids=[doc_id] if doc_id else [],
                why_flagged="Втрата зареєстрованого файла руйнує можливість перевірити його зміст і хеш.",
                next_check="Перевірити переміщення, карантин антивірусу, резервні копії та журнал запусків.",
            )
            continue
        stat = local_path.stat()
        cache = previous_paths.get(safe_relative, {})
        if cache.get("size") == stat.st_size and cache.get("mtime_ns") == stat.st_mtime_ns and cache.get("sha256"):
            actual_hash = cache["sha256"]
        else:
            actual_hash = computed_hashes.get(str(local_path)) or sha256_file(local_path)
        current_paths[safe_relative] = {
            "sha256": actual_hash,
            "size": stat.st_size,
            "mtime_ns": stat.st_mtime_ns,
            "doc_id": doc_id,
            "first_seen": cache.get("first_seen", generated_at),
            "last_seen": generated_at,
        }
        if HASH_RE.fullmatch(registered_hash) and actual_hash != registered_hash:
            findings.add(
                "FILE_HASH_MISMATCH",
                "Фактичний файл не відповідає хешу в Реєстрі",
                "critical",
                "high",
                "filesystem",
                f"У Реєстрі: {registered_hash}; на диску: {actual_hash}.",
                [fact("registry", register_rel, "Файли.SHA-256", registered_hash), fact("filesystem", safe_relative, "Фактичний SHA-256", actual_hash)],
                doc_ids=[doc_id] if doc_id else [],
                why_flagged="Зміна байтів після реєстрації є прямою ознакою заміни або пошкодження локальної копії.",
                next_check="Негайно зберегти обидві версії, перевірити резервну копію й джерело первинного завантаження.",
            )
        if cache.get("sha256") and cache["sha256"] != actual_hash:
            findings.add(
                "FILE_CHANGED_BETWEEN_SCANS",
                "Файл змінився між двома перевірками",
                "critical",
                "high",
                "longitudinal",
                f"Попередній SHA-256: {cache['sha256']}; поточний: {actual_hash}.",
                [fact("snapshot", safe_relative, "Попередній SHA-256", cache["sha256"], cache.get("last_seen")), fact("filesystem", safe_relative, "Поточний SHA-256", actual_hash)],
                doc_ids=[doc_id] if doc_id else [],
                why_flagged="Вміст за тим самим шляхом змінився після попередньої фіксації.",
                next_check="Зупинити автоматичне очищення, створити незмінювані копії та встановити процес, який переписав файл.",
            )
    for safe_relative, cache in previous_paths.items():
        if safe_relative not in current_paths and cache.get("doc_id") and not (root / safe_relative).exists():
            findings.add(
                "FILE_DISAPPEARED_BETWEEN_SCANS",
                "Раніше зафіксований файл зник",
                "critical",
                "high",
                "longitudinal",
                f"Файл {safe_relative} був доступний {cache.get('last_seen')}, але зараз відсутній.",
                [fact("snapshot", safe_relative, "Остання наявність", cache.get("last_seen")), fact("filesystem", safe_relative, "Поточна наявність", "відсутній")],
                doc_ids=[cache.get("doc_id")],
                why_flagged="Зникнення матеріалу після його фіксації потребує відновлення ланцюга зберігання.",
                next_check="Перевірити журнали переміщення, кошик, резервні копії та чи створено новий шлях із тим самим хешем.",
            )

    # Existing cabinet verification artifacts remain useful as imported evidence, without launching a verifier.
    cabinet_report = root / "tmp" / "browser_verification" / "cabinet_check.json"
    cabinet = read_json(cabinet_report, {})
    for row in cabinet.get("rows", []) if isinstance(cabinet, dict) else []:
        status = clean(row.get("status")).casefold()
        if status not in {"mismatch", "cross_proceeding"}:
            continue
        doc_id = clean(row.get("doc_id"))
        primary = clean(row.get("primary_proceeding"))
        cabinet_proceeding = clean(row.get("cabinet_proceeding"))
        findings.add(
            "IMPORTED_CABINET_DISCREPANCY",
            "Попередня звірка кабінету зафіксувала розбіжність",
            "high",
            "medium",
            "cabinet_evidence",
            f"Статус: {status}; первинне провадження: {primary or 'не вказано'}; кабінет: {cabinet_proceeding or 'не вказано'}.",
            [fact("cabinet_report", str(cabinet_report.relative_to(root)), "status", status, cabinet.get("checked_at")), fact("cabinet_report", str(cabinet_report.relative_to(root)), "evidence_note", row.get("evidence_note"), cabinet.get("checked_at"))],
            doc_ids=[doc_id] if doc_id else [],
            proceedings={"primary": primary, "cabinet": cabinet_proceeding},
            why_flagged="Це вже зафіксована різниця між локальними матеріалами та станом кабінету на дату попередньої перевірки.",
            next_check="Звірити з актуальним скріншотом або сторінкою лише для цього конкретного документа.",
        )

    findings.items.sort(key=lambda item: (SEVERITY_ORDER.get(item["severity"], 99), item["rule_id"], item["anomaly_id"]))
    counts = Counter(item["severity"] for item in findings.items if item["status"] == "open")
    risk_points = sum(SEVERITY_WEIGHT.get(item["severity"], 0) for item in findings.items if item["status"] == "open")
    summary = {
        "total": len(findings.items),
        "open": sum(item["status"] == "open" for item in findings.items),
        "critical": counts.get("critical", 0),
        "high": counts.get("high", 0),
        "medium": counts.get("medium", 0),
        "low": counts.get("low", 0),
        "info": counts.get("info", 0),
        "risk_score": min(100, risk_points),
    }
    report = {
        "schema_version": 1,
        "detector_version": DETECTOR_VERSION,
        "generated_at": generated_at,
        "register": register_rel,
        "notice": NEUTRAL_NOTICE,
        "summary": summary,
        "rules_evaluated": [
            "CROSS_PROCEEDING", "REGISTERED_MISMATCH", "ATTACHMENT_COUNT_MISMATCH",
            "MISSING_OR_UNVERIFIED_COMPONENT", "DOCUMENT_AFTER_SUBMISSION", "PRIMARY_PROCEEDING_CONFLICT",
            "MULTIPLE_PROCEEDINGS_IN_DOCUMENT", "DOCUMENT_TYPE_CONFLICT", "EVENT_SEQUENCE_CONTRADICTION",
            "SAME_FILE_DIFFERENT_CONTEXT", "SAME_NAME_DIFFERENT_CONTENT", "MULTIPLE_MAIN_VERSIONS",
            "REGISTERED_FILE_MISSING", "FILE_HASH_MISMATCH", "FILE_CHANGED_BETWEEN_SCANS",
            "FILE_DISAPPEARED_BETWEEN_SCANS", "ID_COLLISION", "IMPORTED_CABINET_DISCREPANCY",
        ],
        "findings": findings.items,
    }
    output_dir = root / "tmp" / "caseflow_anomalies"
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    archive_path = output_dir / "runs" / f"{stamp}.json"
    latest_path = output_dir / "latest.json"
    report["output"] = str(latest_path.relative_to(root))
    write_json(archive_path, report)
    write_json(latest_path, report)
    write_json(snapshot_state_path, {"runs": int(previous_state.get("runs", 0)) + 1, "updated_at": generated_at, "paths": current_paths})
    write_json(text_cache_path, text_cache)
    return report


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Детермінований контроль нестиковок VARTA")
    parser.add_argument("--root", type=Path, required=True)
    parser.add_argument("--register", type=Path)
    args = parser.parse_args(argv)
    root = args.root.resolve()
    register_path = find_register(root, args.register)
    report = run_detector(root, register_path)
    print(json.dumps({"output": report["output"], "register": report["register"], **report["summary"]}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
