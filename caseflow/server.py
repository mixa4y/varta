from __future__ import annotations

import argparse
import base64
import ctypes
import hashlib
import importlib
import importlib.util
import io
import json
import mimetypes
import os
import re
import secrets
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
import webbrowser
import zipfile
from contextlib import contextmanager, redirect_stderr, redirect_stdout
from ctypes import wintypes
from datetime import datetime, timedelta, timezone
from email.parser import BytesHeaderParser
from email.policy import default as email_policy
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path, PurePosixPath
from typing import Any

from case_docket.application import (
    ApplicationError,
    ContactService,
    GetContactQuery,
    GetContactsContextQuery,
    ListContactsQuery,
    SystemClock,
    UuidProvider,
)
from case_docket.repository import SQLiteRepository, SQLiteUnitOfWorkFactory

from caseflow.api_v1 import (
    API_PREFIX,
    API_VERSION,
    RequestValidationError,
    application_error_envelope,
    error_envelope,
    error_status,
    match_contact_route,
    parse_assign_contact_role,
    parse_create_contact,
    parse_update_contact,
    success_envelope,
)


APP_DIR = Path(__file__).resolve().parent
STATIC_DIR = APP_DIR / "static"


def load_app_manifest() -> dict:
    try:
        payload = json.loads((APP_DIR / "version.json").read_text(encoding="utf-8-sig"))
        if payload.get("product") == "VARTA" and payload.get("version"):
            return payload
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        pass
    return {"product": "VARTA", "version": "0.0.0", "data_schema": 1}


APP_MANIFEST = load_app_manifest()
MAX_UPLOAD_BYTES = 12 * 1024 * 1024 * 1024
MAX_FORM_FIELD_BYTES = 64 * 1024
MAX_MULTIPART_HEADER_BYTES = 64 * 1024
MAX_MULTIPART_PARTS = 20_000
GOOGLE_SCOPE = "https://www.googleapis.com/auth/drive.file"
GOOGLE_AUTH_URL = "https://accounts.google.com/o/oauth2/v2/auth"
GOOGLE_TOKEN_URL = "https://oauth2.googleapis.com/token"
GOOGLE_DRIVE_API = "https://www.googleapis.com/drive/v3"
GOOGLE_DRIVE_UPLOAD = "https://www.googleapis.com/upload/drive/v3"
VARTA_VERSION = str(APP_MANIFEST["version"])
# Compatibility alias for imported CaseFlow tests and release tooling.
CASEFLOW_VERSION = VARTA_VERSION
SUPPORTED_DOCUMENT_EXTENSIONS = {".pdf", ".html", ".htm", ".docx", ".doc", ".rtf", ".txt"}
EVIDENCE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff", ".p7s"}
ANOMALY_WEIGHTS = {"critical": 100, "high": 40, "medium": 15, "low": 5, "info": 0}
DOCUMENT_WORK_STATUSES = {"completed", "in_progress", "waiting", "needs_review"}


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def read_json(path: Path, default):
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return default


def write_json(path: Path, value) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(path)


def sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def safe_segment(value: str, fallback: str = "БЕЗ_НАЗВИ") -> str:
    value = " ".join(str(value or "").strip().split())
    for char in '<>:"/\\|?*':
        value = value.replace(char, "_")
    value = value.strip(" ._")[:120]
    return value or fallback


def run_worker(script: Path, worker_args: list[str], cwd: Path, timeout: int) -> subprocess.CompletedProcess[str]:
    """Run a pipeline worker from source or directly from a frozen EXE bundle."""
    if not getattr(sys, "frozen", False):
        return subprocess.run(
            [sys.executable, str(script), *worker_args],
            cwd=cwd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout,
        )

    if script.stem not in {"caseflow_process", "anomaly_detector"}:
        raise ValueError(f"Невідомий вбудований worker: {script.stem}")
    worker_module = importlib.import_module(f"caseflow.{script.stem}")

    stdout = io.StringIO()
    stderr = io.StringIO()
    return_code = 0
    try:
        with redirect_stdout(stdout), redirect_stderr(stderr):
            return_code = int(worker_module.main(worker_args) or 0)
    except SystemExit as exc:
        return_code = int(exc.code or 0)
    except Exception as exc:  # noqa: BLE001
        stderr.write(f"{type(exc).__name__}: {exc}\n")
        return_code = 1
    return subprocess.CompletedProcess(
        [sys.executable, script.stem, *worker_args],
        return_code,
        stdout.getvalue(),
        stderr.getvalue(),
    )


def unique_path(folder: Path, name: str) -> Path:
    candidate = folder / safe_segment(name, "file")
    if not candidate.exists():
        return candidate
    stem, suffix = candidate.stem, candidate.suffix
    for index in range(2, 10000):
        alternative = folder / f"{stem}__{index}{suffix}"
        if not alternative.exists():
            return alternative
    raise RuntimeError(f"Не вдалося підібрати унікальну назву для {name}")


def safe_upload_path(folder: Path, raw_name: str) -> Path:
    """Preserve a browser-provided relative folder without permitting traversal."""
    pure = PurePosixPath(str(raw_name or "").replace("\\", "/"))
    if pure.is_absolute() or not pure.parts or any(part in {"", ".", ".."} for part in pure.parts):
        raise ValueError(f"Небезпечний відносний шлях файла: {raw_name}")
    parts = [safe_segment(part, "file") for part in pure.parts]
    relative = Path(*parts)
    parent = folder.joinpath(*relative.parts[:-1])
    parent.mkdir(parents=True, exist_ok=True)
    candidate = unique_path(parent, relative.name).resolve()
    resolved_folder = folder.resolve()
    if resolved_folder not in candidate.parents:
        raise ValueError(f"Шлях файла виходить за межі пакета: {raw_name}")
    return candidate


class MultipartStream:
    """Bounded streaming reader for a single multipart/form-data request."""

    def __init__(self, stream, length: int):
        self.stream = stream
        self.remaining = length
        self.buffer = bytearray()

    def _fill(self, minimum: int = 1) -> bool:
        while len(self.buffer) < minimum and self.remaining > 0:
            block = self.stream.read(min(1024 * 1024, self.remaining))
            if not block:
                raise ValueError("Неочікуваний кінець multipart-запиту")
            self.remaining -= len(block)
            self.buffer.extend(block)
        return len(self.buffer) >= minimum

    def readline(self, limit: int) -> bytes:
        while True:
            newline = self.buffer.find(b"\n")
            if newline >= 0:
                size = newline + 1
                if size > limit:
                    raise ValueError("Завеликий заголовок multipart-запиту")
                line = bytes(self.buffer[:size])
                del self.buffer[:size]
                return line
            if len(self.buffer) > limit:
                raise ValueError("Завеликий заголовок multipart-запиту")
            if self.remaining <= 0:
                if not self.buffer:
                    return b""
                line = bytes(self.buffer)
                self.buffer.clear()
                return line
            self._fill(len(self.buffer) + 1)

    def copy_part(self, boundary: bytes, output, max_bytes: int | None = None) -> tuple[bool, int]:
        marker = b"\r\n--" + boundary
        written = 0

        def write(data: bytes) -> None:
            nonlocal written
            if max_bytes is not None and written + len(data) > max_bytes:
                raise ValueError("Завелике текстове поле multipart-запиту")
            output.write(data)
            written += len(data)

        while True:
            marker_at = self.buffer.find(marker)
            if marker_at >= 0:
                required = marker_at + len(marker) + 2
                self._fill(required)
                if len(self.buffer) < required:
                    raise ValueError("Незавершена межа multipart-запиту")
                suffix = bytes(self.buffer[marker_at + len(marker) : required])
                if suffix in {b"\r\n", b"--"}:
                    write(bytes(self.buffer[:marker_at]))
                    del self.buffer[:required]
                    if suffix == b"--":
                        self._fill(2)
                        if self.buffer.startswith(b"\r\n"):
                            del self.buffer[:2]
                        return True, written
                    return False, written
                # A boundary-like byte sequence inside a file is ordinary payload.
                writable = marker_at + 2
            else:
                writable = max(0, len(self.buffer) - len(marker) - 1)

            if writable:
                write(bytes(self.buffer[:writable]))
                del self.buffer[:writable]
                continue
            if self.remaining <= 0:
                raise ValueError("Не знайдено завершальну межу multipart-запиту")
            self._fill(len(self.buffer) + 1)

    def discard_remaining(self) -> None:
        self.buffer.clear()
        while self.remaining > 0:
            block = self.stream.read(min(1024 * 1024, self.remaining))
            if not block:
                raise ValueError("Неочікуваний кінець multipart-запиту")
            self.remaining -= len(block)


def multipart_boundary(content_type: str) -> bytes:
    try:
        raw_header = b"Content-Type: " + content_type.encode("latin-1") + b"\r\n\r\n"
        message = BytesHeaderParser(policy=email_policy).parsebytes(raw_header)
        boundary = message.get_boundary()
        if message.get_content_type() != "multipart/form-data" or not boundary:
            raise ValueError
        encoded = boundary.encode("ascii")
    except (UnicodeEncodeError, ValueError):
        raise ValueError("Некоректний Content-Type multipart-запиту") from None
    if not 1 <= len(encoded) <= 70 or any(byte < 32 or byte > 126 for byte in encoded):
        raise ValueError("Некоректна межа multipart-запиту")
    return encoded


def parse_multipart_form(
    stream,
    content_type: str,
    content_length: int,
    temporary_directory: Path,
) -> tuple[dict[str, str], list[dict]]:
    """Parse form fields in memory and stream file fields into temporary files."""
    boundary = multipart_boundary(content_type)
    reader = MultipartStream(stream, content_length)
    if reader.readline(len(boundary) + 8) != b"--" + boundary + b"\r\n":
        raise ValueError("Некоректний початок multipart-запиту")

    fields: dict[str, str] = {}
    files: list[dict] = []
    final_boundary = False
    part_count = 0
    while not final_boundary:
        part_count += 1
        if part_count > MAX_MULTIPART_PARTS:
            raise ValueError("Забагато частин у multipart-запиті")

        header_lines = bytearray()
        while True:
            line = reader.readline(MAX_MULTIPART_HEADER_BYTES)
            if line == b"\r\n":
                break
            if not line:
                raise ValueError("Незавершені заголовки multipart-запиту")
            header_lines.extend(line)
            if len(header_lines) > MAX_MULTIPART_HEADER_BYTES:
                raise ValueError("Завеликі заголовки multipart-запиту")

        headers = BytesHeaderParser(policy=email_policy).parsebytes(bytes(header_lines))
        if headers.get_content_disposition() != "form-data":
            raise ValueError("Частина multipart-запиту не є form-data")
        field_name = headers.get_param("name", header="content-disposition")
        if not field_name:
            raise ValueError("Частина multipart-запиту не має імені поля")
        filename = headers.get_filename()

        if filename is None:
            payload = io.BytesIO()
            final_boundary, _ = reader.copy_part(
                boundary,
                payload,
                max_bytes=MAX_FORM_FIELD_BYTES,
            )
            charset = headers.get_content_charset("utf-8") or "utf-8"
            try:
                value = payload.getvalue().decode(charset)
            except (LookupError, UnicodeDecodeError):
                raise ValueError(f"Некоректне кодування поля {field_name}") from None
            fields.setdefault(str(field_name), value)
            continue

        temporary_path = temporary_directory / f"part_{len(files) + 1:06}.bin"
        with temporary_path.open("wb") as output:
            final_boundary, size = reader.copy_part(boundary, output)
        files.append(
            {
                "field": str(field_name),
                "filename": str(filename),
                "path": temporary_path,
                "bytes": size,
            }
        )

    reader.discard_remaining()
    return fields, files


class BusyError(RuntimeError):
    pass


class DataBlob(ctypes.Structure):
    _fields_ = [("cbData", wintypes.DWORD), ("pbData", ctypes.POINTER(ctypes.c_byte))]


def dpapi_protect(data: bytes) -> bytes:
    if os.name != "nt":
        raise RuntimeError("DPAPI доступний лише у Windows")
    buffer = ctypes.create_string_buffer(data)
    source = DataBlob(len(data), ctypes.cast(buffer, ctypes.POINTER(ctypes.c_byte)))
    destination = DataBlob()
    if not ctypes.windll.crypt32.CryptProtectData(
        ctypes.byref(source), "VARTA Google OAuth", None, None, None, 0, ctypes.byref(destination)
    ):
        raise ctypes.WinError()
    try:
        return ctypes.string_at(destination.pbData, destination.cbData)
    finally:
        ctypes.windll.kernel32.LocalFree(destination.pbData)


def dpapi_unprotect(data: bytes) -> bytes:
    if os.name != "nt":
        raise RuntimeError("DPAPI доступний лише у Windows")
    buffer = ctypes.create_string_buffer(data)
    source = DataBlob(len(data), ctypes.cast(buffer, ctypes.POINTER(ctypes.c_byte)))
    destination = DataBlob()
    if not ctypes.windll.crypt32.CryptUnprotectData(
        ctypes.byref(source), None, None, None, None, 0, ctypes.byref(destination)
    ):
        raise ctypes.WinError()
    try:
        return ctypes.string_at(destination.pbData, destination.cbData)
    finally:
        ctypes.windll.kernel32.LocalFree(destination.pbData)


class CaseFlowState:
    def __init__(self, root: Path, host: str, port: int):
        self.root = root.resolve()
        self.host = host
        self.port = port
        self.csrf_token = secrets.token_urlsafe(32)
        self.oauth_pending: dict[str, str] = {}
        self.lock = threading.RLock()
        self.job_lock = threading.Lock()
        self.active_job: dict | None = None
        self._repository: SQLiteRepository | None = None
        self._contact_service = ContactService(
            SQLiteUnitOfWorkFactory(self.database_path),
            UuidProvider(),
            SystemClock(),
        )
        self.config_path = self.root / ".caseflow" / "config.json"
        self.token_path = self.root / ".caseflow" / "secrets" / "google_token.dpapi"
        self.google_secret_path = self.root / ".caseflow" / "secrets" / "google_client_secret.dpapi"
        self.drive_index_path = self.root / ".caseflow" / "drive_index.json"
        self.job_path = self.root / ".caseflow" / "active_job.json"
        self.config = read_json(
            self.config_path,
            {
                "case_number": self.root.name,
                "google": {"client_id": ""},
                "ui": {"panel_opacity": 82},
            },
        )
        legacy_secret = self.config.get("google", {}).get("client_secret")
        if legacy_secret:
            try:
                self.save_google_secret(str(legacy_secret))
                self.config["google"].pop("client_secret", None)
                self.save_config()
            except OSError:
                pass

    @property
    def database_path(self) -> Path:
        return self.root / ".caseflow" / "varta.sqlite3"

    @property
    def repository(self) -> SQLiteRepository:
        with self.lock:
            if self._repository is None:
                self.database_path.parent.mkdir(parents=True, exist_ok=True)
                self._repository = SQLiteRepository(self.database_path)
            return self._repository

    @property
    def contact_service(self) -> ContactService:
        return self._contact_service

    def prepare_database(self) -> None:
        """Complete legacy bootstrap before HTTP threads open short-lived UoWs."""
        _ = self.repository

    def close(self) -> None:
        with self.lock:
            if self._repository is not None:
                self._repository.close()
                self._repository = None

    def save_config(self) -> None:
        write_json(self.config_path, self.config)

    def save_google_token(self, token: dict) -> None:
        payload = json.dumps(token, ensure_ascii=False).encode("utf-8")
        protected = dpapi_protect(payload)
        self.token_path.parent.mkdir(parents=True, exist_ok=True)
        self.token_path.write_bytes(base64.b64encode(protected))

    def load_google_token(self) -> dict | None:
        try:
            protected = base64.b64decode(self.token_path.read_bytes())
            return json.loads(dpapi_unprotect(protected).decode("utf-8"))
        except (FileNotFoundError, OSError, ValueError, json.JSONDecodeError):
            return None

    def save_google_secret(self, secret: str) -> None:
        self.google_secret_path.parent.mkdir(parents=True, exist_ok=True)
        self.google_secret_path.write_bytes(base64.b64encode(dpapi_protect(secret.encode("utf-8"))))

    def load_google_secret(self) -> str:
        try:
            protected = base64.b64decode(self.google_secret_path.read_bytes())
            return dpapi_unprotect(protected).decode("utf-8")
        except (FileNotFoundError, OSError, ValueError, UnicodeDecodeError):
            return ""

    def google_connected(self) -> bool:
        token = self.load_google_token()
        return bool(token and (token.get("refresh_token") or token.get("access_token")))

    @contextmanager
    def exclusive_job(self, kind: str):
        if not self.job_lock.acquire(blocking=False):
            active = self.active_job or read_json(self.job_path, {})
            raise BusyError(f"Уже виконується операція: {active.get('kind', 'невідома')}")
        descriptor = None
        metadata = {"kind": kind, "pid": os.getpid(), "started_at": now_iso(), "root": str(self.root)}
        try:
            self.job_path.parent.mkdir(parents=True, exist_ok=True)
            try:
                descriptor = os.open(str(self.job_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            except FileExistsError as exc:
                active = read_json(self.job_path, {})
                raise BusyError(
                    f"Робоча папка заблокована операцією {active.get('kind', 'невідома')} "
                    f"(PID {active.get('pid', '?')}, від {active.get('started_at', '?')})."
                ) from exc
            os.write(descriptor, json.dumps(metadata, ensure_ascii=False, indent=2).encode("utf-8"))
            os.close(descriptor)
            descriptor = None
            self.active_job = metadata
            yield metadata
        finally:
            if descriptor is not None:
                os.close(descriptor)
            self.active_job = None
            try:
                if self.job_path.exists():
                    current = read_json(self.job_path, {})
                    if current.get("pid") == os.getpid():
                        self.job_path.unlink()
            finally:
                self.job_lock.release()

    def access_token(self) -> str:
        token = self.load_google_token()
        if not token:
            raise RuntimeError("Google Drive ще не підключено")
        expires_at = token.get("expires_at")
        if token.get("access_token") and expires_at:
            if datetime.fromisoformat(expires_at) > datetime.now(timezone.utc) + timedelta(seconds=60):
                return token["access_token"]
        refresh_token = token.get("refresh_token")
        if not refresh_token:
            raise RuntimeError("Немає refresh token; підключіть Google Drive повторно")
        google = self.config.get("google", {})
        form = {
            "client_id": google.get("client_id", ""),
            "refresh_token": refresh_token,
            "grant_type": "refresh_token",
        }
        client_secret = self.load_google_secret()
        if client_secret:
            form["client_secret"] = client_secret
        refreshed = post_form(GOOGLE_TOKEN_URL, form)
        refreshed["refresh_token"] = refresh_token
        refreshed["expires_at"] = (
            datetime.now(timezone.utc) + timedelta(seconds=int(refreshed.get("expires_in", 3600)))
        ).isoformat()
        self.save_google_token(refreshed)
        return refreshed["access_token"]


def urlopen_json(request: urllib.request.Request) -> dict:
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            raw = response.read()
            return json.loads(raw.decode("utf-8")) if raw else {}
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Google API {exc.code}: {detail[:1200]}") from exc


def post_form(url: str, fields: dict) -> dict:
    data = urllib.parse.urlencode(fields).encode("utf-8")
    return urlopen_json(
        urllib.request.Request(url, data=data, headers={"Content-Type": "application/x-www-form-urlencoded"})
    )


def google_json(method: str, url: str, token: str, payload: dict | None = None) -> dict:
    data = json.dumps(payload).encode("utf-8") if payload is not None else None
    headers = {"Authorization": f"Bearer {token}"}
    if data is not None:
        headers["Content-Type"] = "application/json; charset=UTF-8"
    return urlopen_json(urllib.request.Request(url, data=data, headers=headers, method=method))


def create_drive_folder(token: str, name: str, parent_id: str | None) -> dict:
    metadata: dict[str, Any] = {
        "name": name,
        "mimeType": "application/vnd.google-apps.folder",
    }
    if parent_id:
        metadata["parents"] = [parent_id]
    return google_json("POST", f"{GOOGLE_DRIVE_API}/files?fields=id,name,webViewLink", token, metadata)


def resumable_upload(token: str, path: Path, name: str, parent_id: str, file_id: str | None = None) -> dict:
    mime = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    metadata = {"name": name, "parents": [parent_id]}
    if file_id:
        metadata.pop("parents", None)
        endpoint = f"{GOOGLE_DRIVE_UPLOAD}/files/{urllib.parse.quote(file_id)}?uploadType=resumable&fields=id,name,webViewLink,modifiedTime"
        method = "PATCH"
    else:
        endpoint = f"{GOOGLE_DRIVE_UPLOAD}/files?uploadType=resumable&fields=id,name,webViewLink,modifiedTime"
        method = "POST"
    request = urllib.request.Request(
        endpoint,
        data=json.dumps(metadata).encode("utf-8"),
        method=method,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json; charset=UTF-8",
            "X-Upload-Content-Type": mime,
            "X-Upload-Content-Length": str(path.stat().st_size),
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            location = response.headers.get("Location")
    except urllib.error.HTTPError as exc:
        raise RuntimeError(exc.read().decode("utf-8", errors="replace")) from exc
    if not location:
        raise RuntimeError("Google Drive не повернув адресу resumable upload")
    data = path.read_bytes()
    upload_request = urllib.request.Request(
        location,
        data=data,
        method="PUT",
        headers={"Authorization": f"Bearer {token}", "Content-Type": mime, "Content-Length": str(len(data))},
    )
    return urlopen_json(upload_request)


def sync_to_drive(state: CaseFlowState, selections: list[str]) -> dict:
    allowed = {"00_INBOX", "01_ОПРАЦЬОВАНО", "03_РЕЄСТР"}
    roots = [name for name in selections if name in allowed and (state.root / name).exists()]
    if not roots:
        raise RuntimeError("Не вибрано жодної папки для синхронізації")
    token = state.access_token()
    index = read_json(state.drive_index_path, {"folders": {}, "files": {}})
    folders = index.setdefault("folders", {})
    files_index = index.setdefault("files", {})
    case_name = f"VARTA__{safe_segment(state.config.get('case_number', state.root.name), state.root.name)}"
    if "__ROOT__" not in folders:
        folders["__ROOT__"] = create_drive_folder(token, case_name, None)["id"]

    def folder_id(relative: Path) -> str:
        key = relative.as_posix()
        if key in folders:
            return folders[key]
        parent = folder_id(relative.parent) if relative.parent != Path(".") else folders["__ROOT__"]
        folders[key] = create_drive_folder(token, relative.name, parent)["id"]
        write_json(state.drive_index_path, index)
        return folders[key]

    uploaded = skipped = 0
    links = []
    for root_name in roots:
        source_root = state.root / root_name
        folder_id(Path(root_name))
        for path in sorted(source_root.rglob("*"), key=lambda item: str(item).casefold()):
            if not path.is_file() or path.name.startswith("~$"):
                continue
            relative = path.relative_to(state.root)
            digest = sha256_path(path)
            previous = files_index.get(relative.as_posix(), {})
            if previous.get("sha256") == digest:
                skipped += 1
                continue
            parent = folder_id(relative.parent)
            result = resumable_upload(token, path, path.name, parent, previous.get("id"))
            files_index[relative.as_posix()] = {
                "id": result["id"],
                "sha256": digest,
                "webViewLink": result.get("webViewLink"),
                "synced_at": now_iso(),
            }
            links.append(result.get("webViewLink"))
            uploaded += 1
            write_json(state.drive_index_path, index)
    return {"uploaded": uploaded, "skipped": skipped, "rootFolderId": folders["__ROOT__"], "links": [x for x in links if x]}


def latest_register(root: Path) -> Path | None:
    pointer = root / "03_РЕЄСТР" / "ОСТАННІЙ_РЕЄСТР.txt"
    if pointer.exists():
        try:
            pointed = Path(pointer.read_text(encoding="utf-8-sig").strip())
            if not pointed.is_absolute():
                pointed = root / pointed
            if pointed.exists() and pointed.suffix.lower() == ".xlsx":
                return pointed.resolve()
        except OSError:
            pass
    candidates = [
        path
        for directory in (root / "03_РЕЄСТР" / "exports", root / "03_РЕЄСТР")
        if directory.exists()
        for path in directory.glob("*.xlsx")
        if not path.name.startswith("~$")
    ]
    return max(candidates, key=lambda path: path.stat().st_mtime).resolve() if candidates else None


def document_work_status(completeness: str, next_action: str) -> str:
    """Map explicit register fields to a small, UI-friendly workflow status."""
    normalized = str(completeness or "").strip().upper()
    action = str(next_action or "").strip()
    if any(marker in normalized for marker in ("ПЕРЕВІР", "НЕПОВН", "ПОШКОДЖ")):
        return "needs_review"
    if "КОМПЛЕКТ" in normalized:
        return "completed"
    if action and any(marker in action.casefold() for marker in ("очіку", "отримати", "контрол", "дочек")):
        return "waiting"
    if action:
        return "in_progress"
    return "in_progress"


def find_7zip(configured: str = "") -> Path | None:
    """Locate a 7-Zip executable used for read-only RAR inventory and extraction."""
    candidates = [
        configured,
        os.environ.get("VARTA_7Z", ""),
        os.environ.get("CASEFLOW_7Z", ""),  # legacy compatibility
    ]
    for variable in ("ProgramFiles", "ProgramFiles(x86)"):
        base = os.environ.get(variable)
        if base:
            candidates.append(str(Path(base) / "7-Zip" / "7z.exe"))
    candidates.extend(
        [
            r"C:\Program Files\3uToolsV3\files\patchtools\7z-64\7z.exe",
            r"C:\Program Files\3uToolsV3\files\patchtools\7z-32\7z.exe",
            r"C:\Program Files\Lenovo\Lenovo Bootable Generator\7z.exe",
        ]
    )
    command = shutil.which("7z") or shutil.which("7z.exe")
    if command:
        candidates.append(command)
    for candidate in candidates:
        if candidate:
            path = Path(candidate).expanduser()
            if path.is_file():
                return path.resolve()
    return None


def archive_component(member_name: str) -> tuple[str, str]:
    member = PurePosixPath(member_name.replace("\\", "/"))
    section = member.parts[0].casefold() if member.parts else ""
    if member_name.casefold().endswith(".p7s"):
        return "Підпис КЕП", section
    if section == "main":
        return "Основний документ", section
    if section == "atch":
        return "Додаток", section
    if section == "tech":
        return "Технічний документ", section
    return "Файл архіву", section


def seven_zip_tree_entries(path: Path, seven_zip: Path) -> list[dict]:
    """Read a RAR/ZIP directory with 7-Zip without extracting archive contents."""
    result = subprocess.run(
        [str(seven_zip), "l", "-slt", "-ba", "-sccUTF-8", str(path)],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=120,
        check=False,
    )
    if result.returncode not in {0, 1}:
        return []
    entries: list[dict] = []
    record: dict[str, str] = {}

    def append_record() -> None:
        member_name = record.get("Path", "")
        if not member_name or record.get("Folder") == "+":
            return
        member = PurePosixPath(member_name.replace("\\", "/"))
        if member.is_absolute() or not member.parts or any(part in {"", ".", ".."} for part in member.parts):
            return
        component, section = archive_component(member_name)
        try:
            size = int(record.get("Size", "0") or 0)
        except ValueError:
            size = 0
        entries.append(
            {
                "component": component,
                "name": member.name,
                "path": f"{path.name} › {member.as_posix()}",
                "bytes": size,
                "integrity": f"У {path.suffix.lstrip('.').upper()}",
                "section": section,
            }
        )

    for line in result.stdout.splitlines() + [""]:
        if not line.strip():
            if record:
                append_record()
                record = {}
            continue
        key, separator, value = line.partition(" = ")
        if separator:
            record[key.strip()] = value.strip()
    return entries[:2000]


def archive_tree_entries(path: Path, seven_zip: Path | None = None) -> list[dict]:
    """Inventory ZIP/RAR for the UI without extracting or trusting member paths."""
    if path.suffix.casefold() == ".rar":
        return seven_zip_tree_entries(path, seven_zip) if seven_zip else []
    entries = []
    try:
        with zipfile.ZipFile(path) as archive:
            for item in archive.infolist()[:2000]:
                if item.is_dir():
                    continue
                member_name = item.filename
                if not item.flag_bits & 0x800:
                    try:
                        candidate = item.filename.encode("cp437").decode("windows-1251")
                        if sum("А" <= char <= "я" or char in "ІіЇїЄєҐґ" for char in candidate) > sum("А" <= char <= "я" or char in "ІіЇїЄєҐґ" for char in item.filename):
                            member_name = candidate
                    except (UnicodeEncodeError, UnicodeDecodeError):
                        pass
                member = PurePosixPath(member_name.replace("\\", "/"))
                if member.is_absolute() or any(part in {"", ".", ".."} for part in member.parts):
                    continue
                component, section = archive_component(member_name)
                entries.append({
                    "component": component,
                    "name": member.name,
                    "path": f"{path.name} › {member.as_posix()}",
                    "bytes": item.file_size,
                    "integrity": "У ZIP",
                    "section": section,
                })
    except (OSError, zipfile.BadZipFile):
        return []
    return entries


def build_document_tree(root: Path) -> dict:
    """Build a read-only proceeding/document/file tree from the latest register and INBOX."""
    register = latest_register(root)
    config = read_json(root / ".caseflow" / "config.json", {})
    seven_zip = find_7zip(str(config.get("seven_zip_path", "")))
    status_overrides = read_json(root / ".caseflow" / "document_status.json", {})
    groups: dict[str, dict] = {}

    def group_for(number: str, raw_status: str = "") -> dict:
        key = str(number or "Невизначене провадження").strip()
        return groups.setdefault(
            key,
            {"number": key, "rawStatus": raw_status, "documents": []},
        )

    if register:
        from openpyxl import load_workbook

        workbook = load_workbook(register, read_only=True, data_only=True)
        try:
            def sheet_rows(name: str) -> list[dict]:
                if name not in workbook.sheetnames:
                    return []
                sheet = workbook[name]
                values = sheet.iter_rows(values_only=True)
                headers = [str(value or "").strip() for value in next(values)]
                return [dict(zip(headers, row)) for row in values if any(value is not None for value in row)]

            files_by_doc: dict[str, list[dict]] = {}
            for row in sheet_rows("Файли"):
                doc_id = str(row.get("ID документа") or "").strip()
                if not doc_id:
                    continue
                files_by_doc.setdefault(doc_id, []).append(
                    {
                        "component": row.get("Компонент") or "Інше",
                        "name": row.get("Оригінальна назва") or row.get("Нормалізована назва") or "Файл",
                        "path": row.get("Відносний шлях") or "",
                        "bytes": row.get("Розмір, байт") or 0,
                        "integrity": row.get("Цілісність") or "",
                    }
                )

            proceeding_status = {
                str(row.get("Номер провадження") or "").strip(): str(row.get("Статус") or "").strip()
                for row in sheet_rows("Провадження")
                if row.get("Номер провадження")
            }
            for row in sheet_rows("Документи"):
                doc_id = str(row.get("ID документа") or "").strip()
                if not doc_id:
                    continue
                proceeding = str(row.get("Провадження") or "Невизначене провадження").strip()
                completeness = str(row.get("Статус комплектності") or "").strip()
                next_action = str(row.get("Наступна дія") or "").strip()
                date_value = row.get("Дата документа") or row.get("Дата надходження/подання")
                date_text = date_value.date().isoformat() if isinstance(date_value, datetime) else str(date_value or "")
                group_for(proceeding, proceeding_status.get(proceeding, ""))["documents"].append(
                    {
                        "id": doc_id,
                        "name": row.get("Назва документа") or doc_id,
                        "type": row.get("Тип документа") or "Інше",
                        "flow": row.get("Потік") or "",
                        "date": date_text,
                        "summary": row.get("Опис / пов’язана подія") or "",
                        "completeness": completeness,
                        "nextAction": next_action,
                        "folder": row.get("Відносна папка") or "",
                        "status": document_work_status(completeness, next_action),
                        "files": files_by_doc.get(doc_id, []),
                    }
                )
        finally:
            workbook.close()

    inbox = root / "00_INBOX"
    if inbox.exists():
        for manifest_path in inbox.rglob("caseflow_upload.json"):
            manifest = read_json(manifest_path, {})
            proceeding = str(manifest.get("proceeding_folder") or manifest_path.relative_to(inbox).parts[0])
            files = manifest.get("files", [])
            uploaded_at = str(manifest.get("uploaded_at") or "")
            packet_id = f"INBOX_{manifest_path.parent.name.split('__', 1)[0]}"
            packet_files = []
            for item in files:
                suffix = Path(str(item.get("name", ""))).suffix.casefold()
                component = "RAR-архів" if suffix == ".rar" else "ZIP-архів" if suffix == ".zip" else "Вхідний файл"
                packet_files.append({"component": component, "name": item.get("name", "Файл"), "path": item.get("path", ""), "bytes": item.get("bytes", 0), "integrity": ""})
                local_path = root / str(item.get("path", ""))
                if local_path.suffix.lower() in {".zip", ".rar"} and local_path.is_file():
                    packet_files.extend(archive_tree_entries(local_path, seven_zip))
            single_suffix = Path(str(files[0].get("name", ""))).suffix.casefold() if len(files) == 1 else ""
            group_for(proceeding, "В роботі")["documents"].append(
                {
                    "id": packet_id,
                    "name": files[0].get("name", "Новий пакет") if len(files) == 1 else f"Новий пакет: {len(files)} файлів",
                    "type": f"{single_suffix.lstrip('.').upper()}-архів" if single_suffix in {".zip", ".rar"} else "Вхідний пакет",
                    "flow": manifest.get("flow", ""),
                    "date": uploaded_at[:10],
                    "summary": "Завантажено до 00_INBOX; очікує запуску обробки.",
                    "completeness": "У черзі",
                    "nextAction": "Запустити обробку INBOX.",
                    "folder": str(manifest_path.parent.relative_to(root)),
                    "status": "in_progress",
                    "files": packet_files,
                }
            )

    counts = {status: 0 for status in ("completed", "in_progress", "waiting", "needs_review")}
    proceedings = []
    for group in groups.values():
        for document in group["documents"]:
            override = status_overrides.get(document.get("id", ""), {})
            if override.get("status") in DOCUMENT_WORK_STATUSES:
                document["status"] = override["status"]
                document["statusManual"] = True
                document["statusNote"] = str(override.get("note", ""))
                document["statusUpdatedAt"] = override.get("updated_at")
        group["documents"].sort(key=lambda item: (item.get("date", ""), item.get("id", "")), reverse=True)
        group["counts"] = {status: sum(item["status"] == status for item in group["documents"]) for status in counts}
        for status, value in group["counts"].items():
            counts[status] += value
        proceedings.append(group)
    proceedings.sort(key=lambda item: (item["counts"]["in_progress"] + item["counts"]["needs_review"] > 0, item["number"]), reverse=True)
    return {
        "ok": True,
        "register": str(register.relative_to(root)) if register and root in register.parents else str(register or ""),
        "generatedAt": now_iso(),
        "archiveSupport": {"zip": True, "rar": bool(seven_zip), "sevenZip": str(seven_zip or "")},
        "counts": {**counts, "all": sum(counts.values())},
        "proceedings": proceedings,
    }


def build_preflight(state: CaseFlowState) -> dict:
    """Return a fact-based readiness score before a mutating pipeline run."""
    root = state.root
    checks: list[dict] = []

    def add(code: str, label: str, weight: int, status: str, detail: str, earned: int | None = None) -> None:
        checks.append(
            {
                "code": code,
                "label": label,
                "weight": weight,
                "earned": weight if status == "passed" else (earned or 0),
                "status": status,
                "detail": detail,
            }
        )

    required = [
        "00_INBOX",
        "01_ОПРАЦЬОВАНО",
        "02_РОЗПАКОВАНО",
        "03_РЕЄСТР",
        "99_ПОТРЕБУЄ_ПЕРЕВІРКИ",
        ".caseflow",
    ]
    missing = [name for name in required if not (root / name).is_dir()]
    writable = False
    if not missing:
        probe = root / ".caseflow" / f".preflight_{secrets.token_hex(5)}"
        try:
            probe.write_text("ok", encoding="ascii")
            probe.unlink()
            writable = True
        except OSError:
            writable = False
    add(
        "ROOT_WRITABLE",
        "Структура справи та право запису",
        15,
        "passed" if writable else "blocker",
        "Усі обов’язкові папки доступні для запису." if writable else f"Немає папок або запису: {', '.join(missing) or 'корінь недоступний'}.",
    )

    missing_modules = [name for name in ("openpyxl", "pypdf") if importlib.util.find_spec(name) is None]
    scripts = [state.root / "scripts" / "caseflow_process.py", state.root / "scripts" / "anomaly_detector.py"]
    missing_scripts = [] if getattr(sys, "frozen", False) else [str(path.relative_to(root)) for path in scripts if not path.exists()]
    runtime_status = "passed"
    runtime_detail = f"Python {sys.version_info.major}.{sys.version_info.minor}; openpyxl і pypdf доступні; скрипти на місці."
    if "openpyxl" in missing_modules or missing_scripts:
        runtime_status = "blocker"
        runtime_detail = f"Бракує: {', '.join(missing_modules + missing_scripts)}."
    elif missing_modules:
        runtime_status = "warning"
        runtime_detail = "Немає pypdf: PDF оброблятимуться без надійного читання тексту."
    add("RUNTIME", "Python, залежності та скрипти", 15, runtime_status, runtime_detail, 10 if runtime_status == "warning" else 0)

    register = latest_register(root)
    register_valid = False
    if register:
        try:
            with zipfile.ZipFile(register) as archive:
                register_valid = "xl/workbook.xml" in archive.namelist()
        except (OSError, zipfile.BadZipFile):
            register_valid = False
    if register_valid:
        add(
            "REGISTER_BASELINE",
            "Останній повний Реєстр",
            20,
            "passed",
            f"Базова книга: {register.relative_to(root) if register and root in register.parents else register}.",
        )
    elif register is None:
        add(
            "REGISTER_BASELINE",
            "Останній повний Реєстр",
            20,
            "warning",
            "Це чиста інсталяція: перший запуск конвеєра створить новий Реєстр .xlsx.",
            15,
        )
    else:
        add(
            "REGISTER_BASELINE",
            "Останній повний Реєстр",
            20,
            "blocker",
            "Знайдена книга .xlsx пошкоджена або невалідна.",
        )

    inbox = root / "00_INBOX"
    index_path = root / ".caseflow" / "index.json"
    index_valid = True
    index: dict[str, Any] = {"hashes": {}}
    if index_path.exists():
        try:
            loaded_index = json.loads(index_path.read_text(encoding="utf-8-sig"))
            if isinstance(loaded_index, dict) and isinstance(loaded_index.get("hashes", {}), dict):
                index = loaded_index
            else:
                index_valid = False
        except (OSError, json.JSONDecodeError):
            index_valid = False
    indexed_sources = {
        clean_path
        for entry in index.get("hashes", {}).values()
        if isinstance(entry, dict)
        for clean_path in [str(entry.get("source", "")).replace("\\", "/")]
        if clean_path
    }
    incoming = [path for path in inbox.rglob("*") if path.is_file() and path.name != "caseflow_upload.json"] if inbox.exists() else []
    new_documents = []
    archive_packages = []
    for path in incoming:
        relative = path.relative_to(root).as_posix()
        if path.suffix.lower() in SUPPORTED_DOCUMENT_EXTENSIONS and relative not in indexed_sources:
            new_documents.append(path)
        elif path.suffix.lower() in {".zip", ".rar"}:
            relative_inbox = path.relative_to(inbox)
            try:
                digest = sha256_path(path)[:12]
                unpacked = root / "02_РОЗПАКОВАНО" / relative_inbox.parent / f"{safe_segment(path.stem)}__{digest}"
                if not unpacked.exists():
                    archive_packages.append(path)
            except OSError:
                archive_packages.append(path)
    zip_packages = [path for path in archive_packages if path.suffix.casefold() == ".zip"]
    rar_packages = [path for path in archive_packages if path.suffix.casefold() == ".rar"]
    new_count = len(new_documents) + len(archive_packages)
    if not index_valid:
        new_status = "blocker"
        new_detail = "Пошкоджено .caseflow/index.json; потрібне контрольоване відновлення."
    elif new_count:
        new_status = "passed"
        new_detail = f"Нових основних документів: {len(new_documents)}; ZIP: {len(zip_packages)}; RAR: {len(rar_packages)}."
    else:
        new_status = "blocker"
        new_detail = "Нових підтримуваних документів, ZIP або RAR після попередньої обробки не знайдено."
    add("NEW_INPUT", "Нові матеріали після дедуплікації", 10, new_status, new_detail)

    seven_zip = find_7zip(str(state.config.get("seven_zip_path", "")))
    rar_ready = not rar_packages or bool(seven_zip)
    add(
        "RAR_SUPPORT",
        "Підтримка RAR через 7-Zip",
        0,
        "passed" if rar_ready else "blocker",
        f"7-Zip: {seven_zip}." if seven_zip else ("RAR у новому пакеті немає." if not rar_packages else "Знайдено RAR, але 7-Zip недоступний."),
    )

    metadata_issues = []
    for path in [*new_documents, *archive_packages]:
        relative_parts = path.relative_to(inbox).parts
        if len(relative_parts) < 3:
            metadata_issues.append(str(path.relative_to(root)))
            continue
        folded = {part.casefold() for part in relative_parts}
        if not ({"01_від_суду".casefold(), "02_мої_документи".casefold()} & folded):
            metadata_issues.append(str(path.relative_to(root)))
    metadata_status = "passed" if new_count and not metadata_issues else ("warning" if new_count else "blocker")
    metadata_detail = (
        "Провадження/папка, потік і канал мають достатній контекст."
        if metadata_status == "passed"
        else f"Неповний контекст у {len(metadata_issues)} шляхах; перевірте провадження, потік і канал."
        if new_count
        else "Немає нового пакета для перевірки контексту."
    )
    add("PACKAGE_CONTEXT", "Провадження, потік і канал", 10, metadata_status, metadata_detail, 5 if metadata_status == "warning" else 0)

    total_bytes = sum(path.stat().st_size for path in incoming if path.exists())
    zip_unpacked = 0
    for archive_path in zip_packages[:100]:
        try:
            with zipfile.ZipFile(archive_path) as archive:
                zip_unpacked += sum(item.file_size for item in archive.infolist()[:10000] if not item.is_dir())
        except (OSError, zipfile.BadZipFile):
            pass
    for archive_path in rar_packages[:100]:
        if not seven_zip:
            break
        zip_unpacked += sum(item.get("bytes", 0) for item in seven_zip_tree_entries(archive_path, seven_zip))
    required_bytes = max(250 * 1024 * 1024, total_bytes * 2 + zip_unpacked)
    free_bytes = shutil.disk_usage(root).free
    disk_ok = free_bytes >= required_bytes
    add(
        "DISK_SPACE",
        "Вільне місце для staging, копій і export",
        15,
        "passed" if disk_ok else "blocker",
        f"Вільно {free_bytes / 1073741824:.1f} ГБ; оцінена потреба {required_bytes / 1073741824:.1f} ГБ.",
    )

    lock_metadata = read_json(state.job_path, {}) if state.job_path.exists() else None
    lock_free = not state.job_lock.locked() and not lock_metadata
    add(
        "NO_ACTIVE_JOB",
        "Відсутність паралельного запуску",
        10,
        "passed" if lock_free else "blocker",
        "Активних операцій немає." if lock_free else f"Активна/залишкова операція: {(lock_metadata or state.active_job or {}).get('kind', 'невідома')}.",
    )

    evidence_names = [path.name.casefold() for path in incoming]
    evidence_count = sum(
        path.suffix.lower() in EVIDENCE_EXTENSIONS
        or "картк" in name
        or "протокол" in name
        for path, name in zip(incoming, evidence_names)
    )
    evidence_status = "passed" if evidence_count else "warning"
    add(
        "EVIDENCE_COVERAGE",
        "Локальні картки, КЕП або скриншоти",
        5,
        evidence_status,
        f"Локальних доказових компонентів: {evidence_count}." if evidence_count else "Доказових компонентів не знайдено; частина правил буде not verifiable.",
        2 if evidence_status == "warning" else 0,
    )

    percent = max(0, min(100, sum(item["earned"] for item in checks)))
    blockers = [item for item in checks if item["status"] == "blocker"]
    return {
        "checked_at": now_iso(),
        "percent": percent,
        "can_start": not blockers and percent >= 80,
        "blockers": len(blockers),
        "warnings": sum(item["status"] == "warning" for item in checks),
        "new_documents": len(new_documents),
        "new_zip_packages": len(zip_packages),
        "new_rar_packages": len(rar_packages),
        "register": str(register.relative_to(root)) if register and root in register.parents else str(register) if register else None,
        "checks": checks,
    }


def anomaly_summary(findings: list[dict[str, Any]]) -> dict[str, int]:
    open_items = [item for item in findings if item.get("status", "open") == "open"]
    severities = [
        severity if isinstance(severity := item.get("severity"), str) else ""
        for item in open_items
    ]
    counts = {level: severities.count(level) for level in ANOMALY_WEIGHTS}
    return {
        "total": len(findings),
        "open": len(open_items),
        **counts,
        "risk_score": min(100, sum(ANOMALY_WEIGHTS.get(severity, 0) for severity in severities)),
    }


class Handler(BaseHTTPRequestHandler):
    server_version = f"VARTA/{VARTA_VERSION}"

    @property
    def state(self) -> CaseFlowState:
        return self.server.state  # type: ignore[attr-defined]

    def log_message(self, format: str, *args) -> None:
        log_path = self.state.root / ".caseflow" / "logs" / "server.log"
        log_path.parent.mkdir(parents=True, exist_ok=True)
        with log_path.open("a", encoding="utf-8") as stream:
            stream.write(f"{now_iso()} {self.address_string()} {format % args}\n")

    def send_json(self, status: int, payload: dict) -> None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.end_headers()
        self.wfile.write(data)

    def send_error_json(self, status: int, error: Exception | str) -> None:
        self.send_json(status, {"ok": False, "error": str(error)})

    def send_contract_success(
        self,
        status: int,
        payload: dict[str, object],
        *,
        versioned: bool,
    ) -> None:
        body = success_envelope(payload) if versioned else {"ok": True, **payload}
        self.send_json(status, body)

    def send_application_error(
        self,
        error: ApplicationError | RequestValidationError,
        *,
        versioned: bool,
    ) -> None:
        status = error_status(error)
        if versioned:
            self.send_json(status, application_error_envelope(error))
        else:
            self.send_error_json(status, error.message)

    def send_versioned_error(
        self,
        status: int,
        code: str,
        message: str,
        details: dict[str, object] | None = None,
    ) -> None:
        self.send_json(status, error_envelope(code, message, details))

    def _read_json_value(self) -> object:
        size = int(self.headers.get("Content-Length", "0"))
        if size > 2 * 1024 * 1024:
            raise ValueError("JSON-запит завеликий")
        raw = self.rfile.read(size)
        return json.loads(raw.decode("utf-8")) if raw else {}

    def read_json_body(self) -> dict:
        payload = self._read_json_value()
        if not isinstance(payload, dict):
            raise ValueError("JSON-запит має бути object")
        return payload

    def read_command_body(self) -> object:
        try:
            return self._read_json_value()
        except (json.JSONDecodeError, UnicodeDecodeError, ValueError) as exc:
            raise RequestValidationError(
                "Некоректний JSON body",
                {"field": "body"},
            ) from exc

    def require_csrf(self) -> None:
        if self.headers.get("X-Caseflow-Token") != self.state.csrf_token:
            raise PermissionError("Недійсний локальний токен запиту")

    def do_GET(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        versioned = self._is_versioned_path(parsed.path)
        contact_route = match_contact_route(parsed.path)
        try:
            if parsed.path == "/api/status":
                self.handle_status()
                return
            if parsed.path == f"{API_PREFIX}/status":
                self.handle_api_v1_status()
                return
            if parsed.path == "/api/preflight":
                self.handle_preflight()
                return
            if parsed.path == "/api/documents/tree":
                self.handle_document_tree()
                return
            if contact_route is not None:
                if contact_route.action == "collection":
                    self.handle_contacts(
                        urllib.parse.parse_qs(parsed.query),
                        versioned=contact_route.versioned,
                    )
                    return
                if contact_route.action == "context":
                    self.handle_contacts_context(versioned=contact_route.versioned)
                    return
                if contact_route.action == "detail" and contact_route.contact_id is not None:
                    self.handle_contact(
                        contact_route.contact_id,
                        versioned=contact_route.versioned,
                    )
                    return
                self._send_route_not_found(versioned=contact_route.versioned)
                return
            if parsed.path == "/api/anomalies/latest":
                self.handle_anomalies_latest()
                return
            if parsed.path == "/oauth/google/callback":
                self.handle_google_callback(urllib.parse.parse_qs(parsed.query))
                return
            if versioned:
                self._send_route_not_found(versioned=True)
                return
            self.serve_static(parsed.path)
        except (ApplicationError, RequestValidationError) as exc:
            self.send_application_error(exc, versioned=versioned)
        except Exception as exc:  # noqa: BLE001
            self._send_unexpected_error(exc, versioned=versioned)

    def do_POST(self) -> None:
        route = urllib.parse.urlparse(self.path).path
        versioned = self._is_versioned_path(route)
        contact_route = match_contact_route(route)
        try:
            self.require_csrf()
            if contact_route is not None and contact_route.action == "collection":
                self.handle_contact_create(versioned=contact_route.versioned)
            elif (
                contact_route is not None
                and contact_route.action == "roles"
                and contact_route.contact_id is not None
            ):
                self.handle_contact_role(
                    contact_route.contact_id,
                    versioned=contact_route.versioned,
                )
            elif contact_route is not None:
                self._send_route_not_found(versioned=contact_route.versioned)
            elif route == "/api/upload":
                self.handle_upload()
            elif route == "/api/documents/status":
                self.handle_document_status()
            elif route == "/api/process":
                self.handle_process()
            elif route == "/api/anomalies/run":
                self.handle_anomalies_run()
            elif route == "/api/anomalies/status":
                self.handle_anomaly_status()
            elif route == "/api/settings":
                self.handle_settings()
            elif route == "/api/google/config":
                self.handle_google_config()
            elif route == "/api/google/login":
                self.handle_google_login()
            elif route == "/api/google/disconnect":
                self.handle_google_disconnect()
            elif route == "/api/google/sync":
                self.handle_google_sync()
            elif versioned:
                self._send_route_not_found(versioned=True)
            else:
                self.send_error_json(404, "Маршрут не знайдено")
        except (ApplicationError, RequestValidationError) as exc:
            self.send_application_error(exc, versioned=versioned)
        except BusyError as exc:
            if versioned:
                self.send_versioned_error(409, "busy", "Інша локальна операція ще виконується")
            else:
                self.send_error_json(409, exc)
        except PermissionError as exc:
            if versioned:
                self.send_versioned_error(403, "forbidden", "Недійсний локальний токен запиту")
            else:
                self.send_error_json(403, exc)
        except Exception as exc:  # noqa: BLE001
            self._send_unexpected_error(exc, versioned=versioned)

    def do_PATCH(self) -> None:
        route = urllib.parse.urlparse(self.path).path
        versioned = self._is_versioned_path(route)
        contact_route = match_contact_route(route)
        try:
            self.require_csrf()
            if (
                contact_route is not None
                and contact_route.action == "detail"
                and contact_route.contact_id is not None
            ):
                self.handle_contact_update(
                    contact_route.contact_id,
                    versioned=contact_route.versioned,
                )
            elif contact_route is not None:
                self._send_route_not_found(versioned=contact_route.versioned)
            elif versioned:
                self._send_route_not_found(versioned=True)
            else:
                self.send_error_json(404, "Маршрут не знайдено")
        except (ApplicationError, RequestValidationError) as exc:
            self.send_application_error(exc, versioned=versioned)
        except PermissionError as exc:
            if versioned:
                self.send_versioned_error(403, "forbidden", "Недійсний локальний токен запиту")
            else:
                self.send_error_json(403, exc)
        except Exception as exc:  # noqa: BLE001
            self._send_unexpected_error(exc, versioned=versioned)

    @staticmethod
    def _is_versioned_path(path: str) -> bool:
        return path == API_PREFIX or path.startswith(f"{API_PREFIX}/")

    def _send_route_not_found(self, *, versioned: bool) -> None:
        if versioned:
            self.send_versioned_error(404, "route_not_found", "Маршрут API не знайдено")
        else:
            self.send_error_json(404, "Маршрут не знайдено")

    def _send_unexpected_error(self, error: Exception, *, versioned: bool) -> None:
        if versioned:
            self.send_versioned_error(500, "internal_error", "Внутрішня помилка VARTA")
        else:
            self.send_error_json(400, error)

    def serve_static(self, request_path: str) -> None:
        relative = "index.html" if request_path in {"", "/"} else request_path.lstrip("/")
        candidate = (STATIC_DIR / relative).resolve()
        if STATIC_DIR.resolve() not in candidate.parents and candidate != STATIC_DIR.resolve():
            self.send_error(404)
            return
        if not candidate.exists() or not candidate.is_file():
            self.send_error(404)
            return
        data = candidate.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", mimetypes.guess_type(candidate.name)[0] or "application/octet-stream")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-cache")
        if candidate.name == "legal-case-map.html":
            policy = "default-src 'none'; style-src 'unsafe-inline'; script-src 'unsafe-inline'; img-src 'self' data:; connect-src 'self'; form-action 'self'; base-uri 'none'; object-src 'none'"
        else:
            policy = "default-src 'self'; style-src 'self' 'unsafe-inline'; script-src 'self'; img-src 'self' data:; connect-src 'self'; form-action 'self' https://accounts.google.com"
        self.send_header("Content-Security-Policy", policy)
        self.send_header("X-Content-Type-Options", "nosniff")
        self.end_headers()
        self.wfile.write(data)

    def handle_status(self) -> None:
        inbox = self.state.root / "00_INBOX"
        files = [p for p in inbox.rglob("*") if p.is_file()] if inbox.exists() else []
        recent = sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)[:12]
        runs_dir = self.state.root / "tmp" / "caseflow_runs"
        runs = sorted(runs_dir.glob("*.json"), reverse=True) if runs_dir.exists() else []
        last_run = read_json(runs[0], None) if runs else None
        anomaly_report = read_json(self.state.root / "tmp" / "caseflow_anomalies" / "latest.json", None)
        google = self.state.config.get("google", {})
        with self.state.lock:
            database = self.state.repository.airtable_catalog_counts()
        database["contacts"] = len(self.state.contact_service.list(ListContactsQuery()))
        self.send_json(
            200,
            {
                "ok": True,
                "server": {"product": "VARTA", "version": VARTA_VERSION, "root": str(self.state.root)},
                "csrfToken": self.state.csrf_token,
                "root": str(self.state.root),
                "caseNumber": self.state.config.get("case_number", self.state.root.name),
                "inbox": {"files": len(files), "bytes": sum(p.stat().st_size for p in files)},
                "recent": [
                    {"name": p.name, "path": str(p.relative_to(self.state.root)), "bytes": p.stat().st_size, "modified": datetime.fromtimestamp(p.stat().st_mtime).isoformat(timespec="seconds")}
                    for p in recent
                ],
                "lastRun": last_run,
                "activeJob": self.state.active_job or (read_json(self.state.job_path, None) if self.state.job_path.exists() else None),
                "anomalies": {
                    "available": bool(anomaly_report),
                    "generatedAt": anomaly_report.get("generated_at") if anomaly_report else None,
                    "register": anomaly_report.get("register") if anomaly_report else None,
                    "summary": anomaly_report.get("summary") if anomaly_report else None,
                },
                "google": {"configured": bool(google.get("client_id")), "connected": self.state.google_connected(), "clientId": google.get("client_id", "")},
                "archiveSupport": {
                    "zip": True,
                    "rar": bool(find_7zip(str(self.state.config.get("seven_zip_path", "")))),
                    "sevenZip": str(find_7zip(str(self.state.config.get("seven_zip_path", ""))) or ""),
                },
                "database": database,
                "ui": self.state.config.get("ui", {}),
            },
        )

    def handle_preflight(self) -> None:
        self.send_json(200, {"ok": True, **build_preflight(self.state)})

    def handle_document_tree(self) -> None:
        self.send_json(200, build_document_tree(self.state.root))

    def handle_contacts(self, query: dict[str, list[str]], *, versioned: bool) -> None:
        search = str(query.get("q", [""])[0])
        contacts = self.state.contact_service.list(ListContactsQuery(search or None))
        self.send_contract_success(
            200,
            {
                "contacts": [contact.to_dict() for contact in contacts],
                "count": len(contacts),
            },
            versioned=versioned,
        )

    def handle_contact(self, contact_id: str, *, versioned: bool) -> None:
        contact = self.state.contact_service.get(
            GetContactQuery(urllib.parse.unquote(contact_id))
        )
        self.send_contract_success(
            200,
            {"contact": contact.to_dict()},
            versioned=versioned,
        )

    def handle_contacts_context(self, *, versioned: bool) -> None:
        context = self.state.contact_service.context(GetContactsContextQuery())
        self.send_contract_success(200, context.to_dict(), versioned=versioned)

    def handle_contact_create(self, *, versioned: bool) -> None:
        command = parse_create_contact(self.read_command_body())
        contact = self.state.contact_service.create(command)
        self.send_contract_success(
            201,
            {"contact": contact.to_dict()},
            versioned=versioned,
        )

    def handle_contact_update(self, contact_id: str, *, versioned: bool) -> None:
        command = parse_update_contact(
            urllib.parse.unquote(contact_id),
            self.read_command_body(),
        )
        contact = self.state.contact_service.update(command)
        self.send_contract_success(
            200,
            {"contact": contact.to_dict()},
            versioned=versioned,
        )

    def handle_contact_role(self, contact_id: str, *, versioned: bool) -> None:
        command = parse_assign_contact_role(
            urllib.parse.unquote(contact_id),
            self.read_command_body(),
        )
        participant_id, contact = self.state.contact_service.assign_role(command)
        self.send_contract_success(
            201,
            {"id": participant_id, "contact": contact.to_dict()},
            versioned=versioned,
        )

    def handle_document_status(self) -> None:
        payload = self.read_json_body()
        doc_id = str(payload.get("docId", "")).strip()
        status = str(payload.get("status", "")).strip()
        note = str(payload.get("note", "")).strip()[:1000]
        if not re.fullmatch(r"[0-9A-Za-zА-Яа-яІіЇїЄєҐґ_.:-]{3,160}", doc_id):
            raise ValueError("Некоректний ідентифікатор документа")
        if status not in DOCUMENT_WORK_STATUSES:
            raise ValueError("Невідомий статус документа")
        tree = build_document_tree(self.state.root)
        known_ids = {
            str(document.get("id", ""))
            for proceeding in tree.get("proceedings", [])
            for document in proceeding.get("documents", [])
        }
        if doc_id not in known_ids:
            raise ValueError("Документ не знайдено у поточному дереві")
        with self.state.lock:
            path = self.state.root / ".caseflow" / "document_status.json"
            statuses = read_json(path, {})
            statuses[doc_id] = {"status": status, "note": note, "updated_at": now_iso()}
            write_json(path, statuses)
        self.send_json(200, {"ok": True, "docId": doc_id, "status": status, "note": note})

    def handle_anomalies_latest(self) -> None:
        path = self.state.root / "tmp" / "caseflow_anomalies" / "latest.json"
        report = read_json(path, None)
        if not report:
            self.send_json(200, {"ok": True, "available": False, "findings": [], "summary": anomaly_summary([])})
            return
        self.send_json(200, {"ok": True, "available": True, **report})

    def handle_upload(self) -> None:
        with self.state.exclusive_job("upload"):
            self._handle_upload_unlocked()

    def _handle_upload_unlocked(self) -> None:
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0 or length > MAX_UPLOAD_BYTES:
            raise ValueError("Некоректний або завеликий пакет")
        temporary_root = self.state.root / ".caseflow" / "tmp" / "uploads"
        temporary_root.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(prefix="request_", dir=temporary_root) as temporary:
            fields, uploads = parse_multipart_form(
                self.rfile,
                self.headers.get("Content-Type", ""),
                length,
                Path(temporary),
            )
            proceeding = safe_segment(
                fields.get("proceeding", "НОВЕ_ПРОВАДЖЕННЯ"),
                "НОВЕ_ПРОВАДЖЕННЯ",
            )
            flow = fields.get("flow", "02_МОЇ_ДОКУМЕНТИ")
            if flow not in {"01_ВІД_СУДУ", "02_МОЇ_ДОКУМЕНТИ"}:
                raise ValueError("Невідомий потік")
            channel = safe_segment(fields.get("channel", "ІНШЕ").upper(), "ІНШЕ")
            options = json.loads(fields.get("options", "{}"))
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            destination = self.state.root / "00_INBOX" / proceeding / flow / f"{stamp}__{channel}"
            destination.mkdir(parents=True, exist_ok=False)
            saved = []
            try:
                for item in uploads:
                    if item["field"] != "files" or not item["filename"]:
                        continue
                    target = safe_upload_path(destination, item["filename"])
                    item["path"].replace(target)
                    saved.append(
                        {
                            "name": str(target.relative_to(destination)),
                            "path": str(target.relative_to(self.state.root)),
                            "bytes": target.stat().st_size,
                        }
                    )
                if not saved:
                    raise ValueError("Файли не отримано")
                manifest = {
                    "uploaded_at": now_iso(),
                    "proceeding_folder": proceeding,
                    "flow": flow,
                    "channel": channel,
                    "options": options,
                    "files": saved,
                }
                write_json(destination / "caseflow_upload.json", manifest)
            except Exception:
                shutil.rmtree(destination, ignore_errors=True)
                raise
        self.send_json(
            200,
            {
                "ok": True,
                "destination": str(destination.relative_to(self.state.root)),
                "saved": saved,
            },
        )

    def handle_api_v1_status(self) -> None:
        self.send_json(
            200,
            success_envelope(
                {
                    "server": {"product": "VARTA", "version": VARTA_VERSION},
                    "api": {"version": API_VERSION},
                }
            ),
        )

    def handle_process(self) -> None:
        payload = self.read_json_body()
        settings = payload.get("settings", {})
        preflight = build_preflight(self.state)
        if not preflight["can_start"]:
            blockers = "; ".join(item["detail"] for item in preflight["checks"] if item["status"] == "blocker")
            raise RuntimeError(f"Запуск заблоковано передстартовою перевіркою: {blockers}")
        with self.state.exclusive_job("process"):
            script = self.state.root / "scripts" / "caseflow_process.py"
            if not script.exists():
                script = APP_DIR / "caseflow_process.py"
            worker_args = ["--root", str(self.state.root), "--settings-json", json.dumps(settings, ensure_ascii=False)]
            result = run_worker(script, worker_args, self.state.root, 60 * 60)
            if result.returncode != 0:
                raise RuntimeError((result.stderr or result.stdout or "Помилка обробки")[-4000:])
            lines = [line for line in result.stdout.splitlines() if line.strip()]
            response = json.loads(lines[-1]) if lines else {"message": "Опрацювання завершено"}
            anomaly = None
            anomaly_error = None
            if settings.get("runAnomalyScan", True) and response.get("register"):
                try:
                    anomaly = self._execute_anomaly(response["register"])
                except Exception as exc:  # noqa: BLE001
                    anomaly_error = str(exc)
        self.send_json(200, {"ok": True, **response, "preflight": preflight, "anomalies": anomaly, "anomaly_error": anomaly_error})

    def _execute_anomaly(self, register: str | None = None) -> dict:
        script = self.state.root / "scripts" / "anomaly_detector.py"
        if not script.exists():
            script = APP_DIR / "anomaly_detector.py"
        if not script.exists() and not getattr(sys, "frozen", False):
            raise FileNotFoundError("Не знайдено scripts/anomaly_detector.py")
        worker_args = ["--root", str(self.state.root)]
        if register:
            register_path = Path(register)
            if not register_path.is_absolute():
                register_path = self.state.root / register_path
            register_path = register_path.resolve()
            if self.state.root not in register_path.parents or register_path.suffix.lower() != ".xlsx":
                raise ValueError("Некоректний шлях Реєстру для контролю нестиковок")
            worker_args.extend(["--register", str(register_path)])
        result = run_worker(script, worker_args, self.state.root, 60 * 60)
        if result.returncode != 0:
            raise RuntimeError((result.stderr or result.stdout or "Помилка контролю нестиковок")[-4000:])
        report = read_json(self.state.root / "tmp" / "caseflow_anomalies" / "latest.json", None)
        if not report:
            raise RuntimeError("Аналізатор не створив latest.json")
        return report

    def handle_anomalies_run(self) -> None:
        self.read_json_body()
        with self.state.exclusive_job("anomaly_scan"):
            report = self._execute_anomaly()
        self.send_json(200, {"ok": True, "available": True, **report})

    def handle_anomaly_status(self) -> None:
        payload = self.read_json_body()
        fingerprint = str(payload.get("fingerprint", "")).strip().upper()
        status = str(payload.get("status", "")).strip()
        note = str(payload.get("note", "")).strip()[:1000]
        if not re.fullmatch(r"[0-9A-F]{24}", fingerprint):
            raise ValueError("Некоректний fingerprint")
        allowed = {"open", "acknowledged", "resolved", "false_positive"}
        if status not in allowed:
            raise ValueError("Невідомий статус ручної перевірки")
        with self.state.lock:
            status_path = self.state.root / ".caseflow" / "anomaly_status.json"
            statuses = read_json(status_path, {})
            statuses[fingerprint] = {"status": status, "note": note, "updated_at": now_iso()}
            write_json(status_path, statuses)
            latest_path = self.state.root / "tmp" / "caseflow_anomalies" / "latest.json"
            report = read_json(latest_path, None)
            if report:
                found = False
                for item in report.get("findings", []):
                    if item.get("fingerprint") == fingerprint:
                        item["status"] = status
                        item["status_note"] = note
                        found = True
                if not found:
                    raise ValueError("Картку нестиковки не знайдено в останньому звіті")
                report["summary"] = anomaly_summary(report.get("findings", []))
                write_json(latest_path, report)
        self.send_json(200, {"ok": True, "status": status, "summary": report.get("summary") if report else None})

    def handle_settings(self) -> None:
        payload = self.read_json_body()
        if "caseNumber" in payload:
            self.state.config["case_number"] = safe_segment(payload["caseNumber"], self.state.root.name)
        if "panelOpacity" in payload:
            self.state.config.setdefault("ui", {})["panel_opacity"] = max(45, min(98, int(payload["panelOpacity"])))
        self.state.save_config()
        self.send_json(200, {"ok": True})

    def handle_google_config(self) -> None:
        payload = self.read_json_body()
        client_id = str(payload.get("clientId", "")).strip()
        client_secret = str(payload.get("clientSecret", "")).strip()
        if client_id and not client_id.endswith(".apps.googleusercontent.com"):
            raise ValueError("Client ID має закінчуватися на .apps.googleusercontent.com")
        self.state.config["google"] = {"client_id": client_id}
        if client_secret:
            self.state.save_google_secret(client_secret)
        elif self.state.google_secret_path.exists():
            self.state.google_secret_path.unlink()
        self.state.save_config()
        self.send_json(200, {"ok": True, "configured": bool(client_id)})

    def handle_google_login(self) -> None:
        google = self.state.config.get("google", {})
        client_id = google.get("client_id", "")
        if not client_id:
            raise ValueError("Спочатку збережіть OAuth Client ID")
        verifier = secrets.token_urlsafe(64)[:96]
        challenge = base64.urlsafe_b64encode(hashlib.sha256(verifier.encode("ascii")).digest()).decode("ascii").rstrip("=")
        state_value = secrets.token_urlsafe(32)
        redirect_uri = f"http://127.0.0.1:{self.state.port}/oauth/google/callback"
        self.state.oauth_pending = {"state": state_value, "verifier": verifier, "redirect_uri": redirect_uri, "created": str(time.time())}
        params = {
            "client_id": client_id,
            "redirect_uri": redirect_uri,
            "response_type": "code",
            "scope": GOOGLE_SCOPE,
            "access_type": "offline",
            "prompt": "consent",
            "state": state_value,
            "code_challenge": challenge,
            "code_challenge_method": "S256",
        }
        self.send_json(200, {"ok": True, "url": f"{GOOGLE_AUTH_URL}?{urllib.parse.urlencode(params)}"})

    def handle_google_callback(self, query: dict[str, list[str]]) -> None:
        pending = self.state.oauth_pending
        if not pending or query.get("state", [""])[0] != pending.get("state"):
            self.send_html_message("Підключення відхилено", "Недійсний або прострочений параметр state.", False)
            return
        if query.get("error"):
            self.send_html_message("Google Drive не підключено", query["error"][0], False)
            return
        code = query.get("code", [""])[0]
        google = self.state.config.get("google", {})
        form = {
            "client_id": google.get("client_id", ""),
            "code": code,
            "code_verifier": pending["verifier"],
            "grant_type": "authorization_code",
            "redirect_uri": pending["redirect_uri"],
        }
        client_secret = self.state.load_google_secret()
        if client_secret:
            form["client_secret"] = client_secret
        try:
            token = post_form(GOOGLE_TOKEN_URL, form)
            token["expires_at"] = (datetime.now(timezone.utc) + timedelta(seconds=int(token.get("expires_in", 3600)))).isoformat()
            self.state.save_google_token(token)
            self.state.oauth_pending = {}
            self.send_html_message("Google Drive підключено", "Можна закрити це вікно й повернутися до VARTA.", True)
        except Exception as exc:  # noqa: BLE001
            self.send_html_message("Помилка Google OAuth", str(exc), False)

    def send_html_message(self, title: str, message: str, success: bool) -> None:
        color = "#35d07f" if success else "#ff6b6b"
        html = f"<!doctype html><meta charset='utf-8'><title>{title}</title><body style='font:16px \"Segoe UI Variable Text\",\"Segoe UI\",Arial,sans-serif;background:#0d1624;color:#f5f7fb;display:grid;place-items:center;height:100vh;margin:0'><main style='max-width:620px;padding:32px;border:1px solid #30445f;border-radius:22px;background:#17263a'><h1 style='color:{color}'>{title}</h1><p>{message}</p></main></body>".encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(html)))
        self.end_headers()
        self.wfile.write(html)

    def handle_google_disconnect(self) -> None:
        if self.state.token_path.exists():
            self.state.token_path.unlink()
        self.send_json(200, {"ok": True})

    def handle_google_sync(self) -> None:
        payload = self.read_json_body()
        with self.state.exclusive_job("drive_sync"):
            result = sync_to_drive(self.state, list(payload.get("folders", [])))
        self.send_json(200, {"ok": True, **result})


def main() -> int:
    parser = argparse.ArgumentParser(description="Локальна вебв’юха VARTA")
    executable_dir = Path(sys.executable).resolve().parent
    frozen_root = executable_dir
    for candidate in (executable_dir, executable_dir.parent, executable_dir.parent.parent):
        if (candidate / "03_РЕЄСТР").exists() or (candidate / "00_INBOX").exists():
            frozen_root = candidate
            break
    parser.add_argument("--root", type=Path, default=frozen_root if getattr(sys, "frozen", False) else APP_DIR.parent)
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8768)
    parser.add_argument("--open", action="store_true")
    parser.add_argument("--no-open", action="store_true", help=argparse.SUPPRESS)
    args = parser.parse_args()
    if args.host not in {"127.0.0.1", "localhost"}:
        raise ValueError("За замовчуванням VARTA дозволено слухати лише loopback")
    root = args.root.resolve()
    for folder in ["00_INBOX", "01_ОПРАЦЬОВАНО", "02_РОЗПАКОВАНО", "03_РЕЄСТР", "99_ПОТРЕБУЄ_ПЕРЕВІРКИ", "tmp/caseflow_runs", ".caseflow/logs"]:
        (root / folder).mkdir(parents=True, exist_ok=True)
    state = CaseFlowState(root, args.host, args.port)
    state.prepare_database()
    server = ThreadingHTTPServer((args.host, args.port), Handler)
    server.state = state  # type: ignore[attr-defined]
    pid_path = root / ".caseflow" / "server.pid"
    pid_path.write_text(str(os.getpid()), encoding="ascii")
    url = f"http://127.0.0.1:{args.port}/"
    print(json.dumps({"ready": True, "url": url, "root": str(root)}, ensure_ascii=False), flush=True)
    if (args.open or getattr(sys, "frozen", False)) and not args.no_open:
        threading.Timer(0.7, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()
        state.close()
        try:
            if pid_path.read_text(encoding="ascii").strip() == str(os.getpid()):
                pid_path.unlink()
        except (FileNotFoundError, OSError):
            pass
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
