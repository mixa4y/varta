from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl


ABSOLUTE_PATH = re.compile(r"(?<![A-Za-z0-9_])[A-Za-z]:\\[^\r\n]*")


def workbook_signature(workbook: openpyxl.Workbook) -> dict:
    signature: dict[str, object] = {"sheets": {}}
    for sheet in workbook.worksheets:
        formulas = 0
        values = 0
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    values += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
        signature["sheets"][sheet.title] = {
            "dimension": sheet.calculate_dimension(),
            "tables": sorted(sheet.tables.keys()),
            "formulas": formulas,
            "nonempty_cells": values,
            "freeze": str(sheet.freeze_panes) if sheet.freeze_panes else None,
            "conditional_formats": len(sheet.conditional_formatting),
        }
    return signature


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Створити копію XLSX без абсолютних локальних Windows-шляхів."
    )
    parser.add_argument("--source", required=True, type=Path, help="Вхідний XLSX-реєстр.")
    parser.add_argument(
        "--backup-dir",
        type=Path,
        help="Каталог резервної копії; типово BACKUP біля вхідного файла.",
    )
    parser.add_argument(
        "--name-prefix",
        help="Префікс вихідної назви; типово безпечна назва вхідного файла.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    source = args.source.resolve()
    if not source.is_file():
        raise FileNotFoundError(source)
    if source.suffix.lower() != ".xlsx":
        raise ValueError("--source має вказувати на XLSX-файл")

    backup_dir = args.backup_dir.resolve() if args.backup_dir else source.parent / "BACKUP"
    safe_prefix = re.sub(r"[^0-9A-Za-zА-Яа-яІіЇїЄєҐґ._-]+", "_", args.name_prefix or source.stem)
    safe_prefix = safe_prefix.strip(" ._") or "reiestr"

    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    destination = backup_dir / f"{safe_prefix}__sanityzovanyi_bekap_{timestamp}.xlsx"

    workbook = openpyxl.load_workbook(source)
    before = workbook_signature(workbook)
    sanitized: list[dict[str, str]] = []

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and not value.startswith("=") and ABSOLUTE_PATH.search(value):
                    replacement = ABSOLUTE_PATH.sub("[LOCAL_PATH_REMOVED]", value)
                    cell.value = None if replacement == "[LOCAL_PATH_REMOVED]" else replacement
                    sanitized.append({"sheet": sheet.title, "cell": cell.coordinate})

                hyperlink = cell.hyperlink
                if hyperlink and hyperlink.target and ABSOLUTE_PATH.search(hyperlink.target):
                    cell.hyperlink = None
                    sanitized.append({"sheet": sheet.title, "cell": cell.coordinate + " (hyperlink)"})

    workbook.save(destination)
    workbook.close()

    check = openpyxl.load_workbook(destination, data_only=False, read_only=False)
    after = workbook_signature(check)
    residual: list[dict[str, str]] = []
    formula_errors: list[dict[str, str]] = []
    for sheet in check.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and ABSOLUTE_PATH.search(value):
                    residual.append({"sheet": sheet.title, "cell": cell.coordinate})
                if isinstance(value, str) and value.startswith("#"):
                    formula_errors.append({"sheet": sheet.title, "cell": cell.coordinate, "value": value})
    check.close()

    structural_before = {
        sheet: {key: value for key, value in details.items() if key != "nonempty_cells"}
        for sheet, details in before["sheets"].items()
    }
    structural_after = {
        sheet: {key: value for key, value in details.items() if key != "nonempty_cells"}
        for sheet, details in after["sheets"].items()
    }
    result = {
        "source": source.name,
        "backup": destination.name,
        "backup_size": destination.stat().st_size,
        "sanitized_count": len(sanitized),
        "residual_absolute_paths": residual,
        "formula_error_literals": formula_errors,
        "structure_preserved": structural_before == structural_after,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
