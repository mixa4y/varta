from __future__ import annotations

import argparse
import json
import shutil
from pathlib import Path

from openpyxl import load_workbook


def items(payload: object) -> list[dict]:
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("items", "rows"):
            if isinstance(payload.get(key), list):
                return payload[key]
    raise ValueError("Summary sidecar must be an array or contain items/rows")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("summaries", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    if args.input.resolve() == args.output.resolve():
        raise ValueError("Output workbook must be a new file")
    payload = items(json.loads(args.summaries.read_text(encoding="utf-8")))
    by_id: dict[str, dict] = {}
    for item in payload:
        doc_id = str(item.get("doc_id", "")).strip()
        confidence = str(item.get("confidence", "")).strip().lower()
        summary = str(item.get("short_summary", "")).strip()
        if not doc_id or not summary or confidence not in {"high", "medium", "low"}:
            raise ValueError(f"Invalid summary item: {doc_id}")
        if doc_id in by_id:
            raise ValueError(f"Duplicate summary doc_id: {doc_id}")
        by_id[doc_id] = item

    args.output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(args.input, args.output)
    workbook = load_workbook(args.output, data_only=False)
    sheet = workbook["Документи"]
    headers = {str(sheet.cell(1, col).value or "").strip(): col for col in range(1, sheet.max_column + 1)}
    id_col = headers.get("ID документа")
    summary_col = headers.get("Опис / пов’язана подія")
    if not id_col or not summary_col:
        raise ValueError("Required document columns were not found")
    workbook_ids: set[str] = set()
    applied: list[str] = []
    skipped: list[str] = []
    for row in range(2, sheet.max_row + 1):
        doc_id = str(sheet.cell(row, id_col).value or "").strip()
        if not doc_id:
            continue
        if doc_id in workbook_ids:
            raise ValueError(f"Duplicate workbook doc_id: {doc_id}")
        workbook_ids.add(doc_id)
        item = by_id.get(doc_id)
        if not item:
            continue
        confidence = str(item["confidence"]).lower()
        if confidence == "low":
            skipped.append(doc_id)
            continue
        sheet.cell(row, summary_col).value = str(item["short_summary"]).strip()
        sheet.cell(row, summary_col).alignment = sheet.cell(row, summary_col).alignment.copy(wrap_text=True)
        applied.append(doc_id)
    unknown = sorted(set(by_id) - workbook_ids)
    if unknown:
        raise ValueError(f"Unknown summary doc_ids: {unknown}")
    workbook.save(args.output)
    workbook.close()
    print(json.dumps({"output": str(args.output), "applied": applied, "skipped": skipped}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
