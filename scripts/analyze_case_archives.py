from __future__ import annotations

import argparse
import hashlib
import json
import zipfile
from collections import defaultdict
from io import BytesIO
from pathlib import Path, PurePosixPath

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
MAX_DEPTH = 5
MAX_ENTRY_BYTES = 50 * 1024 * 1024
MAX_TOTAL_BYTES = 500 * 1024 * 1024


def sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest().upper()


def unsafe_name(name: str) -> bool:
    pure = PurePosixPath(name.replace("\\", "/"))
    return pure.is_absolute() or ".." in pure.parts


def compound_suffix(name: str) -> str:
    lower = name.lower()
    if lower.endswith(".p7s.2"):
        return ".p7s.2"
    return Path(name).suffix.lower()


def component_hint(name: str) -> str:
    lower = name.casefold()
    suffix = compound_suffix(name)
    if suffix in {".p7s", ".p7s.2"}:
        return "ПІДПИС"
    if "картка_руху" in lower or "картка руху" in lower:
        return "КАРТКА_РУХУ"
    if "протокол_перевірки_кеп" in lower or "протокол перевірки кеп" in lower:
        return "ПРОТОКОЛ_КЕП"
    if "реєстраційна картка" in lower:
        return "КАРТКА_РЕЄСТРАЦІЇ"
    if suffix in {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff"}:
        return "ЗОБРАЖЕННЯ"
    if suffix == ".zip":
        return "ВКЛАДЕНИЙ_ZIP"
    return "ДОКУМЕНТ"


def existing_hashes(register_path: Path) -> dict[str, list[dict]]:
    workbook = load_workbook(register_path, read_only=False, data_only=False)
    sheet = workbook["Файли"]
    result: dict[str, list[dict]] = defaultdict(list)
    for row in range(2, sheet.max_row + 1):
        digest = sheet.cell(row, 12).value
        if not digest:
            continue
        result[str(digest).upper()].append(
            {
                "file_id": sheet.cell(row, 1).value,
                "doc_id": sheet.cell(row, 2).value,
                "component": sheet.cell(row, 3).value,
                "normalized_name": sheet.cell(row, 6).value,
            }
        )
    workbook.close()
    return result


def inspect_zip(
    data: bytes,
    container: str,
    known: dict[str, list[dict]],
    package_hashes: set[str],
    total: list[int],
    depth: int = 0,
) -> list[dict]:
    if depth > MAX_DEPTH:
        return [{"container": container, "depth": depth, "error": "maximum ZIP depth exceeded"}]
    records: list[dict] = []
    with zipfile.ZipFile(BytesIO(data)) as archive:
        for info in archive.infolist():
            if info.is_dir():
                continue
            record = {
                "container": container,
                "entry": info.filename,
                "depth": depth,
                "size": info.file_size,
                "unsafe_path": unsafe_name(info.filename),
                "component_hint": component_hint(info.filename),
            }
            if record["unsafe_path"]:
                record["error"] = "unsafe archive path"
                records.append(record)
                continue
            if info.file_size > MAX_ENTRY_BYTES:
                record["error"] = "entry size limit exceeded"
                records.append(record)
                continue
            total[0] += info.file_size
            if total[0] > MAX_TOTAL_BYTES:
                record["error"] = "total uncompressed size limit exceeded"
                records.append(record)
                return records
            payload = archive.read(info)
            digest = sha256(payload)
            package_hashes.add(digest)
            record["sha256"] = digest
            record["existing_matches"] = known.get(digest, [])
            records.append(record)
            if compound_suffix(info.filename) == ".zip":
                try:
                    records.extend(
                        inspect_zip(
                            payload,
                            f"{container}!{info.filename}",
                            known,
                            package_hashes,
                            total,
                            depth + 1,
                        )
                    )
                except Exception as exc:  # noqa: BLE001
                    records.append(
                        {
                            "container": f"{container}!{info.filename}",
                            "depth": depth + 1,
                            "error": str(exc),
                        }
                    )
    return records


def main() -> None:
    parser = argparse.ArgumentParser(description="Hash and compare case ZIP contents without extraction.")
    parser.add_argument("case_folder", type=Path)
    parser.add_argument("--register", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    case_folder = (ROOT / args.case_folder).resolve()
    register = (ROOT / args.register).resolve()
    output = (ROOT / args.output).resolve()
    if ROOT.resolve() not in case_folder.parents:
        raise ValueError("case folder is outside project")
    if ROOT.resolve() not in output.parents:
        raise ValueError("output is outside project")

    known = existing_hashes(register)
    packages: list[dict] = []
    hash_to_packages: dict[str, list[str]] = defaultdict(list)
    for path in sorted(case_folder.rglob("*.zip"), key=lambda item: str(item).casefold()):
        package_hashes: set[str] = set()
        total = [0]
        records = inspect_zip(path.read_bytes(), path.name, known, package_hashes, total)
        matched_doc_ids = sorted(
            {
                str(match["doc_id"])
                for record in records
                for match in record.get("existing_matches", [])
                if match.get("doc_id")
            }
        )
        for digest in package_hashes:
            hash_to_packages[digest].append(path.name)
        packages.append(
            {
                "source_relative": str(path.relative_to(ROOT)).replace("/", "\\"),
                "source_name": path.name,
                "source_sha256": sha256(path.read_bytes()),
                "stream_folder": path.relative_to(case_folder).parts[0],
                "entry_count": len(records),
                "uncompressed_bytes_read": total[0],
                "matched_existing_doc_ids": matched_doc_ids,
                "matched_existing_entries": sum(bool(record.get("existing_matches")) for record in records),
                "records": records,
            }
        )

    shared = {
        digest: names
        for digest, names in hash_to_packages.items()
        if len(set(names)) > 1
    }
    payload = {
        "schema_version": 1,
        "case_folder": str(args.case_folder).replace("/", "\\"),
        "package_count": len(packages),
        "packages": packages,
        "shared_entry_hashes": shared,
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "output": str(output),
                "packages": len(packages),
                "matched_packages": sum(bool(item["matched_existing_doc_ids"]) for item in packages),
                "shared_hashes": len(shared),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
