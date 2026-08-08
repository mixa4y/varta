from __future__ import annotations

import argparse
import copy
import shutil
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


def date_key(value: object) -> tuple[int, datetime]:
    if isinstance(value, datetime):
        return (0, value)
    if isinstance(value, date):
        return (0, datetime.combine(value, time.min))
    return (1, datetime.max)


def snapshot_cell(cell):
    return {
        "value": cell.value,
        "style": copy.copy(cell._style),
        "number_format": cell.number_format,
        "font": copy.copy(cell.font),
        "fill": copy.copy(cell.fill),
        "border": copy.copy(cell.border),
        "alignment": copy.copy(cell.alignment),
        "protection": copy.copy(cell.protection),
        "hyperlink": copy.copy(cell.hyperlink),
        "comment": copy.copy(cell.comment),
    }


def restore_cell(cell, state) -> None:
    cell.value = state["value"]
    cell._style = copy.copy(state["style"])
    cell.number_format = state["number_format"]
    cell.font = copy.copy(state["font"])
    cell.fill = copy.copy(state["fill"])
    cell.border = copy.copy(state["border"])
    cell.alignment = copy.copy(state["alignment"])
    cell.protection = copy.copy(state["protection"])
    cell._hyperlink = copy.copy(state["hyperlink"])
    cell.comment = copy.copy(state["comment"])


def ensure_legacy_validations(wb) -> None:
    """Replace two Excel x14-only dropdowns that openpyxl cannot preserve."""
    ws = wb["Документи"]
    existing_ranges = {str(item.sqref) for item in ws.data_validations.dataValidation}
    replacements = (
        ("Y2:Y200", '"Процесуальний,Технологічний"'),
        ("Z2:Z200", '"match,mismatch,cross_proceeding,not_verifiable,Не звірено"'),
    )
    for target_range, formula in replacements:
        if target_range in existing_ranges:
            continue
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "Оберіть значення зі списку"
        validation.errorTitle = "Неприпустиме значення"
        validation.prompt = "Оберіть значення зі списку"
        validation.promptTitle = "Доступні значення"
        ws.add_data_validation(validation)
        validation.add(target_range)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("backup", type=Path)
    args = parser.parse_args()

    args.backup.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(args.workbook, args.backup)

    wb = load_workbook(args.workbook)
    ws = wb["Хронологія"]

    first_row = 2
    data_rows = [
        row
        for row in range(first_row, ws.max_row + 1)
        if ws.cell(row, 1).value not in (None, "")
    ]
    if not data_rows:
        raise RuntimeError("Аркуш 'Хронологія' не містить подій")

    last_row = max(data_rows)
    if data_rows != list(range(first_row, last_row + 1)):
        raise RuntimeError("У блоці подій є порожні рядки; автоматичне сортування зупинено")

    snapshots = []
    for row in data_rows:
        snapshots.append(
            {
                "original_row": row,
                "height": ws.row_dimensions[row].height,
                "cells": [snapshot_cell(ws.cell(row, col)) for col in range(1, ws.max_column + 1)],
            }
        )

    snapshots.sort(
        key=lambda record: (
            *date_key(record["cells"][1]["value"]),
            record["original_row"],
        )
    )

    for target_row, record in zip(data_rows, snapshots):
        ws.row_dimensions[target_row].height = record["height"]
        for col, state in enumerate(record["cells"], start=1):
            restore_cell(ws.cell(target_row, col), state)

    ensure_legacy_validations(wb)
    wb.save(args.workbook)

    check = load_workbook(args.workbook, read_only=True, data_only=False)
    check_ws = check["Хронологія"]
    ids = [check_ws.cell(row, 1).value for row in data_rows]
    dates = [check_ws.cell(row, 2).value for row in data_rows]
    if len(ids) != len(set(ids)):
        raise RuntimeError("Після сортування виявлено дублікати ID подій")
    if dates != sorted(dates):
        raise RuntimeError("Після збереження хронологія не є монотонною")

    print(f"sorted_events={len(data_rows)}")
    print(f"first={ids[0]}|{dates[0]}")
    print(f"last={ids[-1]}|{dates[-1]}")
    print(f"backup={args.backup.name}")


if __name__ == "__main__":
    main()
