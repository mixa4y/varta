from __future__ import annotations

import hashlib
import json
import re
import sqlite3
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from case_docket.application.legacy_migration import (
    LegacyDryRunCommand,
    LegacyImportCommand,
    LegacyMigrationService,
)
from case_docket.legacy_import import (
    LEGACY_XLSX_HEADERS,
    SQLiteLegacyMigrationAdapter,
    dry_run,
    field_mapping_catalog,
    import_legacy,
)
from case_docket.repository import SQLiteRepository


FIXTURES = Path(__file__).parent / "fixtures" / "c09"


def _append(sheet, headers: tuple[str, ...], values: dict[str, object]) -> None:
    sheet.append(list(headers))
    sheet.append([values.get(header) for header in headers])


def _linked_xlsx(path: Path, *, title: str = "Вигаданий процесуальний документ") -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    proceeding = workbook.create_sheet("Провадження")
    _append(
        proceeding,
        LEGACY_XLSX_HEADERS["Провадження"],
        {
            "Папка": "SYNTHETIC_PROC",
            "Номер провадження": "SYNTHETIC-PROC-A",
            "Тип": "synthetic",
            "Суд / орган": "Вигаданий орган",
            "Статус": "review",
        },
    )
    documents = workbook.create_sheet("Документи")
    _append(
        documents,
        LEGACY_XLSX_HEADERS["Документи"],
        {
            "ID документа": "SYNTHETIC-DOC-A",
            "Провадження": "SYNTHETIC-PROC-A",
            "Назва документа": title,
            "Тип документа": "synthetic_notice",
            "Класифікація": "unverified",
        },
    )
    files = workbook.create_sheet("Файли")
    _append(
        files,
        LEGACY_XLSX_HEADERS["Файли"],
        {
            "ID файла": "SYNTHETIC-FILE-A",
            "ID документа": "SYNTHETIC-DOC-A",
            "Провадження": "SYNTHETIC-PROC-A",
            "Оригінальна назва": "synthetic.txt",
            "SHA-256": "A" * 64,
        },
    )
    timeline = workbook.create_sheet("Хронологія")
    _append(
        timeline,
        LEGACY_XLSX_HEADERS["Хронологія"],
        {
            "ID події": "SYNTHETIC-EVENT-A",
            "Провадження": "SYNTHETIC-PROC-A",
            "ID документа": "SYNTHETIC-DOC-A",
            "Документ / подія": "Вигадана подія",
            "Тип події": "synthetic_event",
        },
    )
    workbook.save(path)
    workbook.close()


def _backup(tmp_path: Path) -> Path:
    path = tmp_path / "verified-backup"
    path.mkdir()
    return path


def _repository(tmp_path: Path) -> SQLiteRepository:
    return SQLiteRepository(tmp_path / "target.sqlite3")


def test_versioned_mapping_catalog_covers_every_known_xlsx_field() -> None:
    catalog = field_mapping_catalog()
    actual = {
        (item.source_scope, item.source_field)
        for item in catalog
        if item.source_scope in LEGACY_XLSX_HEADERS
    }
    expected = {
        (sheet, field)
        for sheet, fields in LEGACY_XLSX_HEADERS.items()
        for field in fields
    }
    assert actual == expected
    assert {item.classification for item in catalog} >= {
        "mapped",
        "unsupported",
        "lossy",
        "ambiguous",
        "derived",
    }


def test_application_service_dry_run_is_zero_write_and_golden(tmp_path: Path) -> None:
    source = tmp_path / "synthetic.caseflow"
    source.write_bytes((FIXTURES / "synthetic-source.json").read_bytes())
    before = source.read_bytes()
    connection = sqlite3.connect(":memory:")
    service = LegacyMigrationService(SQLiteLegacyMigrationAdapter(connection))
    report = service.preview(LegacyDryRunCommand(source, source_key="synthetic-fixture-v1"))

    assert report.zero_db_writes is True
    assert connection.execute(
        "SELECT COUNT(*) FROM sqlite_master WHERE name LIKE 'legacy_import_%'"
    ).fetchone()[0] == 0
    assert source.read_bytes() == before
    golden = json.loads((FIXTURES / "golden-dry-run.json").read_text(encoding="utf-8"))
    assert report.to_dict() == golden
    connection.close()


def test_xlsx_two_pass_first_import_and_reimport_are_idempotent(tmp_path: Path) -> None:
    source = tmp_path / "synthetic.xlsx"
    _linked_xlsx(source)
    before_hash = hashlib.sha256(source.read_bytes()).hexdigest()
    repository = _repository(tmp_path)
    service = LegacyMigrationService(SQLiteLegacyMigrationAdapter(repository._conn))
    command = LegacyImportCommand(source, _backup(tmp_path), source_key="synthetic-xlsx")

    first = service.migrate(command)
    second = service.migrate(command)

    assert (first.records, first.imported, first.updated, first.skipped) == (4, 4, 0, 0)
    assert first.resolved_links == 5 and first.unresolved_links == 0
    assert (second.imported, second.updated, second.skipped) == (0, 0, 4)
    assert repository._conn.execute("SELECT COUNT(*) FROM legacy_import_records").fetchone()[0] == 4
    assert repository._conn.execute("SELECT COUNT(*) FROM legacy_import_links").fetchone()[0] == 5
    assert hashlib.sha256(source.read_bytes()).hexdigest() == before_hash
    assert len(first.record_results) == first.records
    repository.close()


def test_stable_external_reference_updates_instead_of_duplicating(tmp_path: Path) -> None:
    source = tmp_path / "synthetic.xlsx"
    _linked_xlsx(source)
    repository = _repository(tmp_path)
    backup = _backup(tmp_path)
    first = import_legacy(
        repository._conn, source, backup_destination=backup, source_key="update-source"
    )
    assert first.imported == 4

    workbook = load_workbook(source)
    sheet = workbook["Документи"]
    title_column = list(LEGACY_XLSX_HEADERS["Документи"]).index("Назва документа") + 1
    sheet.cell(2, title_column, "Оновлений вигаданий документ")
    workbook.save(source)
    workbook.close()

    second = import_legacy(
        repository._conn, source, backup_destination=backup, source_key="update-source"
    )
    assert (second.imported, second.updated, second.skipped) == (0, 1, 3)
    assert repository._conn.execute("SELECT COUNT(*) FROM legacy_import_records").fetchone()[0] == 4
    payload = repository._conn.execute(
        "SELECT payload_json FROM legacy_import_records WHERE external_ref LIKE '%SYNTHETIC-DOC-A'"
    ).fetchone()[0]
    assert "Оновлений вигаданий документ" in payload
    repository.close()


def test_duplicate_unsupported_and_broken_reference_are_explained(tmp_path: Path) -> None:
    source = tmp_path / "conflict.caseflow"
    source.write_text(
        json.dumps(
            {
                "records": [
                    {"id": "same", "unknown_field": "synthetic"},
                    {"id": "same", "document_id": "missing-document"},
                ]
            }
        ),
        encoding="utf-8",
    )
    report = dry_run(source, source_key="conflict-source")

    assert report.conflicts == 1
    assert report.unresolved_links == 1
    assert len(report.record_results) == report.records == 2
    assert {issue["reason"] for issue in report.issues} >= {
        "duplicate_external_ref",
        "unresolved_link",
    }
    assert any(
        item.classification == "unsupported" and item.source_field == "unknown_field"
        for item in report.field_mappings
    )

    repository = _repository(tmp_path)
    imported = import_legacy(
        repository._conn,
        source,
        backup_destination=_backup(tmp_path),
        source_key="conflict-source",
    )
    assert (imported.imported, imported.skipped) == (1, 1)
    assert repository._conn.execute(
        "SELECT COUNT(*) FROM legacy_import_records WHERE status='manual_review_required'"
    ).fetchone()[0] == 1
    assert repository._conn.execute(
        "SELECT COUNT(*) FROM legacy_import_links WHERE status='manual_review_required'"
    ).fetchone()[0] == 1
    repository.close()


def test_caseflow_directory_inventory_skips_sensitive_state_and_preserves_tree(
    tmp_path: Path,
) -> None:
    source = tmp_path / ".caseflow"
    source.mkdir()
    (source / "config.json").write_text(
        '{"id":"synthetic-config","mode":"review"}', encoding="utf-8"
    )
    (source / "index.json").write_text(
        '{"records":[{"id":"synthetic-index","sha256":"' + "B" * 64 + '"}]}',
        encoding="utf-8",
    )
    secrets = source / "secrets"
    secrets.mkdir()
    (secrets / "oauth.json").write_text('{"token":"synthetic-secret"}', encoding="utf-8")
    before = {
        item.relative_to(source).as_posix(): hashlib.sha256(item.read_bytes()).hexdigest()
        for item in source.rglob("*")
        if item.is_file()
    }
    repository = _repository(tmp_path)
    report = import_legacy(
        repository._conn,
        source,
        backup_destination=_backup(tmp_path),
        source_key="runtime-directory",
    )
    after = {
        item.relative_to(source).as_posix(): hashlib.sha256(item.read_bytes()).hexdigest()
        for item in source.rglob("*")
        if item.is_file()
    }
    assert report.records == 2
    assert any(issue["reason"] == "sensitive_runtime_file" for issue in report.issues)
    assert before == after
    assert all("synthetic-secret" not in row[0] for row in repository._conn.execute(
        "SELECT payload_json FROM legacy_import_records"
    ))
    repository.close()


def test_import_transaction_rolls_back_without_source_change(tmp_path: Path) -> None:
    seed = tmp_path / "seed.caseflow"
    seed.write_text('{"records":[]}', encoding="utf-8")
    repository = _repository(tmp_path)
    backup = _backup(tmp_path)
    import_legacy(
        repository._conn, seed, backup_destination=backup, source_key="rollback-source"
    )
    repository._conn.execute(
        """
        CREATE TRIGGER synthetic_c09_failure
        BEFORE INSERT ON legacy_import_records
        WHEN NEW.external_ref LIKE '%FAIL%'
        BEGIN SELECT RAISE(ABORT, 'synthetic C09 failure'); END
        """
    )
    repository._conn.commit()
    source = tmp_path / "rollback.caseflow"
    source.write_text(
        '{"records":[{"id":"GOOD"},{"id":"FAIL"}]}', encoding="utf-8"
    )
    before = source.read_bytes()
    runs_before = repository._conn.execute("SELECT COUNT(*) FROM legacy_import_runs").fetchone()[0]
    with pytest.raises(sqlite3.IntegrityError, match="synthetic C09 failure"):
        import_legacy(
            repository._conn,
            source,
            backup_destination=backup,
            source_key="rollback-source-2",
        )
    assert repository._conn.execute("SELECT COUNT(*) FROM legacy_import_runs").fetchone()[0] == runs_before
    assert repository._conn.execute(
        "SELECT COUNT(*) FROM legacy_import_records WHERE source_key='rollback-source-2'"
    ).fetchone()[0] == 0
    assert source.read_bytes() == before
    repository.close()


@pytest.mark.parametrize(
    "payload",
    ['{"records": null}', '{"items": {"id": "synthetic-record"}}'],
)
def test_caseflow_records_and_items_must_be_arrays(tmp_path: Path, payload: str) -> None:
    source = tmp_path / "invalid.caseflow"
    source.write_text(payload, encoding="utf-8")
    with pytest.raises(ValueError, match="records/items.*array"):
        dry_run(source)


def test_import_requires_verified_backup_destination(tmp_path: Path) -> None:
    source = tmp_path / "synthetic.caseflow"
    source.write_text('{"records": [{"id": "one"}]}', encoding="utf-8")
    repository = _repository(tmp_path)
    with pytest.raises(ValueError, match="backup destination"):
        import_legacy(
            repository._conn,
            source,
            backup_destination=tmp_path / "missing",
        )
    repository.close()


def test_frozen_fixtures_are_synthetic_and_privacy_safe() -> None:
    combined = "\n".join(
        path.read_text(encoding="utf-8") for path in sorted(FIXTURES.iterdir())
    )
    assert not re.search(r"\b\d{1,6}/\d{1,8}/\d{2,4}\b", combined)
    assert not re.search(
        r"BEGIN (?:RSA |OPENSSH )?PRIVATE KEY|access_token|client_secret|dpapi",
        combined,
        re.IGNORECASE,
    )
