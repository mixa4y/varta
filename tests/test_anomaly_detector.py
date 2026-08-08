from __future__ import annotations

import hashlib
import json
import subprocess
import sys
import tempfile
import unittest
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook


WORKSPACE = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(WORKSPACE / "caseflow"))

import anomaly_detector  # noqa: E402
import caseflow_process  # noqa: E402
import server  # noqa: E402


class AnomalyDetectorTests(unittest.TestCase):
    def setUp(self):
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        (self.root / "03_РЕЄСТР" / "exports").mkdir(parents=True)
        (self.root / "01_ОПРАЦЬОВАНО" / "1_111_333_34" / "02_МОЇ_ДОКУМЕНТИ" / "DOC_0001").mkdir(parents=True)
        self.document = self.root / "01_ОПРАЦЬОВАНО" / "1_111_333_34" / "02_МОЇ_ДОКУМЕНТИ" / "DOC_0001" / "заява.txt"
        self.document.write_text("ЗАЯВА\nПровадження 1/111/222/33\n", encoding="utf-8")
        self.digest = hashlib.sha256(self.document.read_bytes()).hexdigest().upper()
        self.register = self.root / "03_РЕЄСТР" / "exports" / "registry.xlsx"
        self._make_register()
        (self.root / "03_РЕЄСТР" / "ОСТАННІЙ_РЕЄСТР.txt").write_text(str(self.register), encoding="utf-8")

    def tearDown(self):
        self.temporary.cleanup()

    def _make_register(self):
        workbook = Workbook()
        workbook.remove(workbook.active)
        docs = workbook.create_sheet("Документи")
        docs.append(
            [
                "ID документа", "Дата документа", "Дата надходження/подання", "Провадження", "Потік",
                "Тип документа", "Основний файл", "Додатків очікується", "Додатків фактично", "Картка руху",
                "Протокол КЕП", "Підпис", "Статус звірки", "Примітки",
            ]
        )
        docs.append(
            [
                "DOC_0001", datetime(2026, 4, 4), datetime(2026, 4, 3, 10, 0), "1/111/222/33", "Мої документи",
                "Заява", "Є", 2, 1, "Є", "Є", "Є", "cross_proceeding",
                "cross_proceeding: первинне 1/111/222/33; фактичне розміщення 1/111/333/34.",
            ]
        )
        files = workbook.create_sheet("Файли")
        files.append(["ID файла", "ID документа", "Компонент", "Оригінальна назва", "Відносний шлях", "Повний локальний шлях", "SHA-256", "Провадження"])
        files.append(["FILE_0001", "DOC_0001", "ОСНОВНИЙ", "заява.txt", str(self.document.relative_to(self.root)), str(self.document), self.digest, "1/111/222/33"])
        timeline = workbook.create_sheet("Хронологія")
        timeline.append(["ID події", "Дата / час", "Провадження", "Тип події", "ID документа", "Документ / подія"])
        timeline.append(["EVT_0001", datetime(2026, 4, 3, 10, 0), "1/111/222/33", "Подання документа", "DOC_0001", "Подано"])
        timeline.append(["EVT_0002", datetime(2026, 4, 3, 9, 0), "1/111/222/33", "Реєстрація документа", "DOC_0001", "Зареєстровано"])
        workbook.save(self.register)

    def test_expected_evidence_cards(self):
        report = anomaly_detector.run_detector(self.root, self.register)
        rules = {item["rule_id"] for item in report["findings"]}
        self.assertIn("CROSS_PROCEEDING", rules)
        self.assertIn("ATTACHMENT_COUNT_MISMATCH", rules)
        self.assertIn("DOCUMENT_AFTER_SUBMISSION", rules)
        self.assertIn("EVENT_SEQUENCE_CONTRADICTION", rules)
        self.assertEqual(report["detector_version"], "1.1.0")
        self.assertTrue(all(item["neutral_notice"] for item in report["findings"]))
        self.assertTrue((self.root / report["output"]).exists())

    def test_changed_file_is_detected_between_scans(self):
        anomaly_detector.run_detector(self.root, self.register)
        self.document.write_text("ЗМІНЕНИЙ ВМІСТ", encoding="utf-8")
        report = anomaly_detector.run_detector(self.root, self.register)
        rules = {item["rule_id"] for item in report["findings"]}
        self.assertIn("FILE_HASH_MISMATCH", rules)
        self.assertIn("FILE_CHANGED_BETWEEN_SCANS", rules)

    def test_status_metadata_is_reapplied(self):
        first = anomaly_detector.run_detector(self.root, self.register)
        fingerprint = first["findings"][0]["fingerprint"]
        status_path = self.root / ".caseflow" / "anomaly_status.json"
        status_path.parent.mkdir(parents=True, exist_ok=True)
        status_path.write_text(json.dumps({fingerprint: {"status": "acknowledged", "note": "перевіряється"}}), encoding="utf-8")
        second = anomaly_detector.run_detector(self.root, self.register)
        item = next(value for value in second["findings"] if value["fingerprint"] == fingerprint)
        self.assertEqual(item["status"], "acknowledged")
        self.assertEqual(item["status_note"], "перевіряється")

    def test_timeline_sort_uses_event_time(self):
        from openpyxl import load_workbook

        workbook = load_workbook(self.register)
        sheet = workbook["Хронологія"]
        caseflow_process.sort_timeline(sheet)
        self.assertEqual(sheet.cell(2, 1).value, "EVT_0002")
        self.assertEqual(sheet.cell(3, 1).value, "EVT_0001")

    def test_upload_path_preserves_subfolders_and_blocks_traversal(self):
        upload = self.root / "00_INBOX" / "packet"
        upload.mkdir(parents=True)
        target = server.safe_upload_path(upload, "вкладена/доказ.txt")
        self.assertEqual(target.parent.name, "вкладена")
        with self.assertRaises(ValueError):
            server.safe_upload_path(upload, "../outside.txt")

    def test_document_tree_statuses_are_based_on_register_fields(self):
        self.assertEqual(server.document_work_status("КОМПЛЕКТНИЙ", "Контролювати відповідь"), "completed")
        self.assertEqual(server.document_work_status("ПЕРЕВІРИТИ", ""), "needs_review")
        self.assertEqual(server.document_work_status("", "Очікувати відповідь суду"), "waiting")
        self.assertEqual(server.document_work_status("", "Подати пояснення"), "in_progress")

    def test_document_tree_contains_registered_document_and_files(self):
        tree = server.build_document_tree(self.root)
        documents = [document for group in tree["proceedings"] for document in group["documents"]]
        document = next(item for item in documents if item["id"] == "DOC_0001")
        self.assertEqual(document["files"][0]["name"], "заява.txt")
        self.assertEqual(tree["counts"]["all"], 1)

    def test_archive_tree_entries_classify_esud_package(self):
        archive_path = self.root / "sample.zip"
        with zipfile.ZipFile(archive_path, "w") as archive:
            archive.writestr("main/document.pdf", b"pdf")
            archive.writestr("atch/evidence.pdf", b"attachment")
            archive.writestr("tech/card.pdf", b"card")
            archive.writestr("main/document.pdf.p7s", b"signature")
            archive.writestr("../unsafe.txt", b"unsafe")
        entries = server.archive_tree_entries(archive_path)
        self.assertEqual([item["component"] for item in entries], ["Основний документ", "Додаток", "Технічний документ", "Підпис КЕП"])

    def test_two_pipeline_runs_accumulate_in_latest_export(self):
        pipeline_root = self.root / "pipeline"
        (pipeline_root / "03_РЕЄСТР").mkdir(parents=True)
        (pipeline_root / ".caseflow").mkdir(parents=True)
        baseline = pipeline_root / "03_РЕЄСТР" / "base.xlsx"
        workbook = caseflow_process.ensure_workbook(None)
        workbook.save(baseline)
        (pipeline_root / "03_РЕЄСТР" / "ОСТАННІЙ_РЕЄСТР.txt").write_text(str(baseline), encoding="utf-8")
        (pipeline_root / ".caseflow" / "config.json").write_text(json.dumps({"case_number": "111/2222/33"}), encoding="utf-8")

        def add_packet(number: int):
            folder = pipeline_root / "00_INBOX" / "1_111_222_33" / "02_МОЇ_ДОКУМЕНТИ" / f"20260722_12000{number}__ЕСУД_МОЇ"
            folder.mkdir(parents=True)
            document = folder / f"заява_{number}.txt"
            document.write_text(f"ЗАЯВА {number}\nПровадження 1/111/222/33\n", encoding="utf-8")
            (folder / "caseflow_upload.json").write_text(
                json.dumps(
                    {
                        "uploaded_at": f"2026-07-22T12:00:0{number}+03:00",
                        "proceeding_folder": "1_111_222_33",
                        "flow": "02_МОЇ_ДОКУМЕНТИ",
                        "channel": "ЕСУД_МОЇ",
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

        def run_pipeline():
            result = subprocess.run(
                [sys.executable, str(WORKSPACE / "caseflow" / "caseflow_process.py"), "--root", str(pipeline_root), "--settings-json", "{}"],
                check=True,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
            )
            self.assertEqual(result.returncode, 0)

        add_packet(1)
        run_pipeline()
        first_export = Path((pipeline_root / "03_РЕЄСТР" / "ОСТАННІЙ_РЕЄСТР.txt").read_text(encoding="utf-8-sig").strip()).resolve()
        add_packet(2)
        run_pipeline()
        second_export = Path((pipeline_root / "03_РЕЄСТР" / "ОСТАННІЙ_РЕЄСТР.txt").read_text(encoding="utf-8-sig").strip()).resolve()
        self.assertNotEqual(first_export, second_export)

        from openpyxl import load_workbook

        result_book = load_workbook(second_export, read_only=True, data_only=False)
        try:
            ids = [result_book["Документи"].cell(row, 1).value for row in range(2, result_book["Документи"].max_row + 1)]
        finally:
            result_book.close()
        self.assertEqual([value for value in ids if value], ["DOC_0001", "DOC_0002"])
        pointer = Path((pipeline_root / "03_РЕЄСТР" / "ОСТАННІЙ_РЕЄСТР.txt").read_text(encoding="utf-8-sig").strip()).resolve()
        self.assertEqual(pointer, second_export)


if __name__ == "__main__":
    unittest.main()
