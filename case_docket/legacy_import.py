"""Read-only XLSX/.caseflow inventory, reconciliation and SQLite adapter.

The adapter deliberately stores migration state outside authoritative domain tables.
Promotion/read-path switching is a separate reconciliation/configuration decision.
"""
from __future__ import annotations

import hashlib
import json
import os
import sqlite3
import uuid
from collections import Counter
from dataclasses import asdict, dataclass, replace
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook

from case_docket.application.legacy_migration import (
    LegacyDryRunCommand,
    LegacyImportCommand,
)


LEGACY_XLSX_SCHEMA_VERSION = "caseflow-register-v1"
LEGACY_CASEFLOW_SCHEMA_VERSION = "caseflow-runtime-json-v1"

LEGACY_XLSX_HEADERS: dict[str, tuple[str, ...]] = {
    "Партії": (
        "ID партії", "Провадження", "Джерело", "№ партії", "Дата завантаження",
        "Очікувано документів", "Фактично документів", "Статус",
        "Дата опрацювання", "Примітки",
    ),
    "Документи": (
        "ID документа", "ID ЄСІТС", "№ документа", "Дата документа",
        "Дата надходження/подання", "Провадження", "Суд / джерело", "Потік",
        "Назва документа", "Тип документа", "Опис / пов’язана подія", "ID партії",
        "Основний файл", "Додатків очікується", "Додатків фактично", "Картка руху",
        "Протокол КЕП", "Підпис", "Статус комплектності", "Наступна дія",
        "Відносна папка", "Google Drive URL", "Airtable record ID", "Примітки",
        "Класифікація", "Статус звірки",
    ),
    "Хронологія": (
        "ID події", "Дата / час", "Провадження", "Напрям / джерело",
        "Учасник / орган", "Тип події", "ID документа", "Документ / подія",
        "Статус / результат", "Наступна дія", "Строк", "ID партії",
        "Посилання / шлях", "Примітки",
    ),
    "Файли": (
        "ID файла", "ID документа", "Компонент", "№ компонента", "Оригінальна назва",
        "Нормалізована назва", "Відносний шлях", "Повний локальний шлях", "Розширення",
        "MIME", "Розмір, байт", "SHA-256", "Дублікат?", "Цілісність",
        "Дата завантаження", "Дата опрацювання", "Google Drive URL",
        "Airtable record ID", "ID партії", "Провадження", "Примітки",
    ),
    "Провадження": (
        "Папка", "Номер провадження", "Тип", "Суд / орган", "Статус",
        "Поточний потік", "Документів від суду", "Моїх документів",
        "Всього документів", "Комплектних", "Потребують перевірки",
        "Остання дата документа", "Примітки",
    ),
    "Довідники": (
        "Провадження", "Потік", "Компонент", "Тип документа", "Статус партії",
        "Основний файл", "Картка руху", "Протокол КЕП", "Підпис", "Цілісність",
        "Дублікат", "Статус провадження", "Комплектність", "Правило назви",
        "judgment_code", "name", "Класифікація", "Статус звірки",
        "Канал надходження", "Пояснення каналу",
    ),
}

_PRIMARY_FIELDS: dict[str, tuple[str, ...]] = {
    "Партії": ("ID партії",),
    "Документи": ("ID документа",),
    "Хронологія": ("ID події",),
    "Файли": ("ID файла",),
    "Провадження": ("Номер провадження", "Папка"),
}

_LINK_FIELDS: dict[tuple[str, str], str] = {
    ("Партії", "Провадження"): "Провадження",
    ("Документи", "Провадження"): "Провадження",
    ("Документи", "ID партії"): "Партії",
    ("Хронологія", "Провадження"): "Провадження",
    ("Хронологія", "ID документа"): "Документи",
    ("Хронологія", "ID партії"): "Партії",
    ("Файли", "ID документа"): "Документи",
    ("Файли", "ID партії"): "Партії",
    ("Файли", "Провадження"): "Провадження",
}

_MAPPED_TARGETS: dict[tuple[str, str], str] = {
    ("Партії", "ID партії"): "ImportBatchDTO.externalReference",
    ("Партії", "Дата завантаження"): "ImportBatchDTO.createdAt",
    ("Партії", "Статус"): "ImportBatchDTO.status",
    ("Документи", "ID документа"): "ExternalReferenceInput.value",
    ("Документи", "Назва документа"): "CreateEvidenceDocumentCommand.title",
    ("Документи", "Тип документа"): "CreateEvidenceDocumentCommand.document_type",
    ("Документи", "Суд / джерело"): "CreateEvidenceDocumentCommand.source",
    ("Документи", "Потік"): "CreateEvidenceDocumentCommand.process_role",
    ("Документи", "Класифікація"): "CreateEvidenceDocumentCommand.classification",
    ("Документи", "Провадження"): "EvidenceMembershipInput.context_id",
    ("Хронологія", "ID події"): "ExternalReferenceInput.value",
    ("Хронологія", "Дата / час"): "CreateEvidenceEventCommand.event_at",
    ("Хронологія", "Тип події"): "CreateEvidenceEventCommand.event_type",
    ("Хронологія", "Документ / подія"): "CreateEvidenceEventCommand.title",
    ("Хронологія", "Статус / результат"): "CreateEvidenceEventCommand.workflow_status",
    ("Файли", "ID файла"): "ManagedFileRecord.externalReference",
    ("Файли", "ID документа"): "EvidenceDocumentDTO.fileIds",
    ("Файли", "Оригінальна назва"): "ManagedFileRecord.original_name",
    ("Файли", "SHA-256"): "ManagedFileRecord.sha256",
    ("Провадження", "Номер провадження"): "CreateWorkspaceProceedingCommand.proceeding_number",
    ("Провадження", "Тип"): "CreateWorkspaceProceedingCommand.relationship_kind",
    ("Провадження", "Суд / орган"): "CreateWorkspaceProceedingCommand.name",
    ("Провадження", "Статус"): "WorkspaceProceedingDTO.status",
}

_DERIVED_FIELDS = {
    "№ партії", "Очікувано документів", "Фактично документів", "Дата опрацювання",
    "Основний файл", "Додатків очікується", "Додатків фактично", "Картка руху",
    "Протокол КЕП", "Підпис", "Статус комплектності", "Нормалізована назва",
    "Розширення", "MIME", "Розмір, байт", "Дублікат?", "Цілісність",
    "Документів від суду", "Моїх документів", "Всього документів", "Комплектних",
    "Потребують перевірки", "Остання дата документа", "Правило назви",
}
_LOSSY_FIELDS = {"Статус звірки", "Класифікація", "Статус / результат", "Статус"}
_AMBIGUOUS_FIELDS = {
    "Опис / пов’язана подія", "Напрям / джерело", "Учасник / орган", "Наступна дія",
    "Посилання / шлях", "Примітки", "Джерело", "Повний локальний шлях",
}
_SENSITIVE_NAMES = {"secrets", "token", "tokens", "oauth", "dpapi"}
_CASEFLOW_LINK_KEYS = {"document_id", "file_id", "event_id", "batch_id", "proceeding_id"}


@dataclass(frozen=True, slots=True)
class FieldMapping:
    source_scope: str
    source_field: str
    classification: str
    target: str | None
    note: str


@dataclass(frozen=True, slots=True)
class SourceTreeEntry:
    relative_path: str
    sha256: str
    size: int


@dataclass(frozen=True, slots=True)
class LegacyRecord:
    external_ref: str
    record_kind: str
    source_location: str
    payload: Mapping[str, Any]


@dataclass(frozen=True, slots=True)
class LegacyLink:
    from_external_ref: str
    source_field: str
    raw_target: str
    target_external_ref: str | None


@dataclass(frozen=True, slots=True)
class RecordResult:
    external_ref: str
    source_location: str
    record_kind: str
    action: str
    reasons: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class LegacyImportReport:
    source_kind: str
    source_name: str
    source_key: str
    format_version: str
    source_sha256: str
    source_tree_sha256: str
    records: int
    proposed: int
    imported: int
    updated: int
    skipped: int
    conflicts: int
    unresolved_links: int
    resolved_links: int
    counts: Mapping[str, int]
    issues: tuple[dict[str, Any], ...]
    record_results: tuple[RecordResult, ...]
    field_mappings: tuple[FieldMapping, ...]
    source_files: tuple[SourceTreeEntry, ...]
    zero_db_writes: bool
    source_unchanged: bool
    read_path_switch_allowed: bool = False

    def to_dict(self) -> dict[str, Any]:
        return _json_value(asdict(self))


def _json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Mapping):
        return {str(key): _json_value(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_value(item) for item in value]
    return value


def _canonical_json(value: Any) -> str:
    return json.dumps(
        _json_value(value), ensure_ascii=False, sort_keys=True, separators=(",", ":")
    )


def _sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _tree(path: Path) -> tuple[SourceTreeEntry, ...]:
    if path.is_file():
        return (SourceTreeEntry(path.name, _file_hash(path), path.stat().st_size),)
    if not path.is_dir() or path.name.casefold() != ".caseflow":
        raise ValueError("Джерело має бути .xlsx, .caseflow JSON file або .caseflow directory")
    entries: list[SourceTreeEntry] = []
    for item in sorted(path.rglob("*"), key=lambda value: value.as_posix().casefold()):
        if item.is_symlink():
            raise ValueError("Symlink/reparse source у .caseflow не підтримується")
        if item.is_file():
            entries.append(
                SourceTreeEntry(item.relative_to(path).as_posix(), _file_hash(item), item.stat().st_size)
            )
    return tuple(entries)


def _tree_hash(entries: Iterable[SourceTreeEntry]) -> str:
    payload = [(item.relative_path, item.sha256, item.size) for item in entries]
    return _sha256_bytes(_canonical_json(payload).encode("utf-8"))


def _source_key(path: Path, kind: str, explicit: str | None) -> str:
    if explicit is not None:
        value = explicit.strip()
        if not value or any(ord(character) < 32 for character in value):
            raise ValueError("source_key має бути непорожнім opaque identifier")
        return value
    stable_hint = f"{kind}:{path.parent.name.casefold()}:{path.name.casefold()}"
    return _sha256_bytes(stable_hint.encode("utf-8"))


def _xlsx_external_ref(sheet: str, row_number: int, payload: Mapping[str, Any]) -> str:
    for field in _PRIMARY_FIELDS.get(sheet, ()):
        value = payload.get(field)
        if value not in (None, ""):
            return f"xlsx:{sheet}:{field}:{str(value).strip()}"
    return f"xlsx:{sheet}:row:{row_number}"


def _read_xlsx(path: Path) -> tuple[list[LegacyRecord], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    records: list[LegacyRecord] = []
    issues: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            iterator = sheet.iter_rows(values_only=True)
            header_row = next(iterator, None)
            if header_row is None:
                continue
            seen_headers: Counter[str] = Counter()
            headers: list[str] = []
            for column, raw_header in enumerate(header_row, 1):
                header = str(raw_header).strip() if raw_header is not None else ""
                if not header:
                    header = f"__column_{column}"
                seen_headers[header] += 1
                effective = header if seen_headers[header] == 1 else f"{header}#{seen_headers[header]}"
                headers.append(effective)
                if seen_headers[header] > 1:
                    issues.append(
                        {
                            "source_location": f"{sheet.title}!1:{column}",
                            "reason": "duplicate_header",
                            "field": header,
                            "action": "manual_review_required",
                        }
                    )
            for row_number, values in enumerate(iterator, 2):
                payload = {
                    header: _json_value(value)
                    for header, value in zip(headers, values)
                    if value not in (None, "")
                }
                if not payload:
                    continue
                records.append(
                    LegacyRecord(
                        external_ref=_xlsx_external_ref(sheet.title, row_number, payload),
                        record_kind=f"xlsx:{sheet.title}",
                        source_location=f"{sheet.title}!{row_number}",
                        payload=payload,
                    )
                )
    finally:
        workbook.close()
    return records, issues


def _json_items(payload: Any) -> list[Any]:
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        if "records" in payload:
            items = payload["records"]
            if not isinstance(items, list):
                raise ValueError(".caseflow records/items мають бути JSON array")
            return items
        if "items" in payload:
            items = payload["items"]
            if not isinstance(items, list):
                raise ValueError(".caseflow records/items мають бути JSON array")
            return items
        return [payload]
    raise ValueError(".caseflow має містити JSON object або array")


def _read_caseflow_json(
    path: Path,
    *,
    label: str,
) -> list[LegacyRecord]:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    records: list[LegacyRecord] = []
    for index, item in enumerate(_json_items(payload)):
        if isinstance(item, Mapping):
            raw = {str(key): _json_value(value) for key, value in item.items()}
            identity = raw.get("id", raw.get("external_id", index))
        else:
            raw = {"value": _json_value(item)}
            identity = index
        records.append(
            LegacyRecord(
                external_ref=f"caseflow:{label}:{identity}",
                record_kind=f"caseflow:{label}",
                source_location=f"{label}#{index}",
                payload=raw,
            )
        )
    return records


def _is_sensitive(relative_path: str) -> bool:
    parts = {part.casefold() for part in Path(relative_path).parts}
    stem = Path(relative_path).stem.casefold()
    return bool(parts & _SENSITIVE_NAMES) or any(token in stem for token in _SENSITIVE_NAMES)


def _read_caseflow_directory(path: Path) -> tuple[list[LegacyRecord], list[dict[str, Any]]]:
    records: list[LegacyRecord] = []
    issues: list[dict[str, Any]] = []
    for item in sorted(path.rglob("*"), key=lambda value: value.as_posix().casefold()):
        if not item.is_file():
            continue
        relative = item.relative_to(path).as_posix()
        if _is_sensitive(relative):
            issues.append(
                {
                    "source_location": relative,
                    "reason": "sensitive_runtime_file",
                    "action": "skipped",
                }
            )
            continue
        if item.suffix.casefold() != ".json":
            issues.append(
                {
                    "source_location": relative,
                    "reason": "unsupported_runtime_file",
                    "action": "skipped",
                }
            )
            continue
        records.extend(_read_caseflow_json(item, label=relative))
    return records, issues


def _mapping(scope: str, field: str) -> FieldMapping:
    base_field = field.split("#", 1)[0]
    target = _MAPPED_TARGETS.get((scope, base_field))
    if target is not None:
        classification = "lossy" if base_field in _LOSSY_FIELDS else "mapped"
        note = "Mapped to application command/DTO; raw value remains preserved."
    elif base_field in _DERIVED_FIELDS:
        classification = "derived"
        note = "Recomputed or verified by target services; legacy value is provenance only."
    elif base_field in _AMBIGUOUS_FIELDS:
        classification = "ambiguous"
        note = "Meaning requires manual review before domain promotion."
    else:
        classification = "unsupported"
        note = "No authoritative v1 command field; value is preserved in raw payload."
    return FieldMapping(scope, base_field, classification, target, note)


def field_mapping_catalog() -> tuple[FieldMapping, ...]:
    mappings = [
        _mapping(sheet, field)
        for sheet, fields in LEGACY_XLSX_HEADERS.items()
        for field in fields
    ]
    mappings.extend(
        (
            FieldMapping(".caseflow", "id", "mapped", "ExternalReferenceInput.value", "Stable identity."),
            FieldMapping(".caseflow", "external_id", "mapped", "ExternalReferenceInput.value", "Stable identity."),
            FieldMapping(".caseflow", "sha256", "mapped", "ManagedFileRecord.sha256", "Verified hash."),
            FieldMapping(".caseflow", "*_id", "ambiguous", "relation pass", "Resolved only against this snapshot."),
            FieldMapping(".caseflow", "*", "unsupported", None, "Preserved as raw compatibility payload."),
        )
    )
    return tuple(mappings)


def _split_reference(value: Any) -> tuple[str, ...]:
    if isinstance(value, (list, tuple)):
        return tuple(str(item).strip() for item in value if str(item).strip())
    text = str(value).strip()
    if not text:
        return ()
    if ";" in text:
        return tuple(part.strip() for part in text.split(";") if part.strip())
    return (text,)


def _links(records: list[LegacyRecord]) -> tuple[list[LegacyLink], list[dict[str, Any]]]:
    aliases: dict[tuple[str, str], str] = {}
    global_aliases: dict[str, list[str]] = {}
    for record in records:
        scope = record.record_kind.split(":", 1)[1]
        for field in _PRIMARY_FIELDS.get(scope, ("id", "external_id")):
            value = record.payload.get(field)
            if value not in (None, ""):
                normalized = str(value).strip().casefold()
                aliases[(scope, normalized)] = record.external_ref
                global_aliases.setdefault(normalized, []).append(record.external_ref)

    links: list[LegacyLink] = []
    issues: list[dict[str, Any]] = []
    for record in records:
        scope = record.record_kind.split(":", 1)[1]
        for field, value in record.payload.items():
            base_field = field.split("#", 1)[0]
            target_scope = _LINK_FIELDS.get((scope, base_field))
            is_caseflow_link = scope not in LEGACY_XLSX_HEADERS and base_field in _CASEFLOW_LINK_KEYS
            if target_scope is None and not is_caseflow_link:
                continue
            for raw_target in _split_reference(value):
                normalized = raw_target.casefold()
                target = aliases.get((target_scope, normalized)) if target_scope else None
                if is_caseflow_link:
                    candidates = global_aliases.get(normalized, [])
                    target = candidates[0] if len(candidates) == 1 else None
                link = LegacyLink(record.external_ref, base_field, raw_target, target)
                links.append(link)
                if target is None:
                    issues.append(
                        {
                            "external_ref": record.external_ref,
                            "source_location": record.source_location,
                            "field": base_field,
                            "reason": "unresolved_link",
                            "target": raw_target,
                            "action": "manual_review_required",
                        }
                    )
    return links, issues


def _snapshot(path: Path, explicit_source_key: str | None = None) -> tuple[
    str, str, str, tuple[SourceTreeEntry, ...], list[LegacyRecord], list[dict[str, Any]]
]:
    path = path.resolve()
    entries = _tree(path)
    if path.is_file() and path.suffix.casefold() == ".xlsx":
        kind = "xlsx"
        version = LEGACY_XLSX_SCHEMA_VERSION
        records, issues = _read_xlsx(path)
    elif path.is_file() and path.suffix.casefold() == ".caseflow":
        kind = "caseflow"
        version = LEGACY_CASEFLOW_SCHEMA_VERSION
        records = _read_caseflow_json(path, label=path.name)
        issues = []
    elif path.is_dir() and path.name.casefold() == ".caseflow":
        kind = "caseflow"
        version = LEGACY_CASEFLOW_SCHEMA_VERSION
        records, issues = _read_caseflow_directory(path)
    else:
        raise ValueError("Підтримуються лише .xlsx і .caseflow")
    return kind, version, _source_key(path, kind, explicit_source_key), entries, records, issues


def _analyze(path: Path, source_key: str | None = None) -> LegacyImportReport:
    kind, version, resolved_key, source_files, records, issues = _snapshot(path, source_key)
    seen: set[str] = set()
    duplicate_locations: set[str] = set()
    for record in records:
        if record.external_ref in seen:
            duplicate_locations.add(record.source_location)
            issues.append(
                {
                    "external_ref": record.external_ref,
                    "source_location": record.source_location,
                    "reason": "duplicate_external_ref",
                    "action": "manual_review_required",
                }
            )
        seen.add(record.external_ref)

    links, link_issues = _links(records)
    issues.extend(link_issues)
    observed_mappings: dict[tuple[str, str], FieldMapping] = {}
    warning_locations: dict[str, list[str]] = {}
    for record in records:
        scope = record.record_kind.split(":", 1)[1]
        for field, value in record.payload.items():
            if value in (None, ""):
                continue
            mapping = _mapping(scope, field) if kind == "xlsx" else _caseflow_mapping(field)
            observed_mappings[(mapping.source_scope, mapping.source_field)] = mapping
            if mapping.classification in {"unsupported", "ambiguous", "lossy"}:
                warning_locations.setdefault(record.source_location, []).append(
                    f"{mapping.classification}:{mapping.source_field}"
                )

    issue_reasons: dict[str, list[str]] = {}
    for issue in issues:
        location = str(issue.get("source_location", ""))
        if location:
            issue_reasons.setdefault(location, []).append(str(issue["reason"]))
    results: list[RecordResult] = []
    for record in records:
        reasons = issue_reasons.get(record.source_location, []) + warning_locations.get(
            record.source_location, []
        )
        if record.source_location in duplicate_locations:
            action = "quarantine"
        elif reasons:
            action = "import_manual_review"
        else:
            action = "import"
        results.append(
            RecordResult(
                record.external_ref,
                record.source_location,
                record.record_kind,
                action,
                tuple(sorted(set(reasons))),
            )
        )

    source_sha = source_files[0].sha256 if len(source_files) == 1 else _tree_hash(source_files)
    counts = Counter(record.record_kind for record in records)
    return LegacyImportReport(
        source_kind=kind,
        source_name=path.name,
        source_key=resolved_key,
        format_version=version,
        source_sha256=source_sha,
        source_tree_sha256=_tree_hash(source_files),
        records=len(records),
        proposed=len(seen),
        imported=0,
        updated=0,
        skipped=0,
        conflicts=len(duplicate_locations),
        unresolved_links=len(link_issues),
        resolved_links=sum(link.target_external_ref is not None for link in links),
        counts=dict(sorted(counts.items())),
        issues=tuple(issues),
        record_results=tuple(results),
        field_mappings=tuple(
            observed_mappings[key] for key in sorted(observed_mappings, key=lambda item: str(item))
        ),
        source_files=source_files,
        zero_db_writes=True,
        source_unchanged=True,
    )


def _caseflow_mapping(field: str) -> FieldMapping:
    if field in {"id", "external_id"}:
        return FieldMapping(".caseflow", field, "mapped", "ExternalReferenceInput.value", "Stable identity.")
    if field.casefold() in {"sha256", "hash"}:
        return FieldMapping(".caseflow", field, "mapped", "ManagedFileRecord.sha256", "Hash provenance.")
    if field in _CASEFLOW_LINK_KEYS:
        return FieldMapping(".caseflow", field, "ambiguous", "relation pass", "Snapshot-local resolution.")
    if "path" in field.casefold() or "url" in field.casefold():
        return FieldMapping(".caseflow", field, "ambiguous", None, "Location is provenance, not identity.")
    return FieldMapping(".caseflow", field, "unsupported", None, "Raw compatibility payload only.")


def dry_run(path: Path, *, source_key: str | None = None) -> LegacyImportReport:
    """Read and reconcile a source without opening or writing a database."""

    return _analyze(path, source_key)


def verify_backup_destination(source: Path, destination: Path) -> Path:
    source = source.resolve()
    destination = destination.resolve()
    if not destination.is_dir() or destination.is_symlink() or not os.access(destination, os.W_OK):
        raise ValueError("Потрібен verified writable backup destination directory")
    if source.is_dir() and (destination == source or destination.is_relative_to(source)):
        raise ValueError("Backup destination має бути поза legacy source tree")
    if source.is_relative_to(destination):
        raise ValueError("Legacy source не може бути всередині backup destination")
    return destination


def _columns(connection: sqlite3.Connection, table: str) -> set[str]:
    return {str(row[1]) for row in connection.execute(f"PRAGMA table_info({table})")}


def _add_column(connection: sqlite3.Connection, table: str, definition: str) -> None:
    name = definition.split()[0]
    if name not in _columns(connection, table):
        connection.execute(f"ALTER TABLE {table} ADD COLUMN {definition}")


def _ensure_schema(connection: sqlite3.Connection) -> None:
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS legacy_import_runs (
            id TEXT PRIMARY KEY,
            source_kind TEXT NOT NULL,
            source_name TEXT NOT NULL,
            source_sha256 TEXT NOT NULL,
            mode TEXT NOT NULL,
            report_json TEXT NOT NULL CHECK (json_valid(report_json)),
            created_at TEXT NOT NULL,
            source_key TEXT,
            actor TEXT,
            backup_name TEXT
        )
        """
    )
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS legacy_import_records (
            id TEXT PRIMARY KEY,
            run_id TEXT NOT NULL,
            source_kind TEXT NOT NULL,
            source_sha256 TEXT NOT NULL,
            external_ref TEXT NOT NULL,
            record_kind TEXT NOT NULL,
            payload_json TEXT NOT NULL CHECK (json_valid(payload_json)),
            status TEXT NOT NULL,
            source_key TEXT,
            payload_sha256 TEXT,
            first_seen_at TEXT,
            last_seen_at TEXT
        )
        """
    )
    for definition in ("source_key TEXT", "actor TEXT", "backup_name TEXT"):
        _add_column(connection, "legacy_import_runs", definition)
    for definition in (
        "source_key TEXT", "payload_sha256 TEXT", "first_seen_at TEXT", "last_seen_at TEXT"
    ):
        _add_column(connection, "legacy_import_records", definition)
    connection.execute(
        """
        UPDATE legacy_import_records
        SET source_key = source_kind || ':' || source_sha256
        WHERE source_key IS NULL OR source_key = ''
        """
    )
    connection.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS ux_legacy_import_record_source_ref
        ON legacy_import_records(source_key, external_ref)
        """
    )
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS legacy_import_links (
            id TEXT PRIMARY KEY,
            run_id TEXT NOT NULL,
            source_key TEXT NOT NULL,
            from_external_ref TEXT NOT NULL,
            source_field TEXT NOT NULL,
            raw_target TEXT NOT NULL,
            target_external_ref TEXT,
            status TEXT NOT NULL,
            UNIQUE(source_key, from_external_ref, source_field, raw_target)
        )
        """
    )
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS legacy_import_issues (
            id TEXT PRIMARY KEY,
            run_id TEXT NOT NULL,
            issue_key TEXT NOT NULL,
            issue_json TEXT NOT NULL CHECK (json_valid(issue_json)),
            UNIQUE(run_id, issue_key)
        )
        """
    )


def _record_status(report: LegacyImportReport, external_ref: str) -> str:
    actions = {
        item.action for item in report.record_results if item.external_ref == external_ref
    }
    return "manual_review_required" if actions & {"quarantine", "import_manual_review"} else "imported"


def _import_snapshot(
    connection: sqlite3.Connection,
    path: Path,
    *,
    backup_destination: Path,
    actor: str,
    source_key: str | None,
) -> LegacyImportReport:
    if not actor.strip():
        raise ValueError("actor має бути непорожнім")
    backup = verify_backup_destination(path, backup_destination)
    plan = _analyze(path, source_key)
    before_tree = plan.source_files
    _, _, _, _, records, _ = _snapshot(path, source_key)
    links, _ = _links(records)
    run_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    imported = updated = skipped = 0
    actual_actions: dict[str, str] = {}
    savepoint = f"legacy_import_{uuid.uuid4().hex}"
    connection.execute(f"SAVEPOINT {savepoint}")
    try:
        _ensure_schema(connection)
        connection.execute(
            """
            INSERT INTO legacy_import_runs(
                id, source_kind, source_name, source_sha256, mode, report_json,
                created_at, source_key, actor, backup_name
            ) VALUES (?, ?, ?, ?, 'import', ?, ?, ?, ?, ?)
            """,
            (
                run_id,
                plan.source_kind,
                plan.source_name,
                plan.source_sha256,
                _canonical_json(plan.to_dict()),
                now,
                plan.source_key,
                actor.strip(),
                backup.name,
            ),
        )

        unique_records: dict[str, LegacyRecord] = {}
        for record in records:
            if record.external_ref in unique_records:
                skipped += 1
                continue
            unique_records[record.external_ref] = record
        for external_ref, record in unique_records.items():
            payload_json = _canonical_json(record.payload)
            payload_hash = _sha256_bytes(payload_json.encode("utf-8"))
            existing = connection.execute(
                """
                SELECT id, payload_sha256, payload_json
                FROM legacy_import_records
                WHERE source_key = ? AND external_ref = ?
                """,
                (plan.source_key, external_ref),
            ).fetchone()
            status = _record_status(plan, external_ref)
            if existing is None:
                record_id = str(uuid.uuid5(uuid.NAMESPACE_URL, f"{plan.source_key}:{external_ref}"))
                connection.execute(
                    """
                    INSERT INTO legacy_import_records(
                        id, run_id, source_kind, source_sha256, external_ref, record_kind,
                        payload_json, status, source_key, payload_sha256, first_seen_at, last_seen_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        record_id, run_id, plan.source_kind, plan.source_sha256,
                        external_ref, record.record_kind, payload_json, status,
                        plan.source_key, payload_hash, now, now,
                    ),
                )
                imported += 1
                actual_actions[external_ref] = "imported"
            else:
                old_hash = existing[1] or _sha256_bytes(str(existing[2]).encode("utf-8"))
                if old_hash == payload_hash:
                    skipped += 1
                    actual_actions[external_ref] = "skipped_unchanged"
                else:
                    connection.execute(
                        """
                        UPDATE legacy_import_records
                        SET run_id = ?, source_sha256 = ?, record_kind = ?, payload_json = ?,
                            status = ?, payload_sha256 = ?, last_seen_at = ?
                        WHERE id = ?
                        """,
                        (
                            run_id, plan.source_sha256, record.record_kind, payload_json,
                            status, payload_hash, now, existing[0],
                        ),
                    )
                    updated += 1
                    actual_actions[external_ref] = "updated"

        for link in links:
            status = "resolved" if link.target_external_ref is not None else "manual_review_required"
            link_id = str(
                uuid.uuid5(
                    uuid.NAMESPACE_URL,
                    f"{plan.source_key}:{link.from_external_ref}:{link.source_field}:{link.raw_target}",
                )
            )
            connection.execute(
                """
                INSERT INTO legacy_import_links(
                    id, run_id, source_key, from_external_ref, source_field,
                    raw_target, target_external_ref, status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(source_key, from_external_ref, source_field, raw_target)
                DO UPDATE SET run_id=excluded.run_id,
                    target_external_ref=excluded.target_external_ref, status=excluded.status
                """,
                (
                    link_id, run_id, plan.source_key, link.from_external_ref,
                    link.source_field, link.raw_target, link.target_external_ref, status,
                ),
            )

        for ordinal, issue in enumerate(plan.issues):
            issue_json = _canonical_json(issue)
            issue_key = _sha256_bytes(f"{ordinal}:{issue_json}".encode("utf-8"))
            connection.execute(
                "INSERT INTO legacy_import_issues VALUES (?, ?, ?, ?)",
                (str(uuid.uuid4()), run_id, issue_key, issue_json),
            )

        actual_results: list[RecordResult] = []
        seen_results: set[str] = set()
        for result in plan.record_results:
            if result.external_ref in seen_results:
                action = "quarantined_duplicate"
            else:
                action = actual_actions.get(result.external_ref, result.action)
                seen_results.add(result.external_ref)
            actual_results.append(replace(result, action=action))
        final = replace(
            plan,
            imported=imported,
            updated=updated,
            skipped=skipped,
            record_results=tuple(actual_results),
            zero_db_writes=False,
        )
        connection.execute(
            "UPDATE legacy_import_runs SET report_json = ? WHERE id = ?",
            (_canonical_json(final.to_dict()), run_id),
        )
        after_tree = _tree(path.resolve())
        if before_tree != after_tree:
            raise RuntimeError("Legacy source змінився під час import; transaction rolled back")
        connection.execute(f"RELEASE SAVEPOINT {savepoint}")
        return final
    except Exception:
        connection.execute(f"ROLLBACK TO SAVEPOINT {savepoint}")
        connection.execute(f"RELEASE SAVEPOINT {savepoint}")
        raise


@dataclass(slots=True)
class SQLiteLegacyMigrationAdapter:
    connection: sqlite3.Connection

    def dry_run(self, command: LegacyDryRunCommand) -> LegacyImportReport:
        return dry_run(command.source, source_key=command.source_key)

    def import_snapshot(self, command: LegacyImportCommand) -> LegacyImportReport:
        return _import_snapshot(
            self.connection,
            command.source,
            backup_destination=command.backup_destination,
            actor=command.actor,
            source_key=command.source_key,
        )


def import_legacy(
    connection: sqlite3.Connection,
    path: Path,
    *,
    backup_destination: Path,
    actor: str = "system:legacy-import",
    source_key: str | None = None,
) -> LegacyImportReport:
    """Compatibility facade over the application command/adapter boundary."""

    adapter = SQLiteLegacyMigrationAdapter(connection)
    return adapter.import_snapshot(
        LegacyImportCommand(path, backup_destination, actor=actor, source_key=source_key)
    )
